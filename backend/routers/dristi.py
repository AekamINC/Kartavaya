"""
dristi.py — Dristi · दृष्टि (Analytics) Router
Cross-module KPIs, trends, and saved dashboards.
Reads from all modules: Graha, Ganit, Manav, Vikray, Vetana, tasks.
"""
import asyncio
import json
import logging
import os
from datetime import date, datetime, time as _dt_time, timedelta, timezone


def _parse_time(s: str) -> _dt_time:
    parts = s.split(":")
    return _dt_time(int(parts[0]), int(parts[1]))

from fastapi import APIRouter, Depends, Header, HTTPException, Response
from pydantic import BaseModel

from auth_router import require_user
from db import get_pool
from middleware.org_resolver import get_org_id
from middleware.subscription import require_module
from middleware.module_levels import held_level
from services.audit_actors import display_name
from services.on_the_rolls import DEFAULT_ALIAS, still_on_the_rolls
from services.report_schedule_window import blocked_reason, is_due
from services import analytics_window as aw

router = APIRouter(prefix="/api/v1/dristi", tags=["dristi-analytics"])
log = logging.getLogger(__name__)

_gate = require_module("dristi")


# ── Cross-module reach ───────────────────────────────────────────────────────
#
# Dristi reads from every other module by design — that is what an analytics
# module is. The problem is that it was gated on `dristi` ALONE, so the
# entitlement that governs the SOURCE data was never consulted.
#
# Concretely, before this: `GET /overview` returned payroll totals from
# `vetana_payroll_runs`, headcount from `manav_employees` and revenue from
# `ganit_invoices` to anyone holding a `dristi` grant. All three of those are
# SENSITIVE modules — withheld by default, audited on platform bypass. And
# `dristi` is in `STAFF_MODULES` while ganit, manav and vetana are deliberately
# not, so `platform_staff` — a role whose whole definition is "the operating
# set, excluding finance and all HR" — could read any customer's payroll and
# revenue through the analytics endpoint, with no audit row.
#
# The fix is per-source, not per-endpoint. Refusing the whole dashboard would
# break it for every legitimate user; instead each block is included only if the
# caller can reach the module it comes from, and the response says which were
# withheld so the UI can render an honest gap rather than a silent zero.

_OVERVIEW_SOURCES = {
    "crm": "graha",
    "deals": "graha",
    "revenue": "ganit",
    "hr": "manav",
    "orders": "vikray",
    "payroll": "vetana",
}


async def reachable_modules(pool, user_id: str, org_id: str, codes) -> set[str]:
    """Which of `codes` this caller may actually read, via the Tier-4 resolver.

    `held_level` returns None when the caller has no platform reach, no org
    admin role and no grant row — which is exactly "may not read this module".
    """
    reachable = set()
    for code in codes:
        if await held_level(pool, user_id, org_id, code) is not None:
            reachable.add(code)
    return reachable


# ── Pydantic Models ──────────────────────────────────────────

class DashboardCreate(BaseModel):
    name: str
    description: str = ""
    widgets: list[dict] = []
    is_default: bool = False


class DashboardUpdate(BaseModel):
    name: str | None = None
    description: str | None = None
    widgets: list[dict] | None = None
    is_default: bool | None = None


# ── Helpers ──────────────────────────────────────────────────

def _month_range(months_back: int = 6):
    today = date.today()
    ranges = []
    for i in range(months_back - 1, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        label = f"{y}-{m:02d}"
        ranges.append(label)
    return ranges


#: Which module each exportable report reads. `_fetch_report_data` is reached
#: by `GET /exports/{type}` and by scheduled reports, and both bypassed every
#: source-module check the GET endpoints have — exporting "hr" returned the
#: employee register, and "revenue" the invoice ledger, behind `dristi` alone.
#: The map itself now lives in services/module_report.py, ONE copy for all
#: three delivery doors (this router, the sweep, and Niyam's report.send) —
#: aliased here so every existing reference keeps reading.
from services.module_report import REPORT_SOURCE_MODULES as _REPORT_SOURCE_MODULES  # noqa: E402


async def _fetch_report_data(pool, org_id: str, report_type: str,
                             win: "aw.Window | None" = None) -> dict:
    """Fetch report data by type, reusing the same queries as the GET endpoints.

    `win` is the D1 retrofit. Where a report has a date column the window
    filters it; where it counts current state (headcount, contacts, tasks on
    hand) the count is left alone, because a stock has no period. A scheduled
    report passes None and therefore behaves exactly as it did before.
    """
    today = date.today()
    month_ago = today - timedelta(days=30)
    # Bound as ($2, $3) by every branch that uses them, so the SQL below reads
    # the same whether or not a window was supplied.
    wargs = [win.start, win.end] if win else []

    if report_type == "overview":
        # `teams` has no `id` column — its primary key is `team_id` (TEXT), and
        # the forward tenancy path is `teams.org_id` (migration 028). The old
        # query joined `tm.id` twice, so it raised UndefinedColumn and this
        # branch 500'd every time: `GET /exports/overview` had never returned,
        # and neither had a scheduled report of type "overview".
        tasks = await pool.fetchval(
            "SELECT COUNT(*) FROM tasks t "
            "JOIN teams tm ON tm.team_id = t.team_id "
            "WHERE tm.org_id=$1::uuid AND tm.deleted_at IS NULL",
            org_id,
        )
        contacts = await pool.fetchval(
            "SELECT COUNT(*) FROM staging.graha_contacts WHERE org_id=$1::uuid AND is_active=TRUE", org_id)
        # The SAME guards as the `/overview` tile this block exports, for the
        # same reason the "revenue" branch below carries them: an export that
        # disagrees with the screen it was taken from is worse than either
        # being wrong alone.
        #
        # `payment_status='paid'` narrowed the leak but did not close it — a
        # draft CAN be marked paid. Live 2026-08-26 for Unicode Group there is
        # exactly one, Rs 2,06,500, and this figure read Rs 27,39,830 against a
        # true Rs 25,33,330. E2E has no paid draft at all, so probing the
        # in-scope orgs one at a time would have certified this as clean.
        revenue = await pool.fetchval(
            "SELECT COALESCE(SUM(total),0) FROM staging.ganit_invoices "
            "WHERE org_id=$1::uuid AND payment_status='paid' AND is_active=TRUE "
            "AND COALESCE(doc_status, '') <> 'draft'"
            + (" AND invoice_date BETWEEN $2::date AND $3::date" if win else ""),
            org_id, *wargs) or 0
        return {"tasks": tasks, "contacts": contacts, "revenue": float(revenue)}
    elif report_type == "revenue":
        # The SAME guards as `revenue_trends` below, which is the on-screen
        # tile this block exports to CSV and PDF. They are two renderings of
        # one number: an export that disagrees with the chart it was taken
        # from is worse than either being wrong alone, because the reader has
        # no way to tell which one to believe.
        rows = await pool.fetch(
            "SELECT DATE_TRUNC('month', invoice_date) AS month, "
            "SUM(total) AS total, COUNT(*) AS count "
            "FROM staging.ganit_invoices WHERE org_id=$1::uuid AND is_active=TRUE "
            "AND COALESCE(doc_status, '') <> 'draft' "
            + ("AND invoice_date BETWEEN $2::date AND $3::date " if win else "")
            + "GROUP BY 1 ORDER BY 1 DESC LIMIT 12", org_id, *wargs)
        return {"monthly": [dict(r) for r in rows]}
    elif report_type == "pipeline":
        rows = await pool.fetch(
            "SELECT stage, COUNT(*) AS count, SUM(value) AS value "
            "FROM staging.graha_deals WHERE org_id=$1::uuid AND is_active=TRUE "
            "GROUP BY stage", org_id)
        return {"stages": [dict(r) for r in rows]}
    elif report_type == "hr":
        # HEADCOUNT IS A STOCK, and this is the CSV and PDF twin of the
        # `/overview` tile — the same question, in a file a partner mails to a
        # client. `is_active` alone does not answer it: the flag is one
        # somebody has to remember to clear, and a leaver deliberately KEEPS it
        # until settlement (`routers/manav.py:1958` — clearing it on the last
        # working day dropped the person out of payroll and stranded an
        # unrecovered salary advance). The fact is
        # `manav_offboarding.last_working_day`.
        #
        # The predicate comes from `services/on_the_rolls.py` rather than being
        # written out here, so this and the tile cannot answer differently.
        # Live 2026-08-26: E2E Test & Associates 83 -> 73; Unicode Group
        # records no exits and stays at 26.
        count = await pool.fetchval(
            "SELECT COUNT(*) FROM staging.manav_employees e "
            "WHERE e.org_id=$1::uuid AND e.is_active=TRUE AND e.status='active'"
            + still_on_the_rolls("e"), org_id)
        return {"active_employees": count}
    elif report_type == "sales":
        rows = await pool.fetch(
            "SELECT status, COUNT(*) AS count, SUM(total) AS total "
            "FROM staging.vikray_orders WHERE org_id=$1::uuid "
            + ("AND order_date BETWEEN $2::date AND $3::date " if win else "")
            + "GROUP BY status", org_id, *wargs)
        return {"orders_by_status": [dict(r) for r in rows]}
    return {"report_type": report_type}


# ── Overview KPIs ────────────────────────────────────────────

@router.get("/overview")
async def overview(
    date_from: str = "",
    date_to: str = "",
    user=Depends(require_user),
    org_id: str = Depends(get_org_id),
    _g=Depends(_gate),
):
    pool = await get_pool()
    # Flows take the window; stocks do not. `tasks`, `crm`, `deals` and `hr`
    # are counts of what exists now — windowing them by `created_at` would
    # answer a different question under the same label. See analytics_window.
    win = aw.parse(date_from, date_to)
    allowed = await reachable_modules(
        pool, user["user_id"], org_id, set(_OVERVIEW_SOURCES.values()),
    )

    tasks_stats = await pool.fetchrow(
        "SELECT COUNT(*) AS total_tasks, "
        "COUNT(*) FILTER (WHERE status='done') AS done_tasks, "
        "COUNT(*) FILTER (WHERE status='in_progress') AS active_tasks, "
        "COUNT(*) FILTER (WHERE due_at < NOW() AND status != 'done') AS overdue_tasks "
        "FROM tasks WHERE team_id IN ("
        "  SELECT team_id FROM teams WHERE org_id=$1::uuid AND deleted_at IS NULL"
        ") AND archived_at IS NULL",
        org_id,
    )

    crm = None
    if "graha" in allowed:
        crm = await pool.fetchrow(
        "SELECT COUNT(*) AS total_contacts, "
        "COUNT(*) FILTER (WHERE contact_type='lead') AS leads, "
        "COUNT(*) FILTER (WHERE contact_type='customer') AS customers "
        "FROM staging.graha_contacts WHERE org_id=$1::uuid AND is_active=TRUE",
        org_id,
    )

    deals = None
    if "graha" in allowed:
        deals = await pool.fetchrow(
        "SELECT COUNT(*) AS total_deals, "
        "COALESCE(SUM(value),0) AS pipeline_value, "
        "COUNT(*) FILTER (WHERE stage='Won') AS won_deals, "
        "COALESCE(SUM(value) FILTER (WHERE stage='Won'),0) AS won_value, "
        "COUNT(*) FILTER (WHERE stage='Lost') AS lost_deals "
        "FROM staging.graha_deals WHERE org_id=$1::uuid",
        org_id,
    )

    revenue = None
    if "ganit" in allowed:
        # This tile is printed DIRECTLY ABOVE `revenue_trends`, and until now it
        # was the only invoice read on the page with no guard at all —
        # `org_id=$1::uuid` and nothing else. So the chart excluded drafts and
        # the number above it did not, and the two disagreed on screen.
        # Measured live 2026-08-26 for E2E Test & Associates: total_invoiced
        # Rs 12,29,86,008.58 against Rs 11,14,93,756.12, a Rs 1,14,92,252.46
        # phantom over 97 draft rows; outstanding Rs 3,86,36,429.46 against
        # Rs 2,71,54,767.00. That org holds zero inactive invoices, so the whole
        # gap is drafts.
        #
        # `outstanding` matters more than the headline: an unissued document is
        # not a receivable, and this figure is what a partner chases a client
        # over. `total_collected` moves too — Rs 10,590 sat as `amount_paid` on
        # three partially-paid drafts.
        #
        # The guards are `revenue_trends`' verbatim, which are in turn
        # `analytics.py`'s client report verbatim. `COALESCE(doc_status, '')`
        # rather than a bare `<>`, the canonical form from
        # `services/gst_period.py`: the column is nullable and NULL <> 'draft'
        # is NULL, which would silently drop every row predating the column.
        revenue = await pool.fetchrow(
        "SELECT COALESCE(SUM(total),0) AS total_invoiced, "
        "COALESCE(SUM(amount_paid),0) AS total_collected, "
        "COALESCE(SUM(total - amount_paid) FILTER (WHERE payment_status NOT IN ('paid','cancelled')),0) AS outstanding "
        "FROM staging.ganit_invoices WHERE org_id=$1::uuid AND is_active=TRUE "
        "AND COALESCE(doc_status, '') <> 'draft'"
        + (" AND invoice_date BETWEEN $2::date AND $3::date" if win else ""),
        *([org_id, win.start, win.end] if win else [org_id]),
    )

    hr = None
    if "manav" in allowed:
        hr = await pool.fetchrow(
        # HEADCOUNT IS A STOCK — this function's own docstring says so — and a
        # stock is "who is on the rolls NOW", which `is_active` alone does not
        # answer. `is_active` is a flag somebody has to remember to clear;
        # `manav_offboarding.last_working_day` is a fact already recorded, and
        # the two disagree by exactly the people who have gone.
        #
        # AND THE FLAG IS NOT STALE — IT IS DELIBERATE. `routers/manav.py:1958`
        # records that offboarding used to set `is_active=FALSE`, which dropped
        # the person out of payroll the same day and left an outstanding salary
        # advance unrecoverable. So a leaver KEEPS the flag until settlement.
        # Live 2026-08-26, E2E: two of the ten still carry advances totalling
        # ₹1,15,000. Nothing about that data may be "cleaned"; the reads are
        # what must ask the right question.
        #
        # The predicate is `analytics/metrics/manav.py::_headcount_asat`'s, at
        # today rather than at a bound date — one definition of "on the rolls",
        # not a second one that drifts. Live: 83 before, 73 after.
        "SELECT COUNT(*) AS headcount, "
        "COUNT(*) FILTER (WHERE department IS NOT NULL AND department != '') AS in_departments "
        "FROM staging.manav_employees e WHERE e.org_id=$1::uuid AND e.is_active=TRUE "
        "AND NOT EXISTS ("
        "  SELECT 1 FROM staging.manav_offboarding x "
        "   WHERE x.org_id = e.org_id AND x.employee_id = e.id "
        "     AND x.status <> 'cancelled' "
        "     AND x.last_working_day < CURRENT_DATE)",
        org_id,
    )

    orders = None
    if "vikray" in allowed:
        orders = await pool.fetchrow(
        "SELECT COUNT(*) AS total_orders, "
        "COALESCE(SUM(total),0) AS order_value, "
        "COUNT(*) FILTER (WHERE status='delivered' OR status='closed') AS fulfilled "
        "FROM staging.vikray_orders WHERE org_id=$1::uuid AND is_active=TRUE"
        + (" AND order_date BETWEEN $2::date AND $3::date" if win else ""),
        *([org_id, win.start, win.end] if win else [org_id]),
    )

    payroll = None
    if "vetana" in allowed:
        payroll = await pool.fetchrow(
        "SELECT COALESCE(SUM(total_net),0) AS ytd_payroll, "
        "COALESCE(SUM(total_pf + total_esi + total_tds),0) AS ytd_statutory "
        # `month` is TEXT 'YYYY-MM', so a lexicographic comparison is also a
        # chronological one and the window needs no cast. The response keys stay
        # `ytd_*` because the frontend reads them by name; the `window` block
        # below is what tells the reader the span is no longer the year.
        "FROM staging.vetana_payroll_runs "
        "WHERE org_id=$1::uuid AND month "
        + ("BETWEEN $2 AND $3" if win else "LIKE $2"),
        *([org_id, win.start.strftime("%Y-%m"), win.end.strftime("%Y-%m")] if win
          else [org_id, f"{date.today().year}-%"]),
    )

    # `withheld` is named rather than left as an empty object so the UI can say
    # "you do not have access to payroll" instead of drawing a zero, which is
    # indistinguishable from a company that paid nobody this year.
    return {
        "tasks": dict(tasks_stats) if tasks_stats else {},
        "crm": dict(crm) if crm else {},
        "deals": dict(deals) if deals else {},
        "revenue": dict(revenue) if revenue else {},
        "hr": dict(hr) if hr else {},
        "orders": dict(orders) if orders else {},
        "payroll": dict(payroll) if payroll else {},
        "withheld": sorted({
            block for block, mod in _OVERVIEW_SOURCES.items() if mod not in allowed
        }),
        "window": aw.describe(
            win,
            windowed=["revenue", "orders", "payroll"],
            as_at=["tasks", "crm", "deals", "hr"],
        ),
    }


# ── Revenue Trends ───────────────────────────────────────────

@router.get("/revenue")
async def revenue_trends(
    months: int = 6,
    date_from: str = "",
    date_to: str = "",
    user=Depends(require_user),
    org_id: str = Depends(get_org_id),
    _g=Depends(_gate),
):
    pool = await get_pool()

    # Every figure here is invoices and expenses — the ledger, not "analytics".
    # Ganit is a sensitive module and a `dristi` grant is not a grant to it.
    if not await reachable_modules(pool, user["user_id"], org_id, {"ganit"}):
        raise HTTPException(
            403,
            "Revenue analytics reads the accounting ledger. Ask your org admin "
            "for access to the Ganit module.",
        )

    # An explicit window supersedes `months` entirely — the two cannot both be
    # authoritative, and a caller that sends dates means the dates. `months`
    # stays for the callers that already use it, including saved widgets.
    win = aw.parse(date_from, date_to)
    labels = aw.months_between(win) if win else _month_range(min(months, 12))

    # The revenue half of this tile counted DRAFTS as revenue and the expense
    # half four lines below did not — two figures on one chart, filtered
    # differently, subtracted from each other to make `profit`. Measured live
    # 2026-08-25: 102 draft invoices in the table, 87 of them inside this
    # query's default one-year window, Rs 85,16,666.56 of revenue that nobody
    # had been billed for.
    #
    # The guards are `analytics.py`'s client report VERBATIM — `is_active` plus
    # the draft exclusion — because that report is the same money narrowed to
    # one client, and a dashboard that disagrees with the client page about the
    # same rows discredits both. `is_active` also matches the expense query's
    # own shape; live there are zero soft-deleted invoices, so it is a no-op
    # today and a guard for ever.
    #
    # `COALESCE(doc_status, '')` rather than a bare `<>`, the canonical form
    # from `services/gst_period.py`: the column is nullable and NULL <> 'draft'
    # is NULL, which would silently drop every row predating the column.
    invoiced = await pool.fetch(
        "SELECT TO_CHAR(invoice_date, 'YYYY-MM') AS month, "
        "SUM(total) AS invoiced, SUM(amount_paid) AS collected, COUNT(*) AS count "
        "FROM staging.ganit_invoices "
        "WHERE org_id=$1::uuid AND is_active=TRUE "
        "AND COALESCE(doc_status, '') <> 'draft' AND "
        + ("invoice_date BETWEEN $2::date AND $3::date "
           if win else "invoice_date >= (CURRENT_DATE - INTERVAL '1 year') ")
        + "GROUP BY month ORDER BY month",
        *([org_id, win.start, win.end] if win else [org_id]),
    )
    inv_map = {r["month"]: dict(r) for r in invoiced}

    expenses = await pool.fetch(
        "SELECT TO_CHAR(expense_date, 'YYYY-MM') AS month, "
        "SUM(total) AS total_expenses, COUNT(*) AS count "
        "FROM staging.ganit_expenses "
        "WHERE org_id=$1::uuid AND is_active=TRUE AND "
        + ("expense_date BETWEEN $2::date AND $3::date "
           if win else "expense_date >= (CURRENT_DATE - INTERVAL '1 year') ")
        + "GROUP BY month ORDER BY month",
        *([org_id, win.start, win.end] if win else [org_id]),
    )
    exp_map = {r["month"]: dict(r) for r in expenses}

    trend = []
    for m in labels:
        inv = inv_map.get(m, {})
        exp = exp_map.get(m, {})
        trend.append({
            "month": m,
            "invoiced": float(inv.get("invoiced", 0)),
            "collected": float(inv.get("collected", 0)),
            "expenses": float(exp.get("total_expenses", 0)),
            "invoice_count": inv.get("count", 0),
            "profit": float(inv.get("collected", 0)) - float(exp.get("total_expenses", 0)),
        })

    return {
        "trend": trend,
        "labels": labels,
        "window": aw.describe(win, windowed=["trend"], as_at=[]),
    }


# ── Pipeline Analytics ───────────────────────────────────────

@router.get("/pipeline")
async def pipeline_analytics(
    date_from: str = "",
    date_to: str = "",
    user=Depends(require_user),
    org_id: str = Depends(get_org_id),
    _g=Depends(_gate),
):
    pool = await get_pool()
    # `stages` and `conversion` are the pipeline as it stands, which is a stock:
    # a deal sitting in Negotiation is in Negotiation today regardless of the
    # dates asked for. Only the blocks that describe deals RESOLVING — the won
    # trend and the customers behind it — have a period, and `updated_at` is
    # the closest thing to a closed-at date the table has.
    win = aw.parse(date_from, date_to)

    # Every block below is the CRM: deal values, win rates and named customers.
    # Same rule as /revenue and /hr — `dristi` is in STAFF_MODULES, `graha` is
    # reachable by fewer roles than that, and this endpoint was the one place
    # the pipeline could be read without holding it.
    if not await reachable_modules(pool, user["user_id"], org_id, {"graha"}):
        raise HTTPException(
            403,
            "Pipeline analytics reads CRM deals. Ask your org admin for access "
            "to the Graha module.",
        )

    stages = await pool.fetch(
        "SELECT stage, COUNT(*) AS count, COALESCE(SUM(value),0) AS value "
        "FROM staging.graha_deals WHERE org_id=$1::uuid "
        "GROUP BY stage ORDER BY CASE stage "
        "WHEN 'New' THEN 1 WHEN 'Qualified' THEN 2 WHEN 'Proposal' THEN 3 "
        "WHEN 'Negotiation' THEN 4 WHEN 'Won' THEN 5 WHEN 'Lost' THEN 6 ELSE 7 END",
        org_id,
    )

    won_trend = await pool.fetch(
        "SELECT TO_CHAR(updated_at, 'YYYY-MM') AS month, "
        "COUNT(*) AS deals_won, COALESCE(SUM(value),0) AS value_won "
        "FROM staging.graha_deals "
        "WHERE org_id=$1::uuid AND stage='Won' AND "
        + ("updated_at::date BETWEEN $2::date AND $3::date "
           if win else "updated_at >= (CURRENT_DATE - INTERVAL '6 months') ")
        + "GROUP BY month ORDER BY month",
        *([org_id, win.start, win.end] if win else [org_id]),
    )

    conversion = await pool.fetchrow(
        "SELECT "
        "COUNT(*) AS total, "
        "COUNT(*) FILTER (WHERE stage='Won') AS won, "
        "COUNT(*) FILTER (WHERE stage='Lost') AS lost, "
        "CASE WHEN COUNT(*) > 0 THEN "
        "  ROUND(COUNT(*) FILTER (WHERE stage='Won')::numeric / COUNT(*) * 100, 1) "
        "ELSE 0 END AS win_rate "
        "FROM staging.graha_deals WHERE org_id=$1::uuid",
        org_id,
    )

    top_contacts = await pool.fetch(
        "SELECT c.name, c.company, COUNT(d.id) AS deal_count, COALESCE(SUM(d.value),0) AS total_value "
        "FROM staging.graha_deals d "
        "JOIN staging.graha_contacts c ON c.id = d.contact_id "
        "WHERE d.org_id=$1::uuid AND d.stage='Won' "
        + ("AND d.updated_at::date BETWEEN $2::date AND $3::date " if win else "")
        + "GROUP BY c.id, c.name, c.company "
        "ORDER BY total_value DESC LIMIT 10",
        *([org_id, win.start, win.end] if win else [org_id]),
    )

    return {
        "stages": [dict(r) for r in stages],
        "won_trend": [dict(r) for r in won_trend],
        "conversion": dict(conversion) if conversion else {},
        "top_contacts": [dict(r) for r in top_contacts],
        "window": aw.describe(
            win, windowed=["won_trend", "top_contacts"], as_at=["stages", "conversion"]),
    }


# ── HR Analytics ─────────────────────────────────────────────

@router.get("/hr")
async def hr_analytics(
    date_from: str = "",
    date_to: str = "",
    user=Depends(require_user),
    org_id: str = Depends(get_org_id),
    _g=Depends(_gate),
):
    pool = await get_pool()
    win = aw.parse(date_from, date_to)

    # Every block below is HR or payroll. Unlike /overview there is no
    # non-sensitive remainder worth returning, so this refuses outright rather
    # than serving an empty shell — a `dristi` grant is not a grant to the
    # salary register.
    allowed = await reachable_modules(pool, user["user_id"], org_id, {"manav", "vetana"})
    if "manav" not in allowed:
        raise HTTPException(
            403,
            "HR analytics reads employee records. Ask your org admin for access "
            "to the Manav module.",
        )

    dept_breakdown = await pool.fetch(
        # "How many people are in Accounts" is a STOCK — who is there NOW — so
        # it carries the same guard as the headcount tile on `/overview`, from
        # the same module, and cannot drift away from it. Ten of E2E's people
        # left up to seven weeks ago and were still being counted into the
        # departments they left: live 2026-08-26, Accounts 8->6, Payroll 8->6,
        # Taxation 8->7, Compliance 8->7, and Administration, Advisory, Audit
        # and IT 7->6 each. Unicode Group records no exits and does not move.
        "SELECT COALESCE(NULLIF(e.department,''), 'Unassigned') AS department, COUNT(*) AS count "
        "FROM staging.manav_employees e "
        "WHERE e.org_id=$1::uuid AND e.is_active=TRUE"
        + still_on_the_rolls("e")
        + " GROUP BY e.department ORDER BY count DESC",
        org_id,
    )

    # Payroll is a separate entitlement from HR — reaching Manav does not mean
    # reaching what people are paid.
    payroll_trend = []
    if "vetana" in allowed:
        payroll_trend = await pool.fetch(
        "SELECT month, total_gross, total_net, total_pf, total_esi, total_tds, employee_count "
        "FROM staging.vetana_payroll_runs "
        "WHERE org_id=$1::uuid "
        + ("AND month BETWEEN $2 AND $3 " if win else "")
        + "ORDER BY month DESC LIMIT 12",
        *([org_id, win.start.strftime("%Y-%m"), win.end.strftime("%Y-%m")] if win else [org_id]),
    )

    leave_stats = await pool.fetchrow(
        "SELECT COUNT(*) AS total_leaves, "
        "COUNT(*) FILTER (WHERE status='approved') AS approved, "
        "COUNT(*) FILTER (WHERE status='pending') AS pending, "
        "COUNT(*) FILTER (WHERE status='rejected') AS rejected "
        "FROM staging.manav_leave_requests WHERE org_id=$1::uuid AND "
        + ("start_date BETWEEN $2::date AND $3::date"
           if win else "start_date >= DATE_TRUNC('year', CURRENT_DATE)"),
        *([org_id, win.start, win.end] if win else [org_id]),
    )

    attendance = await pool.fetchrow(
        "SELECT COUNT(DISTINCT employee_id) AS tracked, "
        "COUNT(*) FILTER (WHERE status='present') AS present_days, "
        "COUNT(*) FILTER (WHERE status='absent') AS absent_days "
        "FROM staging.manav_attendance "
        "WHERE org_id=$1::uuid AND "
        + ("date BETWEEN $2::date AND $3::date"
           if win else "date >= (CURRENT_DATE - INTERVAL '30 days')"),
        *([org_id, win.start, win.end] if win else [org_id]),
    )

    return {
        "departments": [dict(r) for r in dept_breakdown],
        "payroll_trend": [dict(r) for r in payroll_trend],
        "leave_stats": dict(leave_stats) if leave_stats else {},
        # The key keeps its name because the frontend reads it by name; when a
        # window is supplied it covers that window, and `window.windowed` says so.
        "attendance_30d": dict(attendance) if attendance else {},
        "window": aw.describe(
            win,
            windowed=["payroll_trend", "leave_stats", "attendance_30d"],
            as_at=["departments"],
        ),
    }


# ── Sales Analytics ──────────────────────────────────────────

@router.get("/sales")
async def sales_analytics(
    date_from: str = "",
    date_to: str = "",
    user=Depends(require_user),
    org_id: str = Depends(get_org_id),
    _g=Depends(_gate),
):
    pool = await get_pool()
    # The leaderboard is deliberately left alone: it is bound to each target's
    # own period_start/period_end, and a second window over the top of that
    # would produce attainment against a target nobody set for those dates.
    win = aw.parse(date_from, date_to)

    # Orders and targets are Vikray; the leaderboard additionally reads won
    # deals out of Graha. Refuse without Vikray — there is no non-sensitive
    # remainder — and drop the leaderboard alone when Graha is unreachable,
    # the same split /hr uses for payroll.
    allowed = await reachable_modules(pool, user["user_id"], org_id, {"vikray", "graha"})
    if "vikray" not in allowed:
        raise HTTPException(
            403,
            "Sales analytics reads the order book. Ask your org admin for "
            "access to the Vikray module.",
        )

    order_trend = await pool.fetch(
        "SELECT TO_CHAR(order_date, 'YYYY-MM') AS month, "
        "COUNT(*) AS orders, COALESCE(SUM(total),0) AS value "
        "FROM staging.vikray_orders "
        "WHERE org_id=$1::uuid AND is_active=TRUE AND "
        + ("order_date BETWEEN $2::date AND $3::date "
           if win else "order_date >= (CURRENT_DATE - INTERVAL '6 months') ")
        + "GROUP BY month ORDER BY month",
        *([org_id, win.start, win.end] if win else [org_id]),
    )

    status_split = await pool.fetch(
        "SELECT status, COUNT(*) AS count, COALESCE(SUM(total),0) AS value "
        "FROM staging.vikray_orders WHERE org_id=$1::uuid AND is_active=TRUE "
        "GROUP BY status",
        org_id,
    )

    leaderboard = []
    if "graha" in allowed:
        leaderboard = await pool.fetch(
        "SELECT t.salesperson_id, "
        # THE LEADERBOARD NAMES PEOPLE BY NAME. This ended `…, u.email)`, so a
        # salesperson with no name recorded appeared on an analytics
        # leaderboard as their EMAIL ADDRESS — a contact detail rendered as a
        # label, and on a screen Aekam staff read, which inverts the rule that
        # Aekam must not see a customer's member emails. The owner's ruling
        # (2026-08-23): a display-name ladder must never end at an email.
        #
        # MEASURED BEFORE THE CHANGE: 0 of 35 live accounts have neither
        # `full_name` nor `name`, so this rung has never fired on real data.
        # Removing it changes nothing visible; leaving it was a standing risk.
        #
        # The terminal is a STATED label, not blank — a blank name on a
        # leaderboard row reads as a row about nobody, which is false; the row
        # is about a real person whose name we simply do not hold. Ladder owned
        # by `services/audit_actors.display_name()`; it emits no `$n`.
        + display_name("u")
        + " AS name, "
        "t.target_amount, "
        "COALESCE(d.won, 0) AS actual_amount, "
        "CASE WHEN t.target_amount > 0 "
        "  THEN ROUND(COALESCE(d.won,0) / t.target_amount * 100, 1) "
        "  ELSE 0 END AS pct "
        "FROM staging.vikray_targets t "
        "LEFT JOIN users u ON u.user_id = t.salesperson_id "
        "LEFT JOIN LATERAL ("
        "  SELECT COALESCE(SUM(value),0) AS won FROM staging.graha_deals "
        "  WHERE org_id=$1::uuid AND stage='Won' AND owner_id::text = t.salesperson_id "
        "  AND updated_at >= t.period_start AND updated_at < t.period_end + 1"
        ") d ON TRUE "
        "WHERE t.org_id=$1::uuid AND t.period_start <= CURRENT_DATE AND t.period_end >= CURRENT_DATE "
        "ORDER BY pct DESC",
        org_id,
    )

    return {
        "order_trend": [dict(r) for r in order_trend],
        "status_split": [dict(r) for r in status_split],
        "leaderboard": [dict(r) for r in leaderboard],
        # Named so the UI can say the leaderboard is withheld rather than draw
        # an empty board, which reads as "nobody sold anything".
        "withheld": [] if "graha" in allowed else ["leaderboard"],
        "window": aw.describe(
            win, windowed=["order_trend"], as_at=["status_split", "leaderboard"]),
    }


# ── Saved Dashboards CRUD ───────────────────────────────────

@router.get("/dashboards")
async def list_dashboards(
    user=Depends(require_user),
    org_id: str = Depends(get_org_id),
    _g=Depends(_gate),
):
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT * FROM staging.dristi_dashboards "
        "WHERE org_id=$1::uuid AND is_active=TRUE ORDER BY is_default DESC, name",
        org_id,
    )
    return {"data": [dict(r) for r in rows]}


@router.post("/dashboards")
async def create_dashboard(
    body: DashboardCreate,
    user=Depends(require_user),
    org_id: str = Depends(get_org_id),
    _g=Depends(_gate),
):
    pool = await get_pool()
    if body.is_default:
        await pool.execute(
            "UPDATE staging.dristi_dashboards SET is_default=FALSE WHERE org_id=$1::uuid",
            org_id,
        )
    row = await pool.fetchrow(
        "INSERT INTO staging.dristi_dashboards (org_id, name, description, widgets, is_default, created_by) "
        "VALUES ($1::uuid, $2, $3, $4::jsonb, $5, $6) RETURNING *",
        org_id, body.name, body.description, json.dumps(body.widgets), body.is_default, user["user_id"],
    )
    return dict(row)


@router.patch("/dashboards/{dash_id}")
async def update_dashboard(
    dash_id: str,
    body: DashboardUpdate,
    user=Depends(require_user),
    org_id: str = Depends(get_org_id),
    _g=Depends(_gate),
):
    pool = await get_pool()
    updates, vals = [], []
    if body.name is not None:
        vals.append(body.name); updates.append(f"name=${len(vals)}")
    if body.description is not None:
        vals.append(body.description); updates.append(f"description=${len(vals)}")
    if body.widgets is not None:
        vals.append(json.dumps(body.widgets)); updates.append(f"widgets=${len(vals)}::jsonb")
    if body.is_default is not None:
        if body.is_default:
            await pool.execute("UPDATE staging.dristi_dashboards SET is_default=FALSE WHERE org_id=$1::uuid", org_id)
        vals.append(body.is_default); updates.append(f"is_default=${len(vals)}")
    if not updates:
        raise HTTPException(400, "Nothing to update")
    updates.append("updated_at=NOW()")
    vals += [dash_id, org_id]
    row = await pool.fetchrow(
        f"UPDATE staging.dristi_dashboards SET {', '.join(updates)} "
        f"WHERE id=${len(vals)-1}::uuid AND org_id=${len(vals)}::uuid RETURNING *",
        *vals,
    )
    if not row:
        raise HTTPException(404, "Dashboard not found")
    return dict(row)


@router.delete("/dashboards/{dash_id}")
async def delete_dashboard(
    dash_id: str,
    user=Depends(require_user),
    org_id: str = Depends(get_org_id),
    _g=Depends(_gate),
):
    pool = await get_pool()
    result = await pool.execute(
        "UPDATE staging.dristi_dashboards SET is_active=FALSE, updated_at=NOW() "
        "WHERE id=$1::uuid AND org_id=$2::uuid",
        dash_id, org_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Dashboard not found")
    return {"ok": True}


# ── Scheduled Reports ───────────────────────────────────────

class ScheduledReportCreate(BaseModel):
    name: str
    report_type: str = "overview"
    frequency: str = "weekly"
    day_of_week: int | None = None
    day_of_month: int | None = None
    time_utc: str = "08:00"
    file_formats: list[str] = ["pdf"]
    recipients: list[str]
    dashboard_id: str | None = None
    filters: dict = {}


class ScheduledReportUpdate(BaseModel):
    name: str | None = None
    frequency: str | None = None
    day_of_week: int | None = None
    day_of_month: int | None = None
    time_utc: str | None = None
    file_formats: list[str] | None = None
    recipients: list[str] | None = None
    filters: dict | None = None
    is_active: bool | None = None


@router.get("/scheduled-reports", dependencies=[Depends(_gate)])
async def list_scheduled_reports(user=Depends(require_user), org_id=Depends(get_org_id)):
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT sr.*, d.name AS dashboard_name "
        "FROM staging.dristi_scheduled_reports sr "
        "LEFT JOIN staging.dristi_dashboards d ON d.id = sr.dashboard_id "
        "WHERE sr.org_id=$1::uuid ORDER BY sr.created_at DESC",
        org_id,
    )
    return {"data": [dict(r) for r in rows]}


@router.post("/scheduled-reports", dependencies=[Depends(_gate)])
async def create_scheduled_report(
    body: ScheduledReportCreate,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    valid_types = ("overview", "revenue", "pipeline", "hr", "sales", "custom")
    if body.report_type not in valid_types:
        raise HTTPException(400, f"report_type must be one of: {', '.join(valid_types)}")
    valid_freq = ("daily", "weekly", "monthly")
    if body.frequency not in valid_freq:
        raise HTTPException(400, f"frequency must be one of: {', '.join(valid_freq)}")
    if not body.recipients:
        raise HTTPException(400, "At least one recipient is required")

    row = await pool.fetchrow(
        "INSERT INTO staging.dristi_scheduled_reports "
        "(org_id, dashboard_id, name, report_type, frequency, day_of_week, "
        " day_of_month, time_utc, file_formats, recipients, filters, created_by) "
        "VALUES ($1::uuid, NULLIF($2,'')::uuid, $3, $4, $5, $6, $7, $8, $9, $10, $11::jsonb, $12) "
        "RETURNING id, name",
        org_id, body.dashboard_id or "", body.name, body.report_type,
        body.frequency, body.day_of_week, body.day_of_month,
        _parse_time(body.time_utc), body.file_formats, body.recipients,
        json.dumps(body.filters), user["user_id"],
    )
    return {"status": "created", **dict(row)}


@router.patch("/scheduled-reports/{report_id}", dependencies=[Depends(_gate)])
async def update_scheduled_report(
    report_id: str,
    body: ScheduledReportUpdate,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    updates = {k: v for k, v in body.dict(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(400, "No fields to update")
    sets, params = [], [report_id, org_id]
    idx = 3
    for k, v in updates.items():
        if k == "time_utc":
            sets.append(f"{k}=${idx}::time")
        elif k == "filters":
            sets.append(f"{k}=${idx}::jsonb")
            v = json.dumps(v)
        elif k in ("file_formats", "recipients"):
            sets.append(f"{k}=${idx}")
        else:
            sets.append(f"{k}=${idx}")
        params.append(v)
        idx += 1
    await pool.execute(
        f"UPDATE staging.dristi_scheduled_reports SET {', '.join(sets)} "
        f"WHERE id=$1::uuid AND org_id=$2::uuid",
        *params,
    )
    return {"status": "updated"}


@router.delete("/scheduled-reports/{report_id}", dependencies=[Depends(_gate)])
async def delete_scheduled_report(
    report_id: str,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    await pool.execute(
        "DELETE FROM staging.dristi_scheduled_reports "
        "WHERE id=$1::uuid AND org_id=$2::uuid",
        report_id, org_id,
    )
    return {"status": "deleted"}


async def _deliver_scheduled_report(pool, report) -> int:
    """Build one scheduled report, mail it, log it, stamp `last_sent_at`.

    Shared by the person-pressed `run-now` button and the cron sweep, because
    the only difference between those two is WHO decided the report should go
    out — everything after that decision must be identical, or the sweep
    becomes a second, slightly-different way to mail the books.

    Raises on failure after writing the 'failed' log row. The two callers want
    opposite things from a failure — run-now must answer 500 to the person
    waiting, the sweep must count it and carry on to the next schedule — so the
    log row is written here, where the error is, and the HTTP decision is left
    to the caller.

    Returns the number of recipients mailed.
    """
    report_id = str(report["id"])
    org_id = str(report["org_id"])
    try:
        # ONE delivery spine (proposal 65 S4). This used to mail a raw JSON
        # dump of `_fetch_report_data` in a preformatted block; it now renders
        # the SAME letterhead document `services/module_report` gives the
        # module page's download and Niyam's `report.send` — and applies the
        # same two contracts: 'custom' has no module arrangement and refuses
        # in words, and recipients are cut to MEMBERS of the org, because a
        # schedule's recipient list is free text and nothing this product
        # mails on a timer may leave the firm.
        import asyncio as _asyncio

        from services.gst_period import load_org
        from services.module_report import (
            MODULE_TITLES, REPORT_TYPE_MODULES, member_recipients,
            module_arrangement, render_report_html, report_entry,
            schedule_window)

        module = REPORT_TYPE_MODULES.get(report["report_type"])
        if module is None:
            raise ValueError(
                f"a {report['report_type']!r} report has no module "
                f"arrangement to render — custom dashboard delivery is not "
                f"built yet")
        members, skipped = await member_recipients(
            pool, org_id, report["recipients"])
        if not members:
            raise ValueError("none of the schedule's recipients is a member "
                             "of this org — reports mail members only")

        label = MODULE_TITLES.get(module, module)
        win = schedule_window(report["frequency"], report["last_sent_at"])
        layout, _source = await module_arrangement(pool, None, org_id, module)
        gate_cache: dict = {}
        # `report_entry`, NOT `report_widget`. A saved layout may hold a
        # SECTION ({"report": ...}) as well as a metric widget, and
        # `report_widget` handed a section renders "This metric is no longer
        # measured" under the label "None" — a register silently replaced by a
        # wrong sentence, on a document this code EMAILS. It does not raise,
        # which is what makes it dangerous: nobody finds out.
        #
        # `report_entry` dispatches on `is_section` — the one test the
        # validator and the renderer share — so this door and /module-report
        # cannot disagree about what a layout entry is.
        widgets = [await report_entry(pool, org_id, module, win, w,
                                      None, gate_cache)
                   for w in layout]
        org = await load_org(pool, org_id)
        period_line = (f"{win.start.strftime('%d %b %Y')} – "
                       f"{win.end.strftime('%d %b %Y')}")
        # Off the loop: the letterhead's logo embed performs a BLOCKING
        # httpx.get (up to 4 MB).
        html_doc = await _asyncio.to_thread(
            render_report_html, org, label, period_line, widgets)

        # `send_email` is sync (it threads internally) and it is the single
        # choke point that honours OUTBOUND_MODE, `_safe_subject` and
        # `outbound_log` — going through it is what keeps dry runs dry.
        from email_service import send_email

        # Scoped per report, not once around the sweep's loop. This helper is
        # called in a loop that walks EVERY org, so a scope set once outside it
        # would file every send after the first under the previous org — worse
        # than the NULL it replaces, because a wrong org id reads as a fact.
        # Under `run-now` the middleware has already set the same value, so
        # this is a no-op there; under cron there is no request and no
        # middleware, and without it every scheduled send lands on `/outbound`
        # with a NULL org, invisible to every org forever.
        from outbound import org_scope

        with org_scope(org_id):
            for recipient in members:
                send_email(
                    to_email=recipient,
                    subject=f"{report['name']} — {label} report",
                    html_content=html_doc,
                    # 098 asks for the 'unclassified' bucket to be watched
                    # falling; this sender was in it. It is also what tells
                    # `services/email_senders.py` to send from the
                    # notifications address rather than the default one.
                    purpose="report",
                    ref=f"report:{report_id}",
                )

        # `org_id` on the log row was never populated, so every delivery record
        # this table holds is org-NULL and the per-org log view cannot find it.
        await pool.execute(
            "INSERT INTO staging.dristi_report_logs "
            "(scheduled_report_id, org_id, status, recipients_count, error) "
            "VALUES ($1::uuid, $2::uuid, 'sent', $3, $4)",
            report_id, org_id, len(members),
            (f"{skipped} recipient(s) skipped — not members of this org"
             if skipped else None),
        )
        await pool.execute(
            "UPDATE staging.dristi_scheduled_reports SET last_sent_at=NOW() WHERE id=$1::uuid",
            report_id,
        )
        return len(members)
    except Exception as e:
        await pool.execute(
            "INSERT INTO staging.dristi_report_logs "
            "(scheduled_report_id, org_id, status, error) "
            "VALUES ($1::uuid, $2::uuid, 'failed', $3)",
            report_id, str(report["org_id"]), str(e),
        )
        raise


@router.post("/scheduled-reports/{report_id}/run-now", dependencies=[Depends(_gate)])
async def run_report_now(
    report_id: str,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    report = await pool.fetchrow(
        "SELECT * FROM staging.dristi_scheduled_reports "
        "WHERE id=$1::uuid AND org_id=$2::uuid",
        report_id, org_id,
    )
    if not report:
        raise HTTPException(404, "Scheduled report not found")

    # Same source check as the export endpoint. Running a report on demand must
    # not reach further than reading it would.
    required = _REPORT_SOURCE_MODULES.get(report["report_type"], set())
    missing = required - await reachable_modules(pool, user["user_id"], org_id, required)
    if missing:
        raise HTTPException(
            403,
            f"This report reads {', '.join(sorted(missing))}, which you do not "
            f"have access to.",
        )

    try:
        count = await _deliver_scheduled_report(pool, report)
        return {"status": "sent", "recipients": count}
    except Exception as e:
        raise HTTPException(500, f"Report generation failed: {e}")


@router.get("/scheduled-reports/{report_id}/logs", dependencies=[Depends(_gate)])
async def get_report_logs(
    report_id: str,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    # Logs carry recipient addresses and failure reasons, and are keyed on the
    # report id alone — so any report id in any org returned its delivery history.
    if not await pool.fetchval(
        "SELECT 1 FROM staging.dristi_scheduled_reports WHERE id=$1::uuid AND org_id=$2::uuid",
        report_id, org_id,
    ):
        raise HTTPException(404, "Scheduled report not found")
    rows = await pool.fetch(
        "SELECT * FROM staging.dristi_report_logs "
        "WHERE scheduled_report_id=$1::uuid ORDER BY sent_at DESC LIMIT 50",
        report_id,
    )
    return {"logs": [dict(r) for r in rows]}


# ── The sweep ────────────────────────────────────────────────────────────────
#
# ARMING. This endpoint is complete and it is DELIBERATELY INERT until someone
# sets DRISTI_REPORT_SWEEP_ARMED. Unarmed it does every read, every due
# calculation and every entitlement check, and then returns what it WOULD have
# sent without sending anything, writing any log row or moving any
# `last_sent_at`.
#
# That is not caution for its own sake. `outbound.MODE` defaults to "live"
# (`outbound.py:148`) and PRODUCTION HAS OUTBOUND_MODE UNSET, so the first tick
# of an armed sweep in production is real mail to real addresses on rows that
# have been sitting unsent since July. Whether that backlog goes out is the
# owner's decision, and it is not one that should be taken by merging a
# dispatcher. Arming is a separate, deliberate, one-variable act — and the dry
# preview is there so it can be made with the actual list in hand.
#
# Read at call time rather than at import, so the preview can be exercised and
# the arming decision verified without a redeploy of this module's import
# semantics being part of the question.
_SWEEP_ARMED_VAR = "DRISTI_REPORT_SWEEP_ARMED"


def _sweep_armed() -> bool:
    return os.getenv(_SWEEP_ARMED_VAR, "").strip().lower() in ("1", "true", "yes", "on")


@router.post("/scheduled-reports/dispatch", dependencies=[])
async def dispatch_scheduled_reports(
    x_cron_secret: str = Header(""),
    preview: bool = False,
):
    """Cron sweep over `staging.dristi_scheduled_reports`. Every org, one tick.

    NOT the same job as `POST /api/reports/dispatch`, and not a duplicate of it.
    That one walks `public.report_schedules`, which is keyed on `team_id`,
    carries `next_run_at`, and renders a TEAM TIME-TRACKING report — time
    entries, per-member task counts, daily throughput — into a purpose-built
    PDF/Excel mail. This one walks a table keyed on `org_id` whose reports are
    cross-module business figures (invoices, deals, headcount, orders), which
    are gated per source module and have no PDF renderer at all. They share the
    word "report" and nothing else: different tenancy key, different source
    tables, different renderer, different authorisation rule, different
    scheduling state. Folding either into the other would mean carrying both
    sets of behaviour behind one `if`, which is two systems wearing one name.
    What they SHOULD share is one timer and one due-rule, and the due-rule now
    lives in `services/report_schedule_window.py` for whoever schedules that one.

    Authenticated by CRON_SECRET in the `X-Cron-Secret` header, matching
    `routers/scheduler.py`. Not by a query parameter: a secret in a query string
    is written to every access log, proxy log and platform request log the
    request passes through, and those outlive the secret.

    `?preview=true` forces the dry listing even when armed, so an operator can
    ask what the next tick will do without waiting for it.
    """
    # Same constant-time comparison and same secret as every other cron in this
    # product. `!=` on a str short-circuits at the first differing byte, so the
    # time to fail leaks how many leading bytes were correct, and a cron
    # endpoint can be called as often as an attacker likes.
    from utils import secret_matches

    if not secret_matches(x_cron_secret, os.getenv("CRON_SECRET", "")):
        raise HTTPException(403, "Invalid cron secret")

    pool = await get_pool()

    # `now` is captured ONCE, here, and threaded through every due calculation
    # below. A sweep that called datetime.now() per row would use a different
    # instant for each schedule, and a sweep that used date.today() anywhere
    # would be reading the UTC date — which, in the 19:00-22:00 UTC window, is
    # already the previous day in IST and would stamp every report with
    # yesterday's Indian business date.
    now = datetime.now(timezone.utc)

    rows = await pool.fetch(
        "SELECT id, org_id, name, report_type, frequency, day_of_week, "
        "       day_of_month, time_utc, recipients, is_active, "
        "       last_sent_at, created_at, created_by "
        "FROM staging.dristi_scheduled_reports WHERE is_active = TRUE"
    )

    due, skipped = [], []
    for r in rows:
        if not is_due(
            now,
            is_active=bool(r["is_active"]),
            frequency=r["frequency"],
            day_of_week=r["day_of_week"],
            day_of_month=r["day_of_month"],
            time_utc=r["time_utc"],
            last_sent_at=r["last_sent_at"],
            created_at=r["created_at"],
        ):
            continue

        # Entitlement is re-checked on every tick against the schedule's owner,
        # not trusted from creation time. See `blocked_reason`.
        required = _REPORT_SOURCE_MODULES.get(r["report_type"], set())
        reachable = (
            await reachable_modules(pool, r["created_by"], str(r["org_id"]), required)
            if r["created_by"] and required
            else set()
        )
        reason = blocked_reason(r["report_type"], r["created_by"], required, reachable)
        if reason:
            log.warning("Dristi report sweep skipping %s (%s): %s", r["id"], r["name"], reason)
            skipped.append({"id": str(r["id"]), "name": r["name"], "reason": reason})
            continue

        due.append(r)

    armed = _sweep_armed()
    listing = [
        {"id": str(r["id"]), "org_id": str(r["org_id"]), "name": r["name"],
         "report_type": r["report_type"], "recipients": len(r["recipients"] or [])}
        for r in due
    ]

    if preview or not armed:
        log.info(
            "Dristi report sweep: %s active, %s due, %s skipped — NOT SENDING (%s)",
            len(rows), len(due), len(skipped),
            "preview requested" if preview else f"{_SWEEP_ARMED_VAR} is not set",
        )
        return {
            "armed": armed,
            "sent": 0,
            "would_send": listing,
            "skipped": skipped,
            "note": (
                f"No mail was sent and no row was modified. Set {_SWEEP_ARMED_VAR}=true "
                f"to arm this sweep. Note that OUTBOUND_MODE is unset in production and "
                f"outbound defaults to live, so arming there sends real mail on the next tick."
            ),
        }

    sent, failed = 0, []
    for r in due:
        try:
            # A full row is needed by the delivery helper; the listing query
            # above is deliberately narrow so the due decision reads only what
            # it uses.
            full = await pool.fetchrow(
                "SELECT * FROM staging.dristi_scheduled_reports WHERE id=$1::uuid", str(r["id"])
            )
            if not full:
                # Deleted between the two queries. Not an error.
                continue
            sent += await _deliver_scheduled_report(pool, full)
        except Exception as e:
            # One schedule's failure must not stop the other orgs' reports.
            # `_deliver_scheduled_report` has already written the 'failed' log
            # row against this schedule before re-raising.
            log.error("Dristi report sweep failed for %s (%s): %s", r["id"], r["name"], e)
            failed.append({"id": str(r["id"]), "name": r["name"], "error": str(e)})

    log.info(
        "Dristi report sweep: %s due, %s recipients mailed, %s failed, %s skipped",
        len(due), sent, len(failed), len(skipped),
    )
    # A cron that could not do its job must not answer 200 — the convention this
    # product settled on in `routers/scheduler.py`. Every schedule failing is a
    # broken sweep, and a silent 200 is how the last one went unnoticed.
    if failed and len(failed) == len(due):
        raise HTTPException(500, {"job": "dristi-reports", "failed": failed})

    return {"armed": True, "sent": sent, "due": len(due),
            "failed": failed, "skipped": skipped}


# Lifted to `services/report_render.py` so `/api/v1/analytics/run` can share
# the same format negotiation (D2). The aliases keep every call site and every
# test in this router untouched; byte-identical output is pinned by
# tests/test_report_render.py.
from services.report_render import csv_cell as _csv_cell, is_row_list as _is_row_list


@router.get("/exports/{report_type}", dependencies=[Depends(_gate)])
async def export_report(
    report_type: str,
    format: str = "json",
    date_from: str = "",
    date_to: str = "",
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    valid_types = ("overview", "revenue", "pipeline", "hr", "sales")
    if report_type not in valid_types:
        raise HTTPException(400, f"report_type must be one of: {', '.join(valid_types)}")

    required = _REPORT_SOURCE_MODULES.get(report_type, set())
    missing = required - await reachable_modules(pool, user["user_id"], org_id, required)
    if missing:
        raise HTTPException(
            403,
            f"This export reads {', '.join(sorted(missing))}, which you do not "
            f"have access to.",
        )

    win = aw.parse(date_from, date_to)
    data = await _fetch_report_data(pool, org_id, report_type, win)

    # `pipeline` and `hr` are the two STOCK reports: `_fetch_report_data`
    # applies no window to either — a standing pipeline and a headcount are
    # true AS AT today, whatever dates were asked for. Their artefacts must
    # say "as at" rather than implying a period was applied where none was.
    is_stock = report_type in ("pipeline", "hr")

    if format in ("csv", "xlsx", "pdf"):
        # The window — and, since D3b, the org — belong in the artefact, not
        # only in the query that built it: these files get forwarded, and a
        # file that does not say whose numbers it holds or which dates they
        # cover is indistinguishable from anyone else's covering all of them.
        # For CSV the identity rides in the FILENAME alone (a banner row above
        # the header breaks the first parser it meets); xlsx and pdf reuse the
        # same stem so the three formats of one report sort together. The org
        # comes via `load_org`, not a raw SELECT — it R2-signs the logo the
        # PDF letterhead embeds, and a bare `logo_key` URL is not fetchable.
        from services.gst_period import load_org
        from services.report_render import org_slug

        org = await load_org(pool, org_id)
        period = (
            f"as-at-{date.today().isoformat()}" if is_stock
            else f"{win.start.isoformat()}_{win.end.isoformat()}" if win
            else "all-time"
        )
        stem = f"{org_slug(org.get('name'))}_{report_type}_{period}"

    if format == "csv":
        import csv
        import io
        output = io.StringIO()
        if isinstance(data, list) and data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows([{k: _csv_cell(v) for k, v in row.items()} for row in data])
        elif isinstance(data, dict):
            writer = csv.writer(output)
            # A value here is either a scalar or a list of rows, and the two
            # cannot share a shape. `writerow([k, v])` was used for both, so a
            # list of rows went through str() and landed in ONE cell as Python
            # source: measured live on staging, revenue_export.csv contained
            #   monthly,"[{'month': datetime.datetime(2026, 7, 1, 0, 0, ...),
            #              'total': Decimal('311671.60'), 'count': 6}]"
            # which is not openable as a spreadsheet in any useful sense.
            # pipeline and sales had the same shape; overview and hr looked
            # correct only because every value they carry is a scalar.
            scalars = [(k, v) for k, v in data.items() if not _is_row_list(v)]
            tables = [(k, v) for k, v in data.items() if _is_row_list(v)]
            for k, v in scalars:
                writer.writerow([k, _csv_cell(v)])
            for k, rows in tables:
                # Blank line then a titled block, so several tables can share
                # one file and still be readable. Excel keeps the sections
                # visually separate and a parser can split on the empty row.
                if output.getvalue():
                    writer.writerow([])
                writer.writerow([k])
                headers = list(rows[0].keys())
                writer.writerow(headers)
                for row in rows:
                    writer.writerow([_csv_cell(row.get(h)) for h in headers])
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            # The period rides in the filename rather than in a banner row: a
            # comment line above the header breaks the first parser it meets.
            headers={"Content-Disposition": f"attachment; filename={stem}.csv"},
        )

    # ── xlsx and pdf ─────────────────────────────────────────────────────────
    #
    # Both formats were accepted and silently ignored: every value other than
    # `csv` fell through to the JSON return below, so a caller asking for PDF
    # received JSON, and the body labelled itself `"format":"json"` while doing
    # it. Measured live on staging across all five report types.
    #
    # Neither needs a new dependency. `openpyxl` and `weasyprint` are both in
    # requirements.txt, and `doc_render.render_pdf` is described there as "the
    # single PDF path" — the same toolchain behind invoice and payslip PDFs, so
    # a report PDF cannot drift from the documents the firm already sends.
    #
    # The shape rule is the one the CSV branch established: a value is either a
    # scalar or a list of rows, and the two cannot share a presentation. Scalars
    # become a summary block; each row-list becomes its own sheet or its own
    # table.
    if format in ("xlsx", "pdf"):
        scalars, tables = [], []
        if isinstance(data, list) and data:
            tables = [(report_type, data)]
        elif isinstance(data, dict):
            scalars = [(k, v) for k, v in data.items() if not _is_row_list(v)]
            tables = [(k, v) for k, v in data.items() if _is_row_list(v)]

        title = f"{report_type.title()} report"
        # The period line the PDF letterhead and the workbook header both
        # carry. For the two stocks it says "As at <today>" — the letterhead
        # must not imply a period was applied where none was.
        period_line = (
            f"As at {date.today().strftime('%d %b %Y')}" if is_stock
            else f"{win.start.strftime('%d %b %Y')} – {win.end.strftime('%d %b %Y')}" if win
            else "All time"
        )
        generated_at = datetime.now(timezone.utc)
        if format == "xlsx":
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            # The identity header: whose numbers, what report, which period,
            # made when. "Generated" is a real instant with a timezone —
            # `date.today()` said nothing about WHEN the numbers were true,
            # and these files get forwarded across timezones.
            ws.append([org.get("name") or "Organisation name not set"])
            ws["A1"].font = Font(bold=True, size=14)
            ws.append([title])
            ws.append([period_line])
            ws.append([f"Generated {generated_at.isoformat(timespec='seconds')}"])
            ws.append([])
            for k, v in scalars:
                ws.append([k, _csv_cell(v)])
            if not scalars:
                ws.append(["No summary values for this report."])

            for k, rows in tables:
                # Excel sheet names cap at 31 chars and reject []:*?/\
                safe = "".join(c for c in str(k) if c not in "[]:*?/\\")[:31] or "Data"
                sheet = wb.create_sheet(safe)
                headers = list(rows[0].keys())
                sheet.append(headers)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                for row in rows:
                    sheet.append([_csv_cell(row.get(h)) for h in headers])

            buf = io.BytesIO()
            wb.save(buf)
            return Response(
                content=buf.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'},
            )

        # pdf — the branded page (D3b): org name and logo, title, the exact
        # window or an honest "As at", and the generation timestamp; NO GSTIN,
        # NO address, NO phone (owner's ruling, 17 Aug 2026 — a working
        # document the firm may hand to its client, not a tax document).
        # Assembled the way routers/analytics.py's pdf branch assembles its
        # page, through the same doc_render toolchain behind invoice and
        # payslip PDFs — the old anonymous "<h1>" document said nothing about
        # whose numbers it carried, and these files get forwarded.
        from services import doc_render as R
        from services.report_render import analytics_letterhead, pdf_table, summary_table

        def _build_and_render() -> bytes:
            # The whole build runs off the loop, not just WeasyPrint: the
            # letterhead's `embed_logo` performs a BLOCKING httpx.get for the
            # R2-signed logo, and a slow fetch on the loop stalls every
            # request this worker holds. Same fix as routers/analytics.py.
            parts = [analytics_letterhead(org, title, "", period_line)]
            if scalars:
                parts.append(summary_table(scalars))
            parts += [pdf_table(k, rows) for k, rows in tables]
            if not scalars and not tables:
                parts.append("<p>This report produced no rows for this period.</p>")
            generated = generated_at.strftime("%d %b %Y, %H:%M UTC")
            parts.append(R.foot(f"Generated {R.esc(generated)} &middot; Prepared in Kartavya"))

            html_doc = R.document(
                ["".join(parts)], org, title=f"{title} — Kartavaya",
                running=R.running_id(title, org, period_line),
            )
            return R.render_pdf(html_doc)

        return Response(
            content=await asyncio.to_thread(_build_and_render),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{stem}.pdf"'},
        )

    return {"data": data, "format": "json", "window": win.as_dict() if win else None}


# ── Custom Dashboard / Pivot Query Builder ──────────────────

# `module` is what the source belongs to, and is checked against the caller's
# entitlement before the query runs. Without it the pivot builder was a
# general-purpose reader over eight tables — including the invoice ledger and
# the employee register — behind the `dristi` grant alone.
#
# `soft_delete` records whether the table actually has an `is_active` column
# rather than inferring it. All eight do today, which is why the broken
# always-true check that preceded this never misfired; the next source added
# would have been the one to break.
#
# `not_draft` is declared the same way and for the same reason. Only
# `ganit_invoices` has a `doc_status` column, so a draft filter appended to
# every source would turn the other seven into UndefinedColumn 500s — the flag
# is what keeps the next source added from inheriting a predicate its table
# cannot answer.
#
# `on_the_rolls` is the third of these, and it exists because `soft_delete` is
# NOT the same question. `is_active` is a flag somebody has to remember to
# clear, and on `manav_employees` a leaver deliberately keeps it until
# settlement (`routers/manav.py:1958`); `manav_offboarding.last_working_day` is
# the fact. Only `manav_employees` has an exit register, so this too is
# declared per source: hung off `soft_delete` instead, it would correlate
# `manav_offboarding.employee_id` against `vikray_orders.id` and turn six
# perfectly good widgets into 500s.
_ALLOWED_QUERY_TABLES = {
    "invoices": {
        "table": "staging.ganit_invoices",
        "module": "ganit",
        "soft_delete": True,
        "not_draft": True,
        "columns": ["invoice_date", "invoice_type", "total", "subtotal", "amount_paid",
                     "payment_status", "currency", "created_at"],
        "date_col": "invoice_date",
    },
    "deals": {
        "table": "staging.graha_deals",
        "module": "graha",
        "soft_delete": True,
        "columns": ["stage", "value", "expected_close", "created_at", "updated_at"],
        "date_col": "created_at",
    },
    "contacts": {
        "table": "staging.graha_contacts",
        "module": "graha",
        "soft_delete": True,
        "columns": ["contact_type", "source", "company", "lead_score", "created_at"],
        "date_col": "created_at",
    },
    "orders": {
        "table": "staging.vikray_orders",
        "module": "vikray",
        "soft_delete": True,
        "columns": ["order_date", "status", "total", "subtotal", "created_at"],
        "date_col": "order_date",
    },
    "employees": {
        "table": "staging.manav_employees",
        "module": "manav",
        "soft_delete": True,
        "on_the_rolls": True,
        "columns": ["department", "designation", "employment_type", "status",
                     "date_of_joining", "created_at"],
        "date_col": "date_of_joining",
    },
    "expenses": {
        "table": "staging.ganit_expenses",
        "module": "ganit",
        "soft_delete": True,
        "columns": ["category", "amount", "total", "expense_date", "status", "created_at"],
        "date_col": "expense_date",
    },
    "tickets": {
        "table": "staging.graha_tickets",
        "module": "graha",
        "soft_delete": True,
        "columns": ["priority", "status", "category", "created_at", "resolved_at"],
        "date_col": "created_at",
    },
    "events": {
        "table": "staging.prachar_events",
        "module": "prachar",
        "soft_delete": True,
        "columns": ["event_type", "status", "starts_at", "created_at"],
        "date_col": "starts_at",
    },
}


class PivotQuery(BaseModel):
    source: str
    group_by: str = ""
    #: The COLUMN dimension. A pivot is two-dimensional — the rendered reference
    #: (`ScreensThin.jsx`, `DristiPivot`) draws client × quarter with a total per
    #: row, per column and a grand total. With one dimension the "pivot" tab was
    #: a two-column list, which is what `/query` already served the chart cards.
    #: Validated against the same per-source column whitelist as `group_by`, so
    #: it is never a caller-supplied SQL fragment.
    group_by2: str = ""
    measure: str = "count"
    date_from: str = ""
    date_to: str = ""
    filters: dict = {}


@router.post("/query", dependencies=[Depends(_gate)])
async def run_pivot_query(
    body: PivotQuery,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    spec = _ALLOWED_QUERY_TABLES.get(body.source)
    if not spec:
        raise HTTPException(400, f"source must be one of: {', '.join(_ALLOWED_QUERY_TABLES)}")

    pool = await get_pool()
    source_module = spec["module"]
    if not await reachable_modules(pool, user["user_id"], org_id, {source_module}):
        raise HTTPException(
            403,
            f"'{body.source}' reads the {source_module} module, which you do not "
            f"have access to.",
        )

    table = spec["table"]
    allowed = spec["columns"]

    if body.group_by and body.group_by not in allowed:
        raise HTTPException(400, f"group_by must be one of: {', '.join(allowed)}")
    if body.group_by2 and body.group_by2 not in allowed:
        raise HTTPException(400, f"group_by2 must be one of: {', '.join(allowed)}")
    if body.group_by2 and not body.group_by:
        raise HTTPException(400, "group_by2 requires group_by — columns need rows.")
    if body.group_by2 and body.group_by2 == body.group_by:
        raise HTTPException(400, "group_by and group_by2 must differ.")

    measure_sql = "COUNT(*)"
    if body.measure == "sum" and "total" in allowed:
        measure_sql = "COALESCE(SUM(total),0)"
    elif body.measure == "sum" and "value" in allowed:
        measure_sql = "COALESCE(SUM(value),0)"
    elif body.measure == "avg" and "total" in allowed:
        measure_sql = "COALESCE(AVG(total),0)"
    elif body.measure == "avg" and "value" in allowed:
        measure_sql = "COALESCE(AVG(value),0)"

    where = ["org_id=$1::uuid"]
    params: list = [org_id]

    # This was:
    #     if "is_active" in [c for t in [spec] for c in ["is_active"]]:
    # which is a comprehension over the literal ["is_active"] and so is ALWAYS
    # true — `is_active=TRUE` was appended for every source, including the ones
    # whose table has no such column. The intent was plainly to check the
    # source's own column list, and it is declared per source now.
    if spec.get("soft_delete"):
        where.append("is_active=TRUE")

    # `soft_delete` was this query's ONLY predicate beyond `org_id`, so
    # `source=invoices, measure=sum` returned the whole ledger — drafts
    # included. Live 2026-08-26 for E2E Test & Associates that was
    # Rs 12,29,86,008.58 against a draft-free Rs 11,14,93,756.12, the same
    # Rs 1,14,92,252.46 phantom the `/overview` tile was carrying, in a widget a
    # customer builds and pins to their own dashboard.
    #
    # `COALESCE(doc_status, '')` rather than a bare `<>` — the canonical form
    # from `services/gst_period.py`. The column is nullable and NULL <> 'draft'
    # is NULL, which would drop every invoice predating it.
    if spec.get("not_draft"):
        where.append("COALESCE(doc_status, '') <> 'draft'")

    # `::text::date`, not `::date`. `date_from` and `date_to` arrive as ISO
    # STRINGS on the request body. A bare `$2::date` makes Postgres infer the
    # parameter as `date`, and asyncpg then refuses the bind outright:
    #
    #   DataError: invalid input for query argument $2: '2026-01-01'
    #              ('str' object has no attribute 'toordinal')
    #
    # and this cast is built the same way for every source, so EVERY dated
    # pivot raised before reading a row — reproduced live 2026-08-26 on
    # `invoices` and on `deals`, which share nothing but this line. The double
    # cast makes the parameter infer as `text` and casts server-side, which is
    # what `documents.py`, `_tally_rows` and `services/gst_period.py` already
    # do. Nothing about the comparison changes.
    date_col = spec.get("date_col", "created_at")
    if body.date_from:
        params.append(body.date_from)
        where.append(f"{date_col} >= ${len(params)}::text::date")
    if body.date_to:
        params.append(body.date_to)
        where.append(f"{date_col} <= ${len(params)}::text::date")

    for fk, fv in body.filters.items():
        if fk in allowed:
            params.append(str(fv))
            where.append(f"{fk} = ${len(params)}")

    where_clause = " AND ".join(where)

    # The source is aliased so a correlated guard has something to correlate
    # AGAINST, and every source takes the alias rather than only the one that
    # needs it: one statement shape for all eight, and the next source to
    # declare `on_the_rolls` does not also have to remember to change its FROM
    # clause. The column references above stay unqualified — there is exactly
    # one relation in the FROM, so they resolve to it either way.
    source_sql = f"{table} {DEFAULT_ALIAS}"

    # STOCK. `employees` holds one row per PERSON, never one per period, so
    # even the dated pivot — `date_col` is `date_of_joining` — is a cohort cut
    # out of the current register rather than a flow of events, and the source
    # has ALWAYS applied `is_active=TRUE`, so a hand-deactivated employee had
    # already dropped out of it. What the guard changes is that somebody whose
    # last working day has passed drops out too.
    #
    # Live 2026-08-26 for E2E Test & Associates this widget read 83, and
    # grouped by `status` it said `{'active': 83}` for an org where 73 people
    # are employed. A customer builds this one and pins it to their own screen.
    #
    # ── AND ONLY WHEN THERE IS NO WINDOW ────────────────────────────────────
    #
    # `employees` declares `date_col: "date_of_joining"`, and `PivotTab.jsx`
    # posts `date_from`/`date_to` for anything other than "All time". So with a
    # window this source stops being "who is on the rolls" and becomes "who
    # JOINED inside these dates, by department" — a cohort over a past period,
    # which is a FLOW. Guarding it there would erase anybody who joined in the
    # window and has since left, and a hiring chart that silently drops your
    # leavers is worse than one that counts them.
    #
    # Unwindowed the source has no period at all, so it can only mean the
    # present, and the present is a stock. The two readings live in one source
    # because `date_of_joining` is the only date on the table; splitting them
    # into two sources is the better answer and is not this change.
    if spec.get("on_the_rolls") and not (body.date_from or body.date_to):
        where_clause += still_on_the_rolls(DEFAULT_ALIAS)

    if body.group_by and body.group_by2:
        # Both names are whitelist members, never caller text. The cap is on the
        # CELL count rather than the row count: 40 clients x 12 months is 480
        # rows the browser has to pivot, and a cross-tab that wide is unreadable
        # anyway. Ordered by label so the rendered grid is stable between runs.
        rows = await pool.fetch(
            f"SELECT {body.group_by} AS label, {body.group_by2} AS col, "
            f"{measure_sql} AS value "
            f"FROM {source_sql} WHERE {where_clause} "
            f"GROUP BY {body.group_by}, {body.group_by2} "
            f"ORDER BY 1, 2 LIMIT 600",
            *params,
        )
        return {
            "data": [dict(r) for r in rows],
            "source": body.source,
            "measure": body.measure,
            "group_by": body.group_by,
            "group_by2": body.group_by2,
        }

    if body.group_by:
        rows = await pool.fetch(
            f"SELECT {body.group_by} AS label, {measure_sql} AS value "
            f"FROM {source_sql} WHERE {where_clause} "
            f"GROUP BY {body.group_by} ORDER BY value DESC LIMIT 50",
            *params,
        )
        return {"data": [dict(r) for r in rows], "source": body.source, "measure": body.measure}
    else:
        row = await pool.fetchrow(
            f"SELECT {measure_sql} AS value, COUNT(*) AS count FROM {source_sql} WHERE {where_clause}",
            *params,
        )
        return {"data": dict(row) if row else {}, "source": body.source, "measure": body.measure}


@router.get("/widget-types", dependencies=[Depends(_gate)])
async def widget_types(
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    """The pivot builder's vocabulary — only the sources this caller can read.

    Two things this endpoint did not do:

    · It listed every source regardless of entitlement, so the builder offered
      `invoices` and `employees` to someone who gets a 403 the moment they press
      Run. Offering an option that cannot succeed is worse than omitting it.
    · It returned no columns, so the frontend carried its OWN source->columns
      map. The two had already drifted — the copy in `DristiPage.jsx` was
      missing `subtotal`, `amount_paid` and every `created_at`, so those
      groupings were unreachable from the UI although the server allowed them,
      and a column added to the whitelist would never have appeared. The
      whitelist is the only correct source for this list.
    """
    pool = await get_pool()
    reachable = await reachable_modules(
        pool, user["user_id"], org_id,
        {spec["module"] for spec in _ALLOWED_QUERY_TABLES.values()},
    )
    sources = {
        name: {"columns": spec["columns"], "date_col": spec.get("date_col", "created_at")}
        for name, spec in _ALLOWED_QUERY_TABLES.items()
        if spec["module"] in reachable
    }
    return {
        "sources": list(sources.keys()),
        "source_meta": sources,
        "measures": ["count", "sum", "avg"],
        "widget_types": ["number", "bar", "pie", "line", "table"],
        # How many of the eight sources are hidden by entitlement. The builder
        # says so out loud rather than looking like the product only has three.
        "withheld_count": len(_ALLOWED_QUERY_TABLES) - len(sources),
    }
