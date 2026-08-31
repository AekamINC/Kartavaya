"""analytics.py — the universal analytic contract. Proposal 62, phase D2.

One request shape for every analytic in the product:

    GET /api/v1/analytics/run?metric=ganit.invoiced
        &date_from=2026-04-01&date_to=2026-06-30
        &bucket=month&group_by=invoice_type
        &compare=previous_period
        &format=json            # json | csv | xlsx | pdf — SAME query, four renderings

`format` being a parameter rather than a separate endpoint is what makes
"download exactly what I am looking at" true rather than approximately true:
the file runs the same SQL with the same window and the same grouping as the
screen. An unknown format is a 400, never a silent fall-through to JSON —
dristi's export path ignored xlsx/pdf for months precisely that way.

── THE GATE, AND WHY IT IS NOT `held_level` ─────────────────────────────────
This router is deliberately NOT gated on the `dristi` module: a metric carries
its OWN module (proposal 62 fault 2 — buy Ganit, get Ganit's numbers). But
"can reach the module" is two different questions:

  · The CATALOGUE lists names. `held_level` (plain args, never raises, None
    means no) is the right instrument — listing a metric's existence is not
    serving its data.
  · `/run` serves DATA. Only `require_module(code)` checks the subscription
    state and writes the sensitive-module audit row for platform-role reads —
    `held_level` does neither, and a /run gated on it alone would serve
    financials for a lapsed module and leave no audit trail. It is called
    DIRECTLY, the way `routers/search.py` does: the dependency callable runs
    the identical checks FastAPI would run. It reads the authed user off
    `request.state`, which this route's own `Depends(require_user)` has set.
    Never give a helper a `Depends(...)` default — a plain call receives the
    sentinel, and that shape once 500'd every PATCH /api/tasks for ten days.

── FLOWS AND STOCKS (D1's rule, inherited) ──────────────────────────────────
A `flow` metric measures what happened DURING a period and REQUIRES
date_from/date_to here — `aw.parse`'s None-means-unchanged contract belongs
to the RETROFITTED dristi reads and is not reused as a default on a new
endpoint. A `stock` metric is true AS AT an instant: the window is ignored,
and the response says `as_at` rather than pretending a period was applied.
"""
import asyncio
import csv
import io
import json
import logging
import re
import uuid
from datetime import date, datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Request, Response
from pydantic import BaseModel, Field

from analytics.presets import PRESETS, VIZ_TYPES, preset_catalogue
from analytics.registry import REGISTRY, MetricRequest, catalogue_for, load_all, modules_in_registry
from analytics.windowing import BUCKETS, COMPARE_MODES, compare_window
from auth_router import require_user
from db import get_pool
from middleware.module_levels import held_level
from middleware.org_resolver import get_org_id
from middleware.subscription import org_module_refusal, require_module
from services import analytics_window as aw
from services.report_render import csv_cell

router = APIRouter(prefix="/api/v1/analytics", tags=["analytics"])
log = logging.getLogger(__name__)

# Every metrics module registers at import. A broken declaration fails HERE,
# at startup, not at the first request that touches it.
load_all()

FORMATS = ("json", "csv", "xlsx", "pdf")

#: Core PM is deliberately NOT module-gated — `role_tiers.py`'s own words: "the
#: surface a project-only role reaches is precisely the surface that is NOT
#: module-gated — core PM and notifications". `core` is not in ALL_MODULES and
#: no org_modules row for it will ever exist, so `require_module("core")`
#: would refuse every org its own task counts. Membership in the org — which
#: `require_user` + `get_org_id` have already established by the time any
#: route body runs — is exactly the entitlement core PM itself uses.
UNGATED_MODULES = frozenset({"core"})


async def _reachable(pool, user_id: str, org_id: str, also=frozenset(),
                     *, runnable: bool = False) -> set[str]:
    """Which registry modules this caller may SEE — the catalogue intersection.

    The same loop `routers/dristi.reachable_modules` runs, over `held_level`
    (plain arguments, never raises, None = may not read). Re-implemented
    rather than imported: a router must not be imported for a helper (see
    services/gst_period.py's note on that dependency direction).

    `also` adds module codes to ASK ABOUT that the metric registry does not
    name. A `ReportDef` declares its own `reads` and nothing requires those
    modules to carry a Dristi metric; without this, a register reading a
    metric-less module would be permanently withheld from
    `/report-sections` — withheld for a reason that has nothing to do with
    the caller's entitlement, and stated nowhere.

    ⚠ `runnable=True` ADDS THE ORG'S HALF OF THE GATE, and a menu needs it.

    `held_level` answers only "does this PERSON hold a grant", and it returns
    `admin` for any org owner or admin unconditionally. `/run` also asks
    whether the ORG has the module active and its subscription live. Two
    gates, one strictly weaker — and the weaker one drew the menu.

    Suite 12.03 measured it on the reference org 2026-08-31: **the catalogue
    offered 4 metrics `/run` refuses**, every one `varta.*`. That org holds
    twelve active modules and no `module_subscriptions` row for `varta` at
    all, so `varta.sends`, `varta.delivery_rate`, `varta.read_rate` and
    `varta.reply_rate` were pickable in "Add a metric…" and answered 403 the
    moment the widget drew. A menu that lists dishes the kitchen refuses is
    worse than a short menu: the person picks one, gets an error they cannot
    act on, and learns nothing about which module they would have to buy.

    The org half is `subscription.org_module_refusal` — the SAME function
    `require_module` now raises from, so the menu and the door cannot drift.
    It is called rather than `require_module` itself because that would run
    the platform branch once per module, and `platform_audit_needed` writes a
    row for every sensitive module a platform role reads; twelve rows per
    catalogue GET would bury the audit.

    `runnable` is OPT-IN, and `/report-sections` deliberately does not take it
    — see its own docstring, which argues the split it draws between listing a
    register's NAME and serving its ROWS. That argument is about `reads`, and
    this parameter does not touch it.
    """
    asked = (modules_in_registry() | set(also))
    out = set(UNGATED_MODULES & asked)
    for code in asked - UNGATED_MODULES:
        if await held_level(pool, user_id, org_id, code) is None:
            continue
        if runnable and await org_module_refusal(pool, org_id, code) is not None:
            continue
        out.add(code)
    return out


@router.get("/catalogue")
async def catalogue(
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    # `runnable=True`: this list IS the "Add a metric…" menu, so every entry
    # has to be something `/run` will actually answer. See `_reachable`.
    reachable = await _reachable(pool, user["user_id"], org_id, runnable=True)
    metrics = catalogue_for(reachable)
    return {
        "metrics": metrics,
        "buckets": sorted(BUCKETS),
        "compare_modes": sorted(COMPARE_MODES),
        "formats": list(FORMATS),
        # How many declarations entitlement hid — the same honesty line
        # /widget-types draws, so the UI can say "3 more with other modules"
        # instead of looking like the product is small.
        "withheld_count": len(REGISTRY) - len(metrics),
    }


@router.get("/report-sections")
async def report_sections(
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    """The ROW-LEVEL sections this caller may put on a module page.

    `/catalogue` lists the metrics a view may hold; this lists the registers.
    Without it the row-level framework is undiscoverable — a UI cannot offer
    what it cannot enumerate, and `_clean_layout` will not accept a key
    nobody can find.

    ── THE GATE, AND WHY IT IS `held_level` AND NOT `require_module` ─────────
    The same split `/catalogue` and `/run` already draw, applied to the same
    kind of object:

      · This route LISTS NAMES — key, label, one sentence of description. No
        row of anybody's books crosses it. `held_level` (plain args, never
        raises, None means no) is the right instrument, exactly as it is for
        `/catalogue`; listing a register's existence is not serving its rows.
      · The rows go out through `/module-report`, and THAT door runs
        `require_module` per module — subscription state, the
        sensitive-module audit row, platform-role refusal — and then
        `report_section` gates again on `ReportDef.reads` before running a
        single query. A section is never served on the strength of having
        been listed here.

    What this route does mirror from `/run` is WHOSE entitlement decides: the
    DEFINITION's own declared modules, never a blanket dristi grant (proposal
    62 fault 2 — buy Ganit, get Ganit's numbers). The one difference from a
    metric is that `reads` is a SET and ALL of it is required — a section
    joining a second module's table declared that module, and a partial
    export of the books is still an export of the books.
    """
    from services.report_defs import REPORT_DEFS, load_all, sections_for

    # Lazily, here rather than at import: a definition that fails to import
    # must fail where a caller can SAY so (the package's own rule).
    load_all()
    pool = await get_pool()
    # Every module any definition reads is asked about, even one the metric
    # registry never names — otherwise a register over a metric-less module is
    # withheld for a reason that is not the caller's entitlement.
    declared = {code for d in REPORT_DEFS.values() for code in d.reads}
    reachable = await _reachable(pool, user["user_id"], org_id, also=declared)
    sections = sections_for(reachable)
    return {
        "sections": sections,
        # How many declarations entitlement hid — `/catalogue`'s honesty line,
        # so a UI can say "3 more with other modules" instead of looking like
        # the product is small. `sections_for` has already run `load_all`, so
        # REPORT_DEFS is populated by the time it is counted.
        "withheld_count": len(REPORT_DEFS) - len(sections),
    }


@router.get("/run")
async def run(
    request: Request,
    metric: str,
    date_from: str = "",
    date_to: str = "",
    bucket: str = "month",
    group_by: str = "",
    compare: str = "",
    format: str = "json",
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    m = REGISTRY.get(metric)
    if m is None:
        raise HTTPException(404, f"unknown metric: {metric!r} — see /api/v1/analytics/catalogue")

    # The REAL gate, before anything else is inspected or computed: platform
    # reach, org role, per-user grant, subscription state, and the audit row
    # for sensitive modules. Its refusal is the response. Core PM is the one
    # ungated surface (see UNGATED_MODULES) — org membership, already proven
    # by get_org_id, is its whole entitlement.
    if m.module not in UNGATED_MODULES:
        await require_module(m.module)(request, org_id)

    if m.absent:
        # 422: the request is well-formed; the schema cannot answer it
        # honestly. The reason is the payload — proposal 62 §10, a stated
        # absence, never a convincing zero.
        raise HTTPException(422, {"metric": m.key, "absent": m.absent})

    if format not in FORMATS:
        raise HTTPException(400, f"format must be one of: {', '.join(FORMATS)}")
    if bucket not in BUCKETS:
        raise HTTPException(400, f"bucket must be one of: {', '.join(sorted(BUCKETS))}")
    if group_by and group_by not in m.dimensions:
        allowed = ", ".join(m.dimensions) or "(none)"
        raise HTTPException(400, f"group_by for {m.key} must be one of: {allowed}")
    if compare and compare not in COMPARE_MODES:
        raise HTTPException(400, f"compare must be one of: {', '.join(sorted(COMPARE_MODES))}")

    win = aw.parse(date_from, date_to)
    if m.grain == "flow" and win is None:
        raise HTTPException(
            400,
            f"{m.key} measures a flow — date_from and date_to are required. "
            "Stocks (as-at-today metrics) take no window; this is not one.",
        )
    if m.grain == "stock":
        # A stock has no period. The bounds are IGNORED rather than applied to
        # some arbitrary column, and the response says so — a date picker above
        # a headcount must not imply an authority it does not have.
        win = None

    pool = await get_pool()
    req = MetricRequest(org_id=org_id, window=win, bucket=bucket, group_by=group_by or None)
    sql, params = m.sql(req)
    rows = [dict(r) for r in await pool.fetch(sql, *params)]

    compared = None
    if compare and m.grain == "flow":
        cw = compare_window(win, compare)
        csql, cparams = m.sql(MetricRequest(org_id=org_id, window=cw, bucket=bucket,
                                            group_by=group_by or None))
        compared = {
            "mode": compare,
            "window": cw.as_dict(),
            "data": [dict(r) for r in await pool.fetch(csql, *cparams)],
        }

    payload = {
        "metric": m.key,
        "label": m.label,
        "unit": m.unit,
        "grain": m.grain,
        "group_by": group_by or None,
        "bucket": bucket if m.grain == "flow" else None,
        "window": (
            {**win.as_dict(), "windowed": [m.key], "as_at": []}
            if win is not None
            else {"as_at": date.today().isoformat(), "windowed": [], "note":
                  "stock metric — true as at today; any supplied bounds were ignored"}
        ),
        "as_of": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "data": rows,
        "compare": compared,
    }

    if format == "json":
        return payload

    stem = m.key.replace(".", "-") + (
        f"_{win.start.isoformat()}_{win.end.isoformat()}" if win
        else f"_as-at-{date.today().isoformat()}"
    )

    # ── WHAT AN EXPORT OF THIS METRIC IS, WRITTEN INTO THE FILE ────────────
    #
    # ⚠ A METRIC WITH NO ROWS DOWNLOADED AS SEVEN BYTES. `headers = ... else
    # ["value"]` and no rows produced the single line `value\r\n` — a file
    # that opens empty, says nothing, and is indistinguishable from a broken
    # export. Suite 12.10 caught it by weight (`metric-0-csv downloaded as an
    # EMPTY file`, 7 bytes against a floor of 20) and named the class: §1's
    # "a 200 with an empty body".
    #
    # The fix is not a bigger floor. A person who clicks "Download <metric> as
    # CSV" and opens the file has to be able to tell "this metric has no data
    # in this window" from "the export is broken", and neither the metric's
    # name nor the window was in the file at all — even when it DID have rows.
    #
    # The shape is the client report's, which already does this two hundred
    # lines below: identify, then a blank line, then the table.
    #
    # ⚠ NOT the label alone. `as_at`/`window` is the half that makes an empty
    # answer legible — "0 rows for 2026-04-01 to 2026-06-30" is a fact about
    # the question, and "0 rows" on its own is not.
    def _preamble():
        when = (f"{win.start.isoformat()} to {win.end.isoformat()}" if win
                else f"as at {date.today().isoformat()}")
        return [
            ("Metric", m.label or m.key),
            ("Key", m.key),
            ("Period", when),
            ("Rows", len(rows)),
        ]

    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        # ⚠ `_fcell`, NOT `csv_cell`. This path used the bare one, and
        # `routers/pulse.py` says so in a comment — "the formula guard on every
        # cell, WHERE THE TENANT /run USES BARE csv_cell". Somebody found this
        # hole, wrote it down beside the fixed copy, and left it open on the
        # tenant-facing export.
        #
        # It is reachable with ordinary data. Metric labels are per-org text: a
        # deal stage the customer renamed, a client name off a lead form, a
        # product name. `graha.pipeline_by_stage` GROUPs BY `d.stage` and
        # `graha.client_concentration` by client name — a client called
        # `=HYPERLINK("http://…","click")` becomes a live formula in the
        # exported file, on the desk of whoever opens it.
        for k, v in _preamble():
            w.writerow([k, _fcell(v)])
        w.writerow([])
        headers = list(rows[0].keys()) if rows else ["value"]
        w.writerow(headers)
        for r in rows:
            w.writerow([_fcell(r.get(h)) for h in headers])
        return Response(
            content=buf.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{stem}.csv"'},
        )

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        headers = list(rows[0].keys()) if rows else ["value"]
        ws.append([m.label])
        ws["A1"].font = Font(bold=True, size=14)
        period = (f"{win.start.isoformat()} to {win.end.isoformat()}" if win
                  else f"as at {date.today().isoformat()}")
        ws.append([period])
        ws.append([])
        ws.append(headers)
        for cell in ws[4]:
            cell.font = Font(bold=True)
        for r in rows:
            # openpyxl writes an `=`-leading string as a live FORMULA cell, so
            # xlsx needs the guard as much as csv does — see the csv branch.
            ws.append([_fcell(r.get(h)) for h in headers])
        out = io.BytesIO()
        wb.save(out)
        return Response(
            content=out.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'},
        )

    # pdf — the branded page: org name and logo, title, exact window,
    # generation timestamp; NO GSTIN, NO address, NO phone (owner's ruling,
    # 17 Aug 2026 — this is a working document, not a tax document).
    from services import doc_render as R
    from services.gst_period import load_org
    from services.report_render import analytics_letterhead

    org = await load_org(pool, org_id)
    period_line = (f"{win.start.strftime('%d %b %Y')} – {win.end.strftime('%d %b %Y')}"
                   if win else f"As at {date.today().strftime('%d %b %Y')}")

    def _build_and_render() -> bytes:
        # The WHOLE page build runs off the event loop, not just WeasyPrint:
        # `analytics_letterhead` → `letterhead` → `embed_logo` performs a
        # BLOCKING httpx.get for the org's R2-signed logo (up to 4 MB), and a
        # slow fetch on the loop stalls every concurrent request the worker
        # holds. Review finding, 2026-08-17.
        head = analytics_letterhead(org, m.label, "", period_line)

        def _is_num(v) -> bool:
            return isinstance(v, (int, float)) and not isinstance(v, bool)

        headers = list(rows[0].keys()) if rows else []
        body_rows = [
            "<tr>" + "".join(
                f'<td class="{"num" if _is_num(csv_cell(r.get(h))) else ""}">'
                f"{R.esc(str(csv_cell(r.get(h))))}</td>"
                for h in headers
            ) + "</tr>"
            for r in rows
        ]
        data_table = R.table(
            [(h, "num" if rows and _is_num(csv_cell(rows[0].get(h))) else "", "")
             for h in headers],
            body_rows,
        ) if rows else "<p>No rows for this period.</p>"

        generated = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
        page = "".join([
            head,
            data_table,
            R.foot(f"Generated {R.esc(generated)} &middot; Prepared in Kartavya"),
        ])
        html_doc = R.document(
            [page], org, title=f"{m.label} — Kartavaya",
            running=R.running_id(m.label, org, period_line),
        )
        return R.render_pdf(html_doc)

    return Response(
        content=await asyncio.to_thread(_build_and_render),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{stem}.pdf"'},
    )


# ── Saved views (proposal 62 D3) ─────────────────────────────────────────────
#
# One table for every analytics surface (staging.analytics_views, migration
# 149) and one resolution order, stated once:
#
#     personal (user_id = viewer)  >  org (user_id IS NULL)  >  code preset
#
# The module tabs and the Dristi cross-module surface both read these routes;
# `module='dristi'` is the cross-module surface's name. Layouts are validated
# ON SAVE against the registry — a rule the builder cannot express must be
# unwritable (the same promise validate_steps makes for Niyam) — and rows are
# STILL not trusted at render time, because a metric can be retired after a
# view named it; the frontend renders an unknown key as an absent widget.

#: A view is a working screen, not a data warehouse. Thirty widgets is
#: already past what a person reads; past it is a runaway client.
MAX_WIDGETS = 30

#: The cross-module surface's module name. Not in the registry — it is a
#: SURFACE, not an entitlement; reaching it is gated by the dristi module the
#: same way the dristi router gates itself.
CROSS_MODULE = "dristi"


class ViewCreate(BaseModel):
    module: str
    name: str = Field(min_length=1, max_length=80)
    layout: list
    scope: str = "personal"          # personal | org
    is_default: bool = False


class ViewUpdate(BaseModel):
    name: str | None = Field(default=None, min_length=1, max_length=80)
    layout: list | None = None
    is_default: bool | None = None


def _int_in(v, lo: int, hi: int) -> bool:
    """A real int in [lo, hi] — Python's bool passes isinstance(int); a grid
    coordinate must not."""
    return isinstance(v, int) and not isinstance(v, bool) and lo <= v <= hi


#: Keys that mean something on a METRIC widget and nothing on a row-level
#: section. A section is a table by construction — it has no viz to choose, no
#: registry dimension to group by, and `render_report_html` prints the columns
#: its ReportDef returned — so accepting these silently would save a
#: preference the renderer then ignores, which reads to the person who set it
#: as the product forgetting what they asked for.
WIDGET_ONLY_KEYS = ("viz", "group_by", "columns")

#: The geometry a free arrangement (proposal 67) may carry, and its clamps.
#: Shared by both kinds of entry: a section sits on the same 12-column grid a
#: widget does. y's ceiling is a runaway-client clamp, not a layout rule; the
#: print spine (services/module_report) reads none of the three.
GEOMETRY = (("x", 0, 11), ("y", 0, 999), ("h", 1, 8))


def _geometry_into(i: int, src: dict, item: dict, width: int) -> None:
    """x/y/h ride along ONLY when sent, so a legacy entry (w 1–3, no
    geometry) rebuilds byte-identical and a re-save rewrites nothing it did
    not touch."""
    for key, lo, hi in GEOMETRY:
        v = src.get(key)
        if v is None:
            continue
        if not _int_in(v, lo, hi):
            raise HTTPException(
                422, f"widget {i}: {key} must be an int, {lo} to {hi}")
        item[key] = v
    if "x" in item and item["x"] + width > 12:
        raise HTTPException(
            422, f"widget {i}: x+w reaches past the 12-column grid "
                 f"({item['x']}+{width} > 12)")


def _clean_width(i: int, src: dict) -> int:
    width = src.get("w", 1)
    if not _int_in(width, 1, 12):
        raise HTTPException(
            422, f"widget {i}: w must be an int, 1 to 12 grid columns")
    return width


def _clean_section(i: int, w: dict) -> dict:
    """One ROW-LEVEL section entry: `{"report": "<key>"}` plus geometry.

    A rule the builder cannot express must be unwritable — the same promise
    `_clean_layout` already makes for metric widgets, and the reason
    `{"report": …}` was rejected outright until now: `_clean_layout` read
    `metric` off every entry and 422'd on the None, so the whole row-level
    framework was unreachable from the API no matter what else was fixed.

    Validated against `REPORT_DEFS` exactly as a widget is validated against
    `REGISTRY`. NOT against entitlement: what a caller may READ is decided at
    render time by `report_section`'s gate on `ReportDef.reads`, re-evaluated
    every time the report runs, because a grant held on the day a view was
    saved is not a grant held on the day it is opened.
    """
    from services.report_defs import REPORT_DEFS, load_all

    load_all()
    key = w.get("report")
    if key not in REPORT_DEFS:
        raise HTTPException(
            422, f"widget {i}: {key!r} is not a report this product has — "
                 f"see /api/v1/analytics/report-sections")
    if w.get("metric") is not None:
        # An entry naming both is two entries wearing one hat: whichever
        # producer ran it, the other half would be silently discarded.
        raise HTTPException(
            422, f"widget {i}: an entry is a metric or a report, not both")
    for junk in WIDGET_ONLY_KEYS:
        if w.get(junk):
            raise HTTPException(
                422, f"widget {i}: `{junk}` is a metric widget's field; a "
                     f"report section always renders as its own table")
    width = _clean_width(i, w)
    item = {"report": key, "w": width}
    _geometry_into(i, w, item, width)
    return item


def _clean_layout(layout) -> list:
    """Validate and REBUILD every entry — a whitelist, so junk keys never
    reach the row. 422s name the entry index and the offence.

    An entry is a METRIC widget (`{"metric": …}`) or a ROW-LEVEL section
    (`{"report": …}`); `_is_section` is the one test that tells them apart,
    and `/module-report` resolves the saved layout with the SAME test, so a
    layout cannot save as one thing and render as another.
    """
    if not isinstance(layout, list):
        raise HTTPException(422, "layout must be a list of widgets")
    if len(layout) > MAX_WIDGETS:
        raise HTTPException(422, f"a view holds at most {MAX_WIDGETS} widgets")
    out = []
    for i, w in enumerate(layout):
        if not isinstance(w, dict):
            raise HTTPException(422, f"widget {i}: not an object")
        if _is_section(w):
            out.append(_clean_section(i, w))
            continue
        metric = w.get("metric")
        m = REGISTRY.get(metric)
        if m is None:
            raise HTTPException(
                422, f"widget {i}: {metric!r} is not a metric this product "
                     f"has — see /api/v1/analytics/catalogue")
        viz = w.get("viz", "kpi")
        if viz not in VIZ_TYPES:
            raise HTTPException(
                422, f"widget {i}: `{viz}` is not a way to draw a metric. "
                     f"Available: {', '.join(VIZ_TYPES)}")
        width = _clean_width(i, w)
        item = {"metric": metric, "viz": viz, "w": width}
        # Free arrangement (proposal 67), shared with the section branch so
        # both kinds of entry sit on one grid with one set of clamps.
        _geometry_into(i, w, item, width)
        group_by = w.get("group_by")
        if group_by:
            if group_by not in m.dimensions:
                raise HTTPException(
                    422, f"widget {i}: {metric} cannot group by {group_by!r}. "
                         f"Dimensions: {', '.join(m.dimensions) or 'none'}")
            item["group_by"] = group_by
        columns = w.get("columns")
        if columns:
            if viz != "table":
                raise HTTPException(
                    422, f"widget {i}: columns is the table chooser's field; "
                         f"this widget is a {viz}")
            item["columns"] = [str(c)[:64] for c in columns][:12]
        out.append(item)
    return out


def _row_out(r) -> dict:
    layout = r["layout"]
    if isinstance(layout, str):
        layout = json.loads(layout)
    return {
        "id": str(r["id"]),
        "scope": "org" if r["user_id"] is None else "personal",
        "name": r["name"],
        "layout": layout,
        "is_default": r["is_default"],
        "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
    }


def _widget_module(w: dict) -> str | None:
    """Which module a preset widget belongs to.

    ── WHY THIS EXISTS: GET /v1/analytics/views ANSWERED 500 FOR EVERY MODULE ──
    Found 2026-08-30 by Suite 12. `_presets_for` read `REGISTRY[w["metric"]]`
    unconditionally, and a preset layout holds TWO shapes of widget:

        {"metric": "ganit.revenue_this_month"}     a METRIC widget
        {"report": "ganit.member_activity"}        a REPORT widget — NO "metric"

    Three of the shipped presets carry a report widget — founder, finance and
    sales_head (`analytics/presets.py:86,107,149`) — so `w["metric"]` raised
    `KeyError: 'metric'` and the endpoint 500'd. Not for one module: the loop
    walks every preset on every call, so the FIRST report widget it meets kills
    the request whatever module was asked for.

    The consequence is a chain, and it is why this was invisible for so long:
    no views bar renders, so no preset chip renders, so the alert bell — which
    lives on a preset KPI widget — never appears. Saved views and metric alerts
    are unreachable through the product entirely. `analytics_views` and
    `analytics_alerts` each hold ZERO ROWS ACROSS THE WHOLE DATABASE, for all
    time, which reads like a feature nobody uses and is a feature nobody can.

    A report widget names its module the same way a metric does — the segment
    before the first dot — so the module is read from whichever key is present
    rather than assuming the metric one. Returns None for a widget carrying
    neither, which then matches no module and is dropped: an unknown widget
    shape must not take the endpoint down with it a second time.
    """
    if "metric" in w:
        entry = REGISTRY.get(w["metric"])
        if entry is None:
            # ⚠ DROPPED, BUT NEVER SILENTLY. Swapping the old `REGISTRY[...]`
            # for a plain `.get()` would trade a 500 for this codebase's
            # dominant bug class: a preset quietly losing a widget, on a screen
            # whose whole job is to be believed. A preset naming a metric that
            # is not registered is a REAL defect — a rename that missed a
            # caller — and it must stay findable in the log even though it no
            # longer takes the endpoint down with it.
            log.warning(
                "analytics preset references an unregistered metric %r; the widget "
                "is dropped from the layout. This is a defect in PRESETS or a "
                "metric that failed to register, not a caller error.",
                w["metric"],
            )
            return None
        return entry.module
    ref = w.get("report")
    if ref:
        return str(ref).split(".", 1)[0]
    log.warning("analytics preset widget carries neither 'metric' nor 'report': %r", w)
    return None


def _presets_for(module: str, reachable: set) -> list:
    """Presets, cut to what THIS caller may see.

    On a module tab, a preset contributes only its widgets for that module;
    on the cross-module surface, only widgets whose module is reachable. A
    preset with nothing left after the cut is omitted, not shown as a husk.
    """
    out = []
    for key, p in PRESETS.items():
        if module == CROSS_MODULE:
            layout = [w for w in p["layout"]
                      if _widget_module(w) in reachable]
        else:
            layout = [w for w in p["layout"]
                      if _widget_module(w) == module]
        if layout:
            out.append({"key": key, "label": p["label"], "hi": p.get("hi", ""),
                        "why": p["why"], "layout": layout})
    return out


async def _module_reachable_or_403(pool, user_id: str, org_id: str, module: str):
    if module == CROSS_MODULE:
        # The cross-module surface is itself the dristi module's screen.
        if await held_level(pool, user_id, org_id, "dristi") is None:
            raise HTTPException(403, "This view surface needs the Dristi module")
        return
    if module not in modules_in_registry():
        raise HTTPException(404, f"unknown module: {module!r}")
    if module not in UNGATED_MODULES and             await held_level(pool, user_id, org_id, module) is None:
        raise HTTPException(403, f"You do not have access to {module}")


@router.get("/views")
async def list_views(
    module: str,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    await _module_reachable_or_403(pool, user["user_id"], org_id, module)
    reachable = await _reachable(pool, user["user_id"], org_id)
    rows = await pool.fetch(
        "SELECT id, user_id, name, layout, is_default, updated_at "
        "  FROM public.analytics_views "
        " WHERE org_id = $1::uuid AND module = $2::text AND is_active "
        "   AND (user_id IS NULL OR user_id = $3::text) "
        " ORDER BY updated_at DESC",
        org_id, module, user["user_id"])
    personal = [_row_out(r) for r in rows if r["user_id"] is not None]
    org_views = [_row_out(r) for r in rows if r["user_id"] is None]
    presets = _presets_for(module, reachable)

    # The resolution, applied server-side so every surface agrees on it.
    resolved = next((v for v in personal if v["is_default"]), None)
    source = "personal" if resolved else None
    if resolved is None:
        resolved = next((v for v in org_views if v["is_default"]), None)
        source = "org" if resolved else None
    if resolved is None and presets:
        resolved = {"name": presets[0]["label"], "layout": presets[0]["layout"]}
        source = f"preset:{presets[0]['key']}"
    return {
        "personal": personal,
        "org": org_views,
        "presets": presets,
        # source=None means "nothing saved, no preset survives the cut" — the
        # frontend falls back to its built-in arrangement and says so.
        "resolved": {"source": source, **(resolved or {"layout": []})},
    }


@router.post("/views")
async def create_view(
    body: ViewCreate,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    await _module_reachable_or_403(pool, user["user_id"], org_id, body.module)
    layout = _clean_layout(body.layout)
    if body.scope not in ("personal", "org"):
        raise HTTPException(422, "scope is personal or org")
    owner = user["user_id"]
    if body.scope == "org":
        # An org view is what everyone in the org opens: writing one is org
        # administration, same bar as the module settings screens.
        from middleware.roles import admin_org_id
        if not await admin_org_id(user["user_id"], org_id):
            raise HTTPException(403, "Only an org admin can save an org-wide view")
        owner = None
    if body.is_default:
        await pool.execute(
            "UPDATE public.analytics_views SET is_default = FALSE "
            " WHERE org_id = $1::uuid AND module = $2::text "
            "   AND user_id IS NOT DISTINCT FROM $3::text",
            org_id, body.module, owner)
    row = await pool.fetchrow(
        "INSERT INTO public.analytics_views "
        "    (org_id, user_id, module, name, layout, is_default, created_by) "
        "VALUES ($1::uuid, $2::text, $3::text, $4::text, $5::jsonb, $6, $7::text) "
        "RETURNING id, user_id, name, layout, is_default, updated_at",
        org_id, owner, body.module, body.name, json.dumps(layout),
        body.is_default, user["user_id"])
    return _row_out(row)


async def _owned_view_or_404(pool, view_id: str, org_id: str, user) -> dict:
    """The row, if this caller may WRITE it: their own personal view, or an
    org view when they administer the org. Anyone else gets the same 404 a
    nonexistent id gets — a view's existence is not theirs to probe."""
    row = await pool.fetchrow(
        "SELECT id, user_id, module, name, layout, is_default, updated_at "
        "  FROM public.analytics_views "
        " WHERE id = $1::uuid AND org_id = $2::uuid AND is_active",
        view_id, org_id)
    if row is None:
        raise HTTPException(404, "View not found")
    if row["user_id"] is None:
        from middleware.roles import admin_org_id
        if not await admin_org_id(user["user_id"], org_id):
            raise HTTPException(404, "View not found")
    elif row["user_id"] != user["user_id"]:
        raise HTTPException(404, "View not found")
    return row


@router.patch("/views/{view_id}")
async def update_view(
    view_id: str,
    body: ViewUpdate,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    row = await _owned_view_or_404(pool, view_id, org_id, user)
    updates, vals = [], []
    if body.name is not None:
        vals.append(body.name)
        updates.append(f"name = ${len(vals)}::text")
    if body.layout is not None:
        vals.append(json.dumps(_clean_layout(body.layout)))
        updates.append(f"layout = ${len(vals)}::jsonb")
    if body.is_default is not None:
        if body.is_default:
            await pool.execute(
                "UPDATE public.analytics_views SET is_default = FALSE "
                " WHERE org_id = $1::uuid AND module = $2::text "
                "   AND user_id IS NOT DISTINCT FROM $3::text",
                org_id, row["module"], row["user_id"])
        vals.append(body.is_default)
        updates.append(f"is_default = ${len(vals)}")
    if not updates:
        raise HTTPException(400, "Nothing to update")
    updates.append("updated_at = NOW()")
    vals += [view_id, org_id]
    out = await pool.fetchrow(
        f"UPDATE public.analytics_views SET {', '.join(updates)} "
        f" WHERE id = ${len(vals) - 1}::uuid AND org_id = ${len(vals)}::uuid "
        f"RETURNING id, user_id, name, layout, is_default, updated_at",
        *vals)
    return _row_out(out)


@router.delete("/views/{view_id}")
async def delete_view(
    view_id: str,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    await _owned_view_or_404(pool, view_id, org_id, user)
    await pool.execute(
        "UPDATE public.analytics_views SET is_active = FALSE, updated_at = NOW() "
        " WHERE id = $1::uuid AND org_id = $2::uuid",
        view_id, org_id)
    return {"ok": True}


# ── Metric alerts (proposal 62 D7) ───────────────────────────────────────────
#
# The row decides WHEN (which metric, which line); the shipped Niyam template
# decides WHO HEARS. Evaluation happens in the Niyam sweep and runs the
# metric's own registry SQL, so the alert and the dashboard can never
# disagree about what DSO is. Managing alerts is org administration — the
# events they raise go to the org's admins.


class AlertCreate(BaseModel):
    metric: str
    operator: str
    threshold: float
    window_days: int = 30


async def _admin_or_403(user, org_id):
    from middleware.roles import admin_org_id
    if not await admin_org_id(user["user_id"], org_id):
        raise HTTPException(403, "Alerts are managed by an org admin")


@router.get("/alerts")
async def list_alerts(
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    await _admin_or_403(user, org_id)
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT id, metric, operator, threshold, window_days, created_at "
        "  FROM public.analytics_alerts "
        " WHERE org_id = $1::uuid AND is_active ORDER BY created_at",
        org_id)
    out = []
    for r in rows:
        m = REGISTRY.get(r["metric"])
        out.append({
            "id": str(r["id"]),
            "metric": r["metric"],
            # The label rides along so the screen never renders a bare key —
            # and a retired metric says so instead of vanishing silently.
            "label": m.label if m else f"{r['metric']} (no longer measured)",
            "unit": m.unit if m else None,
            "operator": r["operator"],
            "threshold": r["threshold"],
            "window_days": r["window_days"],
        })
    return {"alerts": out}


@router.post("/alerts")
async def create_alert(
    body: AlertCreate,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    await _admin_or_403(user, org_id)
    pool = await get_pool()
    m = REGISTRY.get(body.metric)
    if m is None:
        raise HTTPException(
            422, f"{body.metric!r} is not a metric this product has — see "
                 f"/api/v1/analytics/catalogue")
    if m.absent:
        # An alert on an unmeasurable metric would sit silent for ever and
        # read as "everything is fine" — the exact lie a stated absence
        # exists to prevent.
        raise HTTPException(422, {"metric": m.key, "absent": m.absent})
    if m.module not in UNGATED_MODULES and             await held_level(pool, user["user_id"], org_id, m.module) is None:
        raise HTTPException(403, f"You do not have access to {m.module}")
    if body.operator not in ("gt", "lt"):
        raise HTTPException(422, "operator is gt or lt")
    if not (1 <= body.window_days <= 366):
        raise HTTPException(422, "window_days is between 1 and 366")
    row = await pool.fetchrow(
        "INSERT INTO public.analytics_alerts "
        "    (org_id, metric, operator, threshold, window_days, created_by) "
        "VALUES ($1::uuid, $2::text, $3::text, $4::float8, $5::int, $6::text) "
        "RETURNING id",
        org_id, body.metric, body.operator, float(body.threshold),
        body.window_days, user["user_id"])
    return {"id": str(row["id"]), "metric": body.metric,
            "operator": body.operator, "threshold": body.threshold,
            "window_days": body.window_days}


@router.delete("/alerts/{alert_id}")
async def delete_alert(
    alert_id: str,
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    await _admin_or_403(user, org_id)
    pool = await get_pool()
    done = await pool.execute(
        "UPDATE public.analytics_alerts "
        "   SET is_active = FALSE, updated_at = NOW() "
        " WHERE id = $1::uuid AND org_id = $2::uuid AND is_active",
        alert_id, org_id)
    if done == "UPDATE 0":
        raise HTTPException(404, "Alert not found")
    return {"ok": True}


# ── The blended client report (proposal 60 A5, exports A6) ───────────────────
#
# One page per client: ad spend beside sessions beside leads, deals, invoices
# and payments. The last columns are the reason to own this backend at all —
# Supermetrics cannot show a client's pipeline next to their ad spend,
# because it does not hold the pipeline. We do.
#
# Two disciplines carried over verbatim:
#
#   · CRM money mirrors the REGISTRY's definitions character-for-character
#     (ganit.invoiced's credit-note CASE, ganit.collected's payment-date
#     basis) — tests pin the mirror, because a client page that disagrees
#     with the dashboard about what was invoiced discredits both.
#   · The external columns state their absence. No connected ad account is
#     "not connected", never ₹0 — a zero looks like an answer (62 §10).
#   · The spend column answers to prachar OR ganit (owner ruling 2026-08-18):
#     the registry homes ad spend under prachar (prachar.ad_spend — the
#     module the Meta data originates in), and graha is out of that gate.
#     A graha-only viewer gets the withheld sentence, in words.

#: The PAGE gate: the report is the CRM-beside-spend blend, so the page
#: needs a CRM side. prachar buys the spend COLUMN below, never the page.
CLIENT_REPORT_PAGE_MODULES = frozenset({"graha", "ganit"})

#: The spend column's entitlement (owner ruling 2026-08-18): prachar — the
#: registry's home for ad spend — or ganit. Not graha. One set, read by the
#: gate probe and the spine loop both, so the two can never disagree.
SPEND_COLUMN_MODULES = frozenset({"prachar", "ganit"})


def _client_report_sections(reachable: set) -> list[str]:
    out = []
    if "graha" in reachable:
        out += ["leads", "deals"]
    if "ganit" in reachable:
        out += ["invoices"]
    # spine columns, always listed: each renders a number, a per-client
    # connection absence, or (ads, without SPEND_COLUMN_MODULES) a withheld
    # sentence — never silently dropped
    out += ["ads", "sessions"]
    return out


@router.get("/client-report")
async def client_report(
    request: Request,
    client_id: str,
    date_from: str = "",
    date_to: str = "",
    format: str = "json",
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    pool = await get_pool()
    if format not in FORMATS:
        raise HTTPException(400, f"unknown format: {format!r} — one of {', '.join(FORMATS)}")
    if not date_from or not date_to:
        raise HTTPException(400, "date_from and date_to are required — the report is a period")
    win = aw.parse(date_from, date_to)
    try:
        uuid.UUID(client_id)
    except ValueError:
        # asyncpg would raise on the ::uuid cast and the catch-all would turn
        # a malformed probe into a 500 plus a Sentry event.
        raise HTTPException(400, "client_id must be a uuid")

    # THE gate, not held_level: `require_module` is what /run stands behind,
    # and it does three things a reachability probe does not — refuses non-god
    # platform roles for sensitive modules (ganit), writes the
    # platform.sensitive_module_access audit row, and honours subscription
    # state. Each module is asked separately and a refusal WITHHOLDS its
    # sections rather than killing the page — /overview's rule.
    reachable: set = set()
    for module in sorted(CLIENT_REPORT_PAGE_MODULES | SPEND_COLUMN_MODULES):
        try:
            await require_module(module)(request, org_id)
            reachable.add(module)
        except HTTPException:
            pass
    if not (reachable & CLIENT_REPORT_PAGE_MODULES):
        # The page is the CRM-beside-spend blend; a caller who can read
        # neither CRM side has no page here, whatever the spine holds —
        # prachar included: it entitles the spend column, not the page.
        raise HTTPException(403, "The client report needs Graha or Ganit access")

    client = await pool.fetchrow(
        "SELECT name, created_at FROM public.graha_clients "
        " WHERE id = $1::uuid AND org_id = $2::uuid",
        client_id, org_id)
    if client is None:
        raise HTTPException(404, "No such client in this organisation")

    sections = _client_report_sections(reachable)
    out: dict = {
        "client": {"name": client["name"],
                   "since": client["created_at"].date().isoformat()
                   if client["created_at"] else None},
        "window": win.as_dict(),
        "as_of": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "sections": sections,
    }

    if "leads" in sections:
        # graha.contacts_added's guards VERBATIM (is_active excludes merged
        # duplicates and deletions — "the person existed once, not twice";
        # created_at::date so the window means dates, not midnights), split
        # by this page's own dimension, source.
        rows = await pool.fetch(
            "SELECT COALESCE(NULLIF(TRIM(source), ''), 'No source') AS source, "
            "       COUNT(*)::int AS value "
            "  FROM public.graha_contacts "
            " WHERE client_id = $1::uuid AND org_id = $2::uuid "
            "   AND is_active = TRUE "
            "   AND created_at::date BETWEEN $3::date AND $4::date "
            " GROUP BY 1 HAVING COUNT(*) > 0 ORDER BY value DESC, source",
            client_id, org_id, win.start, win.end)
        out["leads"] = {"total": sum(r["value"] for r in rows),
                        "by_source": [dict(r) for r in rows]}

    if "deals" in sections:
        # The registry's won-value guards VERBATIM (avg_deal_size,
        # won_value_by_month): is_active, won_at::date — a deal won at 14:00
        # on the window's last day is IN the window, and a soft-deleted deal
        # is not a win anywhere else in the product.
        won = await pool.fetchrow(
            "SELECT COUNT(*)::int AS won_count, "
            "       COALESCE(SUM(value), 0)::float AS won_value "
            "  FROM public.graha_deals "
            " WHERE client_id = $1::uuid AND org_id = $2::uuid "
            "   AND is_active = TRUE AND won_at IS NOT NULL "
            "   AND won_at::date BETWEEN $3::date AND $4::date",
            client_id, org_id, win.start, win.end)
        # Open pipeline is a STOCK — as at today, window not applied; the
        # registry's board rule verbatim (pipeline_by_stage: active only,
        # archived out, undecided only).
        pipeline = await pool.fetchval(
            "SELECT COALESCE(SUM(value), 0)::float "
            "  FROM public.graha_deals "
            " WHERE client_id = $1::uuid AND org_id = $2::uuid "
            "   AND is_active = TRUE "
            "   AND won_at IS NULL AND lost_at IS NULL AND archived_at IS NULL",
            client_id, org_id)
        out["deals"] = {"won_count": won["won_count"],
                        "won_value": won["won_value"],
                        "open_pipeline_value": pipeline}

    if "invoices" in sections:
        inv = await pool.fetchrow(
            # ganit.invoiced's definition VERBATIM, narrowed to the client:
            # credit notes are positive rows subtracted in the CASE, drafts
            # excluded, invoice_date is the flow date. Pinned by test against
            # the registry builder so drift fails the suite.
            "SELECT COALESCE(SUM(CASE WHEN invoice_type = 'credit_note' "
            "                         THEN -total ELSE total END), 0)::float AS invoiced, "
            "       COUNT(*)::int AS invoice_count "
            "  FROM public.ganit_invoices "
            " WHERE client_id = $1::uuid AND org_id = $2::uuid "
            "   AND is_active = TRUE AND doc_status <> 'draft' "
            "   AND invoice_date BETWEEN $3::date AND $4::date",
            client_id, org_id, win.start, win.end)
        collected = await pool.fetchval(
            # ganit.collected's basis VERBATIM (payment_date), reached
            # through the invoice join because payments carry no client.
            # `i.org_id` is belt-and-braces on the joined row — fail-closed
            # beats trusting a foreign key one hop away (graha.py's rule).
            "SELECT COALESCE(SUM(p.amount), 0)::float "
            "  FROM public.ganit_payments p "
            "  JOIN public.ganit_invoices i "
            "    ON i.id = p.invoice_id AND i.org_id = $2::uuid "
            " WHERE i.client_id = $1::uuid AND p.org_id = $2::uuid "
            "   AND p.payment_date BETWEEN $3::date AND $4::date",
            client_id, org_id, win.start, win.end)
        outstanding = await pool.fetchval(
            # Stock, as at now: what this client still owes, all time —
            # ganit.outstanding's guards VERBATIM (total minus paid, never
            # balance_due; credit notes out; only rows still owing), narrowed
            # to the client, so this figure and the dashboard's ageing widget
            # can never disagree about the same rows.
            "SELECT COALESCE(SUM(total - COALESCE(amount_paid, 0)), 0)::float "
            "  FROM public.ganit_invoices "
            " WHERE client_id = $1::uuid AND org_id = $2::uuid "
            "   AND is_active = TRUE AND doc_status <> 'draft' "
            "   AND invoice_type <> 'credit_note' "
            "   AND total - COALESCE(amount_paid, 0) > 0",
            client_id, org_id)
        out["invoices"] = {"invoiced": inv["invoiced"],
                          "invoice_count": inv["invoice_count"],
                          "collected": collected,
                          "outstanding": outstanding}

    # ── the spine columns: real numbers or a stated absence ──────────────────
    # Each column asks for an account of ITS OWN source. Without the source
    # filter, a client with only GA4 connected answered the ads column with
    # metric='spend' summed over a GA account — ₹0 presented as a real figure,
    # the precise lie the absence sentence exists to prevent.
    spine_accounts: dict = {}
    withheld: set = set()
    for section, metric_name, source, needs in (
            ("ads", "spend", "meta_ads", "Meta ads account"),
            ("sessions", "sessions", "ga4", "Google Analytics")):
        if section == "ads" and not (reachable & SPEND_COLUMN_MODULES):
            # Owner ruling 2026-08-18: graha is out of the spend gate. The
            # sentence lands BEFORE the account lookup — a withheld column
            # must not read whose ad account exists, and must never render
            # empty (module_report's withheld vocabulary, stated in words).
            withheld.add(section)
            out[section] = {"absent": "Withheld — ad spend needs the "
                                      "prachar or ganit module."}
            continue
        account = await pool.fetchrow(
            "SELECT id, source, name FROM public.analytics_accounts "
            " WHERE client_id = $1::uuid AND org_id = $2::uuid "
            "   AND source = $3::text AND is_active "
            " ORDER BY created_at LIMIT 1",
            client_id, org_id, source)
        if account is None:
            out[section] = {"absent": f"No {needs} is connected for this "
                                      f"client yet — the column fills in "
                                      f"the day one is."}
            continue
        spine_accounts[section] = account
        total = await pool.fetchval(
            "SELECT COALESCE(SUM(value), 0)::float "
            "  FROM public.analytics_metrics_daily "
            " WHERE account_id = $1::uuid AND org_id = $2::uuid "
            "   AND metric = $3::text AND date BETWEEN $4::date AND $5::date",
            str(account["id"]), org_id, metric_name, win.start, win.end)
        out[section] = {"total": total, "source": account["source"],
                        "account_name": account["name"]}

    # ── the monthly blend: the page's chart and the export's table ───────────
    # Seeded with every month the window and the client's lifetime share, so a
    # quiet June appears as an empty row instead of vanishing and shortening
    # the series (months_between's own reason for existing). The clamp to the
    # client's first month keeps an all-time window from prepending decades of
    # pre-history rows.
    fill_start = win.start
    if client["created_at"] is not None:
        since_month = client["created_at"].date().replace(day=1)
        if since_month > fill_start:
            fill_start = since_month
    monthly: dict[str, dict] = {
        label: {} for label in aw.months_between(
            aw.Window(fill_start, win.end), cap=240)
    } if fill_start <= win.end else {}

    def _fold(rows, key):
        for r in rows:
            slot = monthly.setdefault(str(r["period"]), {})
            slot[key] = float(r["value"] or 0)

    if "invoices" in sections:
        _fold(await pool.fetch(
            "SELECT to_char(invoice_date, 'YYYY-MM') AS period, "
            "       SUM(CASE WHEN invoice_type = 'credit_note' "
            "                THEN -total ELSE total END)::float AS value "
            "  FROM public.ganit_invoices "
            " WHERE client_id = $1::uuid AND org_id = $2::uuid "
            "   AND is_active = TRUE AND doc_status <> 'draft' "
            "   AND invoice_date BETWEEN $3::date AND $4::date "
            " GROUP BY 1 ORDER BY 1",
            client_id, org_id, win.start, win.end), "invoiced")
        _fold(await pool.fetch(
            "SELECT to_char(p.payment_date, 'YYYY-MM') AS period, "
            "       SUM(p.amount)::float AS value "
            "  FROM public.ganit_payments p "
            "  JOIN public.ganit_invoices i "
            "    ON i.id = p.invoice_id AND i.org_id = $2::uuid "
            " WHERE i.client_id = $1::uuid AND p.org_id = $2::uuid "
            "   AND p.payment_date BETWEEN $3::date AND $4::date "
            " GROUP BY 1 ORDER BY 1",
            client_id, org_id, win.start, win.end), "collected")
    if "ads" in spine_accounts:
        # The SAME single account the headline summed — two Meta accounts must
        # not make the summary figure and its own month column disagree.
        _fold(await pool.fetch(
            "SELECT to_char(date, 'YYYY-MM') AS period, "
            "       SUM(value)::float AS value "
            "  FROM public.analytics_metrics_daily "
            " WHERE account_id = $1::uuid AND org_id = $2::uuid "
            "   AND metric = 'spend' AND date BETWEEN $3::date AND $4::date "
            " GROUP BY 1 ORDER BY 1",
            str(spine_accounts["ads"]["id"]), org_id, win.start, win.end), "spend")
    out["monthly"] = [{"period": k, **v} for k, v in sorted(monthly.items())]

    if format == "json":
        return out

    # ── files (A6): same query, three renderings, the org's paper ────────────
    # ASCII only — `isalnum()` is Unicode-alnum, and Starlette encodes headers
    # latin-1, so a Devanagari client name in Content-Disposition was a 500 on
    # every export. A fully non-ASCII name falls back to "client"; the name
    # itself still leads the file's first row in every format.
    safe_name = re.sub(r"[^A-Za-z0-9 _-]", "", client["name"]).strip().replace(" ", "-")[:40] or "client"
    stem = f"client-report_{safe_name}_{win.start.isoformat()}_{win.end.isoformat()}"

    def fcell(v):
        """csv_cell plus the formula-injection guard: a client named
        `=HYPERLINK(...)` must open in Excel as text, not execute. Numbers
        pass through as numbers; only strings can start a formula."""
        v = csv_cell(v)
        if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
            return "'" + v
        return v
    headers = ["period", "invoiced", "collected", "spend"]
    table_rows = [[m.get("period", ""), m.get("invoiced", ""),
                   m.get("collected", ""), m.get("spend", "")]
                  for m in out["monthly"]]
    summary_pairs = []
    if "leads" in out:
        summary_pairs.append(("Leads added", out["leads"]["total"]))
    if "deals" in out:
        summary_pairs += [("Deals won", out["deals"]["won_count"]),
                          ("Won value", out["deals"]["won_value"]),
                          ("Open pipeline", out["deals"]["open_pipeline_value"])]
    if "invoices" in out:
        summary_pairs += [("Invoiced", out["invoices"]["invoiced"]),
                          ("Collected", out["invoices"]["collected"]),
                          ("Outstanding", out["invoices"]["outstanding"])]
    for section, label in (("ads", "Ad spend"), ("sessions", "Sessions")):
        block = out.get(section) or {}
        # "withheld" ≠ "not connected": the file must not claim an account
        # is missing when the column was refused for entitlement.
        summary_pairs.append((label, block["total"] if "total" in block
                              else "withheld" if section in withheld
                              else "not connected"))

    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["Client", fcell(client["name"])])
        w.writerow(["Period", f"{win.start.isoformat()} to {win.end.isoformat()}"])
        w.writerow([])
        for k, v in summary_pairs:
            w.writerow([k, fcell(v)])
        w.writerow([])
        w.writerow(headers)
        for r in table_rows:
            w.writerow([fcell(c) for c in r])
        return Response(
            content=buf.getvalue(), media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{stem}.csv"'})

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Client report"
        # fcell here too: openpyxl writes an `=`-leading string as a FORMULA
        # cell, so xlsx is as injectable as csv without the guard.
        ws.append([f"Client report — {client['name']}"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"{win.start.isoformat()} to {win.end.isoformat()}"])
        ws.append([])
        for k, v in summary_pairs:
            ws.append([k, fcell(v)])
        ws.append([])
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for r in table_rows:
            ws.append([fcell(c) for c in r])
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'})

    # pdf — the same branded page /run's pdf uses: org identity block, no
    # GSTIN/address (a working document), Kartavya colophon in the tail.
    # THIS body is the client report's own — summary pairs + the monthly
    # table. The extraction to services/module_report once swapped it with
    # the module report's widget loop, which left `widgets` unbound here and
    # 500'd every pdf on BOTH routes; test_analytics_pdf_branches pins each
    # branch to its own content now.
    from services import doc_render as R
    from services.gst_period import load_org
    from services.report_render import analytics_letterhead

    org = await load_org(pool, org_id)
    period_line = f"{win.start.strftime('%d %b %Y')} — {win.end.strftime('%d %b %Y')}"

    def _build() -> bytes:
        head = analytics_letterhead(
            org, title_en=f"Client report — {client['name']}",
            title_hi="ग्राहक विवरण", period_line=period_line)
        summary_html = R.table(
            [("", "", ""), ("", "num", "")],
            [f"<tr><td>{R.esc(str(k))}</td>"
             f'<td class="num">{R.esc(str(csv_cell(v)))}</td></tr>'
             for k, v in summary_pairs])
        monthly_html = R.table(
            [(h, "num" if h != "period" else "", "") for h in headers],
            ["<tr>" + "".join(
                f'<td class="{"num" if h != "period" else ""}">'
                f"{R.esc(str(csv_cell(m.get(h, '')))) }</td>"
                for h in headers) + "</tr>"
             for m in out["monthly"]],
        ) if out["monthly"] else "<p>No monthly activity in this period.</p>"
        generated = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
        page = "".join([
            head, summary_html, "<h3>Month by month</h3>", monthly_html,
            R.foot(f"Generated {R.esc(generated)} &middot; Prepared in Kartavya"),
        ])
        html_doc = R.document(
            [page], org, title=f"Client report — {client['name']}",
            running=R.running_id(f"Client report — {client['name']}", org, period_line))
        return R.render_pdf(html_doc)

    return Response(
        content=await asyncio.to_thread(_build),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{stem}.pdf"'})


# ── The module report (proposal 65 S2) ───────────────────────────────────────
#
# "One file per module page, same queries as the screen." The report resolves
# the ARRANGEMENT the module's analytics tab would show THIS caller — the same
# precedence /views serves the frontend (personal default > org default >
# preset), finished with the registry-derived default the frontend's
# autoLayout draws when nothing is saved — and runs every widget through the
# metric's own registry builder, exactly as /run does. No SQL is written
# here: every number comes from the builder, org-scoped by its own binds, so
# the file and the screen can never disagree about what a figure means.

# The arrangement resolution, widget runner and letterhead document moved to
# services/module_report.py the day report.send needed them too (65 S4): the
# engine must not import a router, and a second copy would let the emailed
# report disagree with the downloaded one. Names re-imported so this router's
# vocabulary (and its tests' monkeypatch seams) stay put — `report_widget`
# among them, even though the loop below now goes through `report_entry`,
# which dispatches to it.
#
# `is_section` is imported rather than re-implemented for a harder reason:
# `_clean_layout` (far above — Python resolves the name at call time, and the
# module is fully loaded before any request runs) decides what a saved layout
# may CONTAIN, and `report_entry` decides what it MEANS. Two copies of that
# test is how a layout saves as one thing and renders as another.
from services.module_report import (
    DERIVED_WIDGET_CAP, MODULE_TITLES,
    is_section as _is_section,
    module_arrangement as _module_arrangement,
    render_report_html as _render_report_html,
    report_entry as _svc_report_entry,
    report_widget as _svc_report_widget,          # noqa: F401 — re-export
)


def _entry_key(entry) -> str | None:
    """The IDENTITY of one rendered entry — a widget's metric key, or a
    section's report key.

    `report_widget` returns `{metric, label, …}` and `report_section` returns
    `{report, label, …}`; neither fakes the other's key, so anything that
    wants "which thing is this" must ask for both. The window lists below are
    display keys — the frontend prints them beside "for this period" and never
    looks them up in the registry — so a report key is at home there, and
    listing only the metrics would quietly claim a windowed register was not
    windowed.

    Returns None only for a shape that carries neither, which is a producer
    this router does not have; callers filter it out rather than trust it.
    """
    if not isinstance(entry, dict):
        return None
    return entry.get("metric") or entry.get("report")


def _rendered_keys(entries: list, grain: str) -> list:
    """The identity of every entry that actually RENDERED rows at `grain`.

    Two filters, and both matter. `"data" in w` skips the stated absences —
    a withheld or retired entry had no window applied to it because it was
    never run. `_entry_key(...) is not None` drops a shape carrying neither
    identity key rather than letting a bare subscript raise inside the
    response builder, where the failure costs the whole report in every
    format.
    """
    out = []
    for w in entries:
        if "data" not in w or w.get("grain") != grain:
            continue
        key = _entry_key(w)
        if key is not None:
            out.append(key)
    return out


def _fcell(v):
    """The client report's formula guard, at module level for the module
    report: csv_cell plus the `=+-@` neutraliser. A label like
    `=HYPERLINK(...)` must open in Excel as text, not execute — and openpyxl
    writes an `=`-leading string as a live formula cell, so xlsx needs the
    guard as much as csv does."""
    v = csv_cell(v)
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def _sheet_title(label, used: set) -> str:
    """A legal, unique worksheet name: openpyxl refuses `[]:*?/\\` and caps
    titles at 31 characters, and Excel compares names case-insensitively —
    duplicates get ' (2)', ' (3)', … kept inside the cap."""
    base = re.sub(r"[\[\]:*?/\\]", " ", str(label)).strip()[:31] or "Sheet"
    title, n = base, 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[:31 - len(suffix)].rstrip() + suffix
        n += 1
    used.add(title.lower())
    return title


@router.get("/module-report")
async def module_report(
    request: Request,
    module: str,
    date_from: str = "",
    date_to: str = "",
    format: str = "json",
    report: str = "",
    user=Depends(require_user),
    org_id=Depends(get_org_id),
):
    """A module's arrangement — or, with `report=`, ONE named register.

    ── WHY `report=` IS A PARAMETER HERE AND NOT A SECOND ROUTE ──────────────

    Proposal 70 found that /reports is not three reporting surfaces but SIX,
    and that consolidating them needed exactly one thing: a second PRODUCER of
    the widget shape, so that a row-level register could travel down the
    renderer, the PDF engine and the three export branches that already exist.
    `services/module_report.report_section` is that producer and it shipped.

    What was still missing was a DOOR. `/report-sections` lists the registers a
    caller may see; nothing could then ask for one. The only way to render a
    section was to save it into a module's layout first, which means a
    register was reachable only by editing a view — so the catalogue named
    documents nobody could open, and the honest arithmetic in the proposal
    ("net six to six") stayed true.

    A separate route would have needed its own copy of the gate, its own
    window parsing and its own csv/xlsx/pdf branches — four more places for
    the six surfaces to drift apart, which is the thing being consolidated.
    So this is one parameter that swaps the LAYOUT and nothing else:

        layout = [{"report": key}]   instead of the saved arrangement

    Every line below this point is untouched. `report_entry` dispatches on the
    entry's shape, `report_section` re-gates on `ReportDef.reads`, and the four
    format branches discriminate on `label` + `data`/`absent` exactly as they
    already did. No new renderer, no new PDF engine, no new export code — the
    proposal's words, and now literally true.

    ── THE GATE IS NOT WEAKENED BY THIS ─────────────────────────────────────

    `require_module(module)` still runs first, on the module named in the URL.
    Then `report_section` refuses the section unless EVERY code in its `reads`
    passes `_gate` — which is `require_module` again, per module. A section
    whose `module` differs from the URL's therefore has to pass BOTH, and the
    one it would skip (its own, via `report_section`'s "the page's own module"
    shortcut) is precisely the one the URL gate has already asked about. That
    is checked below rather than reasoned about: a `report` whose owning module
    is not the `module` in the URL is refused outright, so the shortcut can
    never be the way past a grant.
    """
    if format not in FORMATS:
        raise HTTPException(400, f"unknown format: {format!r} — one of {', '.join(FORMATS)}")
    if not date_from or not date_to:
        raise HTTPException(400, "date_from and date_to are required — the report is a period")
    win = aw.parse(date_from, date_to)     # malformed, inverted or over the 5-year cap → 400
    if module not in modules_in_registry():
        raise HTTPException(404, f"unknown module: {module!r} — see /api/v1/analytics/catalogue")

    # ── The named register, resolved BEFORE the gate runs. ────────────────
    # A 404 for an unknown key is not a leak: `REPORT_DEFS` is the product's
    # own vocabulary, the same way `REGISTRY` is for /run, and /report-sections
    # already tells a caller which keys exist for them. What would be a leak is
    # letting the key choose the module — hence the equality check.
    section_def = None
    if report:
        from services.report_defs import REPORT_DEFS, load_all
        load_all()
        section_def = REPORT_DEFS.get(report)
        if section_def is None:
            raise HTTPException(
                404, f"unknown report: {report!r} — see "
                     f"/api/v1/analytics/report-sections")
        if section_def.module != module:
            # The register names its own owner. Asking for `manav.…` under
            # `module=core` would reach `report_section`'s "the page's own
            # module" shortcut with the WRONG module and skip the one gate
            # that matters.
            raise HTTPException(
                400, f"report {report!r} belongs to the {section_def.module} "
                     f"module, not {module!r}")

    # THE gate, /run's own: subscription state, the sensitive-module audit
    # row, platform-role refusal. Core PM is the one ungated surface —
    # membership, already proven by get_org_id, is its whole entitlement.
    if module not in UNGATED_MODULES:
        await require_module(module)(request, org_id)

    pool = await get_pool()
    if section_def is not None:
        # ONE entry, and it is a section. `source` says so in the payload so a
        # UI (and a support reader looking at a saved file) can tell a register
        # apart from a module page that happens to contain one.
        layout, source = [{"report": report}], f"section:{report}"
        label = section_def.label
    else:
        layout, source = await _module_arrangement(pool, user["user_id"], org_id, module)
        label = MODULE_TITLES.get(module, module.title())

    # The gate the service seam expects: /run's own door, asked once per
    # foreign module a saved view drags in; a refusal WITHHOLDS the widget.
    async def _gate(code: str) -> bool:
        if code in UNGATED_MODULES:
            return True
        try:
            await require_module(code)(request, org_id)
            return True
        except HTTPException:
            return False

    # One list, two producers, ONE dispatcher. A `{"report": …}` entry runs its
    # ReportDef and comes back in the SAME `{label, data}`/`{label, absent}`
    # shape a widget does — which is why the csv, xlsx and pdf branches below,
    # and `render_report_html`, needed no change at all to print a row-level
    # register. The dispatch lives in the service (`report_entry`) rather than
    # here, because two other doors walk a saved layout too (dristi's
    # scheduled run-now and Niyam's report.send) and a dispatch written in this
    # router is a dispatch those two cannot use.
    gate_cache: dict = {}
    widgets = [await _svc_report_entry(pool, org_id, module, win, w,
                                       _gate, gate_cache)
               for w in layout]

    payload = {
        "module": module,
        "label": label,
        "source": source,
        "window": {
            **win.as_dict(),
            # `_rendered_keys`, never `w["metric"]`: a SECTION has no `metric`
            # key and satisfied both of the old filters, so the subscript
            # raised KeyError the moment a section reached a layout — and this
            # payload is built BEFORE the format branches, so that KeyError
            # took the csv, xlsx and pdf downloads with it, not just
            # `?format=json`.
            "windowed": _rendered_keys(widgets, "flow"),
            "as_at": _rendered_keys(widgets, "stock"),
        },
        "as_of": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "widgets": widgets,
    }
    if format == "json":
        return payload

    # Module codes are ASCII by construction (registry keys are the ratchet),
    # so unlike the client report the stem needs no fallback — and it carries
    # the module and the period, never an org or user id.
    # The filename says WHICH document this is. A register downloaded as
    # `module-report_manav_…` is a file nobody can identify in a downloads
    # folder six weeks later. `report` is a registry key — ASCII, dotted — so
    # only the dot needs replacing to keep the extension unambiguous.
    stem = (f"report_{report.replace('.', '_')}_"
            f"{win.start.isoformat()}_{win.end.isoformat()}" if report else
            f"module-report_{module}_{win.start.isoformat()}_{win.end.isoformat()}")
    period = f"{win.start.isoformat()} to {win.end.isoformat()}"

    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["Module", _fcell(label)])
        w.writerow(["Period", period])
        w.writerow([])
        for wd in widgets:
            w.writerow([_fcell(wd["label"])])
            if "absent" in wd:
                w.writerow([_fcell(f"Not yet measurable — {wd['absent']}")])
            else:
                headers = list(wd["data"][0].keys()) if wd["data"] else ["value"]
                w.writerow(headers)
                for r in wd["data"]:
                    w.writerow([_fcell(r.get(h)) for h in headers])
            w.writerow([])
        return Response(
            content=buf.getvalue(), media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{stem}.csv"'})

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        used: set = set()
        if not widgets:
            ws = wb.active
            ws.title = "Report"
            ws.append([f"Module report — {label}"])
            ws["A1"].font = Font(bold=True, size=14)
            ws.append([period])
        for i, wd in enumerate(widgets):
            title = _sheet_title(wd["label"], used)
            if i == 0:
                ws = wb.active
                ws.title = title
                # Identity on the first sheet — a workbook opened cold must
                # say whose numbers these are and for what period.
                ws.append([f"Module report — {label}"])
                ws["A1"].font = Font(bold=True, size=14)
                ws.append([period])
                ws.append([])
            else:
                ws = wb.create_sheet(title=title)
            ws.append([_fcell(wd["label"])])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            if "absent" in wd:
                ws.append([_fcell(f"Not yet measurable — {wd['absent']}")])
                continue
            headers = list(wd["data"][0].keys()) if wd["data"] else ["value"]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            for r in wd["data"]:
                ws.append([_fcell(r.get(h)) for h in headers])
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'})

    # pdf — the same branded page /run and the client report use: identity
    # block, no GSTIN/address (a working document), one titled table per
    # widget, colophon in the tail. Built OFF the loop — the letterhead's
    # logo fetch blocks (see /run's note).
    from services import doc_render as R
    from services.gst_period import load_org

    org = await load_org(pool, org_id)
    period_line = f"{win.start.strftime('%d %b %Y')} — {win.end.strftime('%d %b %Y')}"

    def _build() -> bytes:
        # The whole page comes from the shared renderer — the same bytes
        # report.send mails — so the download, the email and this pdf can
        # never disagree about a widget or a stated absence.
        html_doc = _render_report_html(org, label, period_line, widgets)
        return R.render_pdf(html_doc)

    return Response(
        content=await asyncio.to_thread(_build),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{stem}.pdf"'})
