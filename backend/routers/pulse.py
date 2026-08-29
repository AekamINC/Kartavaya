"""pulse.py — the Aekam-only product-usage surface (proposal 68).

Who uses Kartavaya and how much — org-level aggregates and org NAMES only,
never a member's name, email or per-person row. The metrics live in
services/pulse.py (PULSE_REGISTRY — deliberately NOT the tenant registry;
its module docstring owns that argument). This router is the door:

  · GATE — the platform console's own: `require_platform_role(*CONSOLE_ROLES)`,
    imported from routers/admin_orgs.py, the same dependency that guards the
    org console. Not require_module: Pulse has no tenant and no subscription
    state, and an org user must meet the same refusal the console gives.
  · AUDIT — every catalog fetch and every export writes a row through
    services/audit.emit, the platform-privacy rule in force: access to a
    platform-wide surface leaves a trace, the discipline the DPDP attendance
    surface already follows.
  · ENVELOPES — /run answers in the tenant /v1/analytics/run shape,
    byte-for-byte on keys, so ViewGrid renders a Pulse widget unchanged.
    Formats (csv/xlsx/pdf) mirror the tenant branches with one difference:
    the identity block is the neutral "Kartavaya — Pulse", never an org
    letterhead — this document is about the platform, not written on any
    tenant's paper.

Every metric here is platform-wide: `MetricRequest.org_id` is passed empty
and no Pulse builder reads it — there is no org to scope to, and the gate
above is the whole entitlement.
"""
import asyncio
import csv
import io
import json
from datetime import date, datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Request, Response
from pydantic import BaseModel

from analytics.registry import MetricRequest
from analytics.windowing import BUCKETS, COMPARE_MODES, compare_window
from db import get_pool
from middleware.roles import require_platform_role
# The console's own reach — ONE definition of "platform staff", admin_orgs'.
from routers.admin_orgs import CONSOLE_ROLES
# The ONE layout whitelist (w 1..12, x/y/h geometry, viz, the 12-column rim)
# and the export helpers. Imported, never copied: drift between this surface's
# validator and the tenant one is the bug the import exists to prevent.
from routers.analytics import FORMATS, _clean_layout, _fcell, _sheet_title
from services import analytics_window as aw
from services.audit import emit as audit_emit
from services.pulse import DEFAULT_LAYOUT, PULSE_REGISTRY, pulse_catalogue

router = APIRouter(prefix="/api/v1/pulse", tags=["pulse"])

#: The gate every route stands behind. tests/test_pulse.py asserts each
#: route's signature carries exactly this dependency.
_pulse_gate = require_platform_role(*CONSOLE_ROLES)

#: The neutral identity every export carries — a name for the letterhead and
#: nothing else: no logo, no address, no GSTIN. Deliberately not any org row.
_PULSE_ORG = {"name": "Kartavaya — Pulse"}

#: Audit actions, in the platform.* family the middleware already writes.
AUDIT_ACCESS = "platform.pulse_access"
AUDIT_EXPORT = "platform.pulse_export"


def _proxy_key() -> str:
    """A registered tenant key `_clean_layout` will accept — see _clean_pulse_
    layout. Resolved lazily and deterministically (first key in sorted
    order); routers.analytics ran load_all() at import, so the registry is
    never empty here."""
    from routers.analytics import REGISTRY
    return min(REGISTRY)


def _clean_pulse_layout(layout) -> list:
    """The tenant whitelist, applied to Pulse keys.

    `_clean_layout` validates metric keys against the SHARED registry, which
    Pulse metrics must stay out of (services/pulse.py's docstring: held_level
    answers 'admin' to org admins for any module code, so a shared-registry
    'pulse' entry would leak platform metrics into tenant catalogues and
    tenant alerts). So the metric KEY is checked here against
    PULSE_REGISTRY, then every widget rides through the imported validator
    under a proxy key — geometry, viz, the 12-column rim, junk-key stripping
    and the widget cap are all THE tenant rules, one copy — and the real key
    is restored on the rebuilt output.
    """
    if not isinstance(layout, list):
        raise HTTPException(422, "layout must be a list of widgets")
    proxy = _proxy_key()
    proxied = []
    for i, w in enumerate(layout):
        if not isinstance(w, dict):
            raise HTTPException(422, f"widget {i}: not an object")
        key = w.get("metric")
        if key not in PULSE_REGISTRY:
            raise HTTPException(
                422, f"widget {i}: {key!r} is not a Pulse metric — see "
                     f"/api/v1/pulse/catalog")
        if w.get("group_by"):
            # No Pulse metric declares dimensions; refused HERE because the
            # proxy's dimensions are not Pulse's to validate against.
            raise HTTPException(
                422, f"widget {i}: Pulse metrics take no group_by")
        proxied.append({**w, "metric": proxy})
    cleaned = _clean_layout(proxied)
    return [{**item, "metric": w["metric"]}
            for w, item in zip(layout, cleaned)]


@router.get("/catalog")
async def catalog(
    request: Request,
    user=Depends(_pulse_gate),
):
    metrics = pulse_catalogue()
    # The trace the platform-privacy rule requires: a platform-wide surface
    # was opened, by whom, when. Fire-and-forget, never blocks the response.
    audit_emit(
        AUDIT_ACCESS, request,
        user_id=user["user_id"],
        resource_type="pulse", resource_id="catalog",
        detail={"metrics": len(metrics)},
    )
    return {
        "metrics": metrics,
        "buckets": sorted(BUCKETS),
        "compare_modes": sorted(COMPARE_MODES),
        "formats": list(FORMATS),
    }


@router.get("/run")
async def run(
    request: Request,
    metric: str,
    date_from: str = "",
    date_to: str = "",
    bucket: str = "month",
    compare: str = "",
    format: str = "json",
    user=Depends(_pulse_gate),
):
    m = PULSE_REGISTRY.get(metric)
    if m is None:
        raise HTTPException(404, f"unknown metric: {metric!r} — see /api/v1/pulse/catalog")
    if m.absent:
        # A stated absence, never a convincing zero — the tenant /run's own
        # refusal, 422 with the reason as the payload.
        raise HTTPException(422, {"metric": m.key, "absent": m.absent})
    if format not in FORMATS:
        raise HTTPException(400, f"format must be one of: {', '.join(FORMATS)}")
    if bucket not in BUCKETS:
        raise HTTPException(400, f"bucket must be one of: {', '.join(sorted(BUCKETS))}")
    if compare and compare not in COMPARE_MODES:
        raise HTTPException(400, f"compare must be one of: {', '.join(sorted(COMPARE_MODES))}")

    win = aw.parse(date_from, date_to)
    if m.grain == "flow" and win is None:
        raise HTTPException(
            400,
            f"{m.key} measures a flow — date_from and date_to are required. "
            "Stocks (as-at-today metrics) take no window; this is not one.",
        )
    if m.grain == "stock":
        win = None

    pool = await get_pool()
    # org_id is empty by design: Pulse metrics are platform-wide and no
    # builder reads it (services/pulse.py header).
    req = MetricRequest(org_id="", window=win, bucket=bucket, group_by=None)
    sql, params = m.sql(req)
    rows = [dict(r) for r in await pool.fetch(sql, *params)]

    compared = None
    if compare and m.grain == "flow":
        cw = compare_window(win, compare)
        csql, cparams = m.sql(MetricRequest(org_id="", window=cw, bucket=bucket))
        compared = {
            "mode": compare,
            "window": cw.as_dict(),
            "data": [dict(r) for r in await pool.fetch(csql, *cparams)],
        }

    # The tenant /run envelope, key for key — ViewGrid renders this unchanged.
    payload = {
        "metric": m.key,
        "label": m.label,
        "unit": m.unit,
        "grain": m.grain,
        "group_by": None,
        "bucket": bucket if m.grain == "flow" else None,
        "window": (
            {**win.as_dict(), "windowed": [m.key], "as_at": []}
            if win is not None
            else {"as_at": date.today().isoformat(), "windowed": [], "note":
                  "stock metric — true as at today; any supplied bounds were ignored"}
        ),
        "as_of": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "data": rows,
        "compare": compared,
    }

    if format == "json":
        return payload

    # An export leaves the building — the trace the audit rule requires.
    audit_emit(
        AUDIT_EXPORT, request,
        user_id=user["user_id"],
        resource_type="pulse", resource_id=m.key,
        detail={"format": format, "rows": len(rows)},
    )

    stem = m.key.replace(".", "-") + (
        f"_{win.start.isoformat()}_{win.end.isoformat()}" if win
        else f"_as-at-{date.today().isoformat()}"
    )
    # _fcell (the formula guard) on every cell, where the tenant /run uses
    # bare csv_cell: org NAMES are tenant-typed text, and a customer named
    # `=HYPERLINK(...)` must open in Excel as text on Aekam's desks too.
    headers = list(rows[0].keys()) if rows else ["value"]

    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for r in rows:
            w.writerow([_fcell(r.get(h)) for h in headers])
        return Response(
            content=buf.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{stem}.csv"'},
        )

    period = (f"{win.start.isoformat()} to {win.end.isoformat()}" if win
              else f"as at {date.today().isoformat()}")

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append([f"Kartavaya — Pulse · {m.label}"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([period])
        ws.append([])
        ws.append(headers)
        for cell in ws[4]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([_fcell(r.get(h)) for h in headers])
        out = io.BytesIO()
        wb.save(out)
        return Response(
            content=out.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'},
        )

    # pdf — the tenant /run page on neutral paper: no load_org, no logo
    # fetch, the "Kartavaya — Pulse" identity block. Built off the loop like
    # every other WeasyPrint branch.
    period_line = (f"{win.start.strftime('%d %b %Y')} – {win.end.strftime('%d %b %Y')}"
                   if win else f"As at {date.today().strftime('%d %b %Y')}")

    def _build() -> bytes:
        from services import doc_render as R
        from services.report_render import analytics_letterhead, pdf_table

        head = analytics_letterhead(_PULSE_ORG, m.label, "", period_line)
        body = pdf_table(m.label, rows) if rows else "<p>No rows for this period.</p>"
        generated = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
        page = "".join([
            head, body,
            R.foot(f"Generated {R.esc(generated)} &middot; Kartavaya — Pulse"),
        ])
        html_doc = R.document(
            [page], _PULSE_ORG, title=f"{m.label} — Kartavaya Pulse",
            running=R.running_id(m.label, _PULSE_ORG, period_line),
        )
        return R.render_pdf(html_doc)

    return Response(
        content=await asyncio.to_thread(_build),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{stem}.pdf"'},
    )


# ── the personal board (migration 155) ───────────────────────────────────────


class PulseViewPut(BaseModel):
    layout: list


@router.get("/view")
async def get_view(
    user=Depends(_pulse_gate),
):
    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT layout, updated_at FROM public.pulse_views "
        " WHERE user_id = $1::text",
        user["user_id"])
    if row is None:
        # The floor is CODE (services/pulse.DEFAULT_LAYOUT), never a row —
        # the presets-are-code rule, so a better default reaches everyone.
        return {"source": "default", "layout": DEFAULT_LAYOUT, "updated_at": None}
    layout = row["layout"]
    if isinstance(layout, str):
        layout = json.loads(layout)
    return {
        "source": "personal",
        "layout": layout,
        "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
    }


@router.put("/view")
async def put_view(
    body: PulseViewPut,
    user=Depends(_pulse_gate),
):
    layout = _clean_pulse_layout(body.layout)
    pool = await get_pool()
    row = await pool.fetchrow(
        "INSERT INTO public.pulse_views (user_id, layout, updated_at) "
        "VALUES ($1::text, $2::jsonb, NOW()) "
        "ON CONFLICT (user_id) DO UPDATE "
        "   SET layout = EXCLUDED.layout, updated_at = NOW() "
        "RETURNING updated_at",
        user["user_id"], json.dumps(layout))
    return {
        "source": "personal",
        "layout": layout,
        "updated_at": row["updated_at"].isoformat() if row and row["updated_at"] else None,
    }


# ── the whole board as one document ──────────────────────────────────────────


async def _board_widget(pool, win, widget: dict) -> dict:
    """One widget of the board, run exactly as /run runs it — or a stated
    absence. The module_report.report_widget contract, on Pulse's registry;
    no per-widget gate, because the surface gate above is the whole
    entitlement and every metric is platform-wide."""
    key = widget.get("metric")
    m = PULSE_REGISTRY.get(key)
    base = {
        "metric": key,
        "label": m.label if m else str(key),
        "viz": widget.get("viz", "kpi"),
        "unit": m.unit if m else None,
        "grain": m.grain if m else None,
    }
    if m is None:
        return {**base, "absent": "This metric is no longer measured — the "
                                  "board names a key the registry has retired."}
    if m.absent:
        return {**base, "absent": m.absent}
    req = MetricRequest(org_id="", window=win if m.grain == "flow" else None)
    sql, params = m.sql(req)
    rows = [dict(r) for r in await pool.fetch(sql, *params)]
    return {**base, "data": rows}


@router.get("/report")
async def report(
    request: Request,
    date_from: str = "",
    date_to: str = "",
    format: str = "json",
    user=Depends(_pulse_gate),
):
    if format not in FORMATS:
        raise HTTPException(400, f"unknown format: {format!r} — one of {', '.join(FORMATS)}")
    if not date_from or not date_to:
        raise HTTPException(400, "date_from and date_to are required — the report is a period")
    win = aw.parse(date_from, date_to)

    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT layout FROM public.pulse_views WHERE user_id = $1::text",
        user["user_id"])
    if row is None:
        layout, source = DEFAULT_LAYOUT, "default"
    else:
        layout = row["layout"]
        if isinstance(layout, str):
            layout = json.loads(layout)
        source = "personal"

    widgets = [await _board_widget(pool, win, w) for w in layout]

    audit_emit(
        AUDIT_EXPORT, request,
        user_id=user["user_id"],
        resource_type="pulse", resource_id="report",
        detail={"format": format, "widgets": len(widgets)},
    )

    payload = {
        "module": "pulse",
        "label": "Pulse",
        "source": source,
        "window": {
            **win.as_dict(),
            "windowed": [w["metric"] for w in widgets
                         if "data" in w and w["grain"] == "flow"],
            "as_at": [w["metric"] for w in widgets
                      if "data" in w and w["grain"] == "stock"],
        },
        "as_of": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "widgets": widgets,
    }
    if format == "json":
        return payload

    stem = f"pulse-report_{win.start.isoformat()}_{win.end.isoformat()}"
    period = f"{win.start.isoformat()} to {win.end.isoformat()}"

    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["Report", "Kartavaya — Pulse"])
        w.writerow(["Period", period])
        w.writerow([])
        for wd in widgets:
            w.writerow([_fcell(wd["label"])])
            if "absent" in wd:
                w.writerow([_fcell(f"Not measured here — {wd['absent']}")])
            else:
                headers = list(wd["data"][0].keys()) if wd["data"] else ["value"]
                w.writerow(headers)
                for r in wd["data"]:
                    w.writerow([_fcell(r.get(h)) for h in headers])
            w.writerow([])
        return Response(
            content=buf.getvalue(), media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{stem}.csv"'})

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        used: set = set()
        if not widgets:
            ws = wb.active
            ws.title = "Pulse"
            ws.append(["Kartavaya — Pulse"])
            ws["A1"].font = Font(bold=True, size=14)
            ws.append([period])
        for i, wd in enumerate(widgets):
            title = _sheet_title(wd["label"], used)
            if i == 0:
                ws = wb.active
                ws.title = title
                ws.append(["Kartavaya — Pulse"])
                ws["A1"].font = Font(bold=True, size=14)
                ws.append([period])
                ws.append([])
            else:
                ws = wb.create_sheet(title=title)
            ws.append([_fcell(wd["label"])])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            if "absent" in wd:
                ws.append([_fcell(f"Not measured here — {wd['absent']}")])
                continue
            headers = list(wd["data"][0].keys()) if wd["data"] else ["value"]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            for r in wd["data"]:
                ws.append([_fcell(r.get(h)) for h in headers])
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'})

    # pdf — the module report's page shape on the neutral identity: one
    # titled table per widget, stated absences in words, no org anywhere.
    period_line = f"{win.start.strftime('%d %b %Y')} — {win.end.strftime('%d %b %Y')}"

    def _build() -> bytes:
        from services import doc_render as R
        from services.report_render import analytics_letterhead, pdf_table

        head = analytics_letterhead(_PULSE_ORG, "Pulse report", "", period_line)
        parts = [head]
        for wd in widgets:
            if "absent" in wd:
                parts.append(R.block(
                    str(wd["label"]),
                    f"<p>Not measured here — {R.esc(str(wd['absent']))}</p>"))
            elif wd.get("data"):
                parts.append(pdf_table(wd["label"], wd["data"]))
            else:
                parts.append(R.block(str(wd["label"]),
                                     "<p>No rows for this period.</p>"))
        generated = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
        parts.append(R.foot(f"Generated {R.esc(generated)} &middot; Kartavaya — Pulse"))
        html_doc = R.document(
            ["".join(parts)], _PULSE_ORG, title="Pulse report — Kartavaya",
            running=R.running_id("Pulse report", _PULSE_ORG, period_line))
        return R.render_pdf(html_doc)

    return Response(
        content=await asyncio.to_thread(_build),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{stem}.pdf"'})
