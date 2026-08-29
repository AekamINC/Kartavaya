"""The CRM report, as a file you can send someone.

Owner, 2026-08-09, approved with the plan in `docs/proposals/47-reports-download.html`:
all five CRM reports were computed and trapped on a screen.

── ONE DOCUMENT, NOT FIVE DOWNLOADS ─────────────────────────────────────────

A firm sends *a report*. So the five report queries are gathered into one dict
and rendered once, in whichever of three formats the caller asked for — and the
three are genuinely different documents:

  · PDF   — presentable, carries the ORGANISATION's name, address and GSTIN
            (not Kartavaya's), for sending to a client or a board.
  · Excel — one sheet per section plus a raw Deals sheet, for working with.
  · CSV   — the raw Deals sheet ONLY. A CSV cannot hold six sections, and
            pretending otherwise produces a file nothing can read.

── THE ORG HEADER IS OPTIONAL, AND THAT IS DELIBERATE ───────────────────────

Name, address, GSTIN and PAN come from `staging.organisations`, the same source
the invoices use. Any of them may be empty and the report still renders with
that line omitted: GSTIN and PAN are NON-MANDATORY in this product and block
nothing. A report that refuses to generate because a tax id is blank would be a
new rule nobody asked for.
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timedelta, timezone

log = logging.getLogger(__name__)

FORMATS = ("pdf", "excel", "csv")


def _esc(v) -> str:
    return (str(v if v is not None else "")
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _inr(v) -> str:
    """Indian grouping — 12,34,567, not 1,234,567. A rupee figure grouped in
    thousands reads as the wrong number to the person it is sent to."""
    try:
        n = int(round(float(v or 0)))
    except (TypeError, ValueError):
        return "0"
    s, neg = str(abs(n)), n < 0
    if len(s) > 3:
        head, tail = s[:-3], s[-3:]
        parts = []
        while len(head) > 2:
            parts.insert(0, head[-2:])
            head = head[:-2]
        if head:
            parts.insert(0, head)
        s = ",".join(parts) + "," + tail
    return ("-" if neg else "") + s


async def gather(pool, org_id: str, days: int, *, include_reps: bool) -> dict:
    """Every figure the report needs, in one dict.

    `include_reps` is the caller's answer to the permission question, not this
    function's: rep performance is admin-only on the screen, and a member who
    cannot see per-person numbers should get a report without that section
    rather than an error.
    """
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)

    org = await pool.fetchrow(
        "SELECT name, gstin, pan, billing_address FROM public.organisations "
        "WHERE id=$1::uuid", org_id)

    # ── TWO COHORTS, NAMED, BECAUSE THEY ARE NOT THE SAME DEALS ─────────────
    #
    # This query used to have ONE window — `created_at > cutoff` in the WHERE —
    # and it counted `won_value` inside it. So "Won value" for the last 90 days
    # meant "the value of deals CREATED in the last 90 days that happen to be
    # Won today". A deal opened in March and won last week was not in it; a deal
    # opened last week and won in two years' time would be. Measured live
    # 2026-08-22, last 90 days: E2E Rs53,13,648 shown against Rs66,37,948 won,
    # Unicode Rs11,22,500 against Rs15,72,500. On the financial year to date
    # Unicode is 39% low — Rs7,30,000 of real wins missing from a figure headed
    # "Won value".
    #
    # The fix is NOT to move the whole query onto `won_at`, because
    # `total_deals` and the conversion rate are honest questions about the
    # CREATED cohort and moving them would swap one mislabelled number for
    # another. Both cohorts are computed here and both are labelled on every
    # renderer:
    #
    #   · `total_deals` / `cohort_won` / `cohort_lost` / `conversion_rate`
    #     — deals OPENED in the window and where they stand TODAY. The
    #       conversion rate has to be built from this pair or it divides wins
    #       from one population by openings from another.
    #   · `won` / `won_value` / `lost` / `avg_cycle_days`
    #     — deals CLOSED in the window, on `won_at` / `lost_at`. This is what
    #       "won in this period" means to the person reading it.
    #
    # `won_at` is filled on 33 of 33 won deals and `lost_at` on 22 of 22 lost
    # ones (live, 2026-08-22, both real orgs and the seeded one), so neither
    # closed figure loses rows to a NULL date. `won_undated` and `lost_undated`
    # carry the exception anyway rather than letting a future NULL quietly
    # shrink a total — the same rule `done_undated` follows on the work report.
    #
    # The WHERE no longer carries `created_at > $2`: a deal won inside the
    # window may have been opened long before it, and filtering the table first
    # would put that deal beyond reach of every FILTER below.
    conversion = await pool.fetchrow(
        "SELECT COUNT(*) FILTER (WHERE created_at > $2) AS total_deals, "
        "  COUNT(*) FILTER (WHERE stage='Won'  AND created_at > $2) AS cohort_won, "
        "  COUNT(*) FILTER (WHERE stage='Lost' AND created_at > $2) AS cohort_lost, "
        "  COUNT(*) FILTER (WHERE stage='Won'  AND won_at  > $2) AS won, "
        "  COUNT(*) FILTER (WHERE stage='Lost' AND lost_at > $2) AS lost, "
        "  COALESCE(SUM(value) FILTER (WHERE stage='Won' AND won_at > $2), 0) "
        "    AS won_value, "
        "  COUNT(*) FILTER (WHERE stage='Won'  AND won_at  IS NULL) AS won_undated, "
        "  COUNT(*) FILTER (WHERE stage='Lost' AND lost_at IS NULL) AS lost_undated, "
        "  AVG(EXTRACT(EPOCH FROM (won_at - created_at))/86400)"
        "    FILTER (WHERE stage='Won' AND won_at > $2)::int AS avg_cycle_days "
        "FROM public.graha_deals WHERE org_id=$1::uuid",
        org_id, cutoff)

    forecast = await pool.fetch(
        "SELECT stage, COUNT(*) AS count, COALESCE(SUM(value),0) AS total_value, "
        "  COALESCE(SUM(value * probability / 100.0),0) AS weighted_value "
        "FROM public.graha_deals "
        "WHERE org_id=$1::uuid AND is_active=TRUE AND stage NOT IN ('Won','Lost') "
        "GROUP BY stage ORDER BY weighted_value DESC",
        org_id)

    velocity = await pool.fetch(
        "SELECT stage, COUNT(*) AS count, COALESCE(SUM(value),0) AS total_value, "
        "  AVG(EXTRACT(EPOCH FROM (updated_at - created_at))/86400)::int AS avg_days_in_stage "
        "FROM public.graha_deals "
        "WHERE org_id=$1::uuid AND is_active=TRUE AND created_at > $2 "
        "GROUP BY stage ORDER BY count DESC",
        org_id, cutoff)

    sources = await pool.fetch(
        "SELECT COALESCE(c.source, 'unknown') AS source, COUNT(*) AS leads, "
        "  COUNT(d.id) AS deals, COUNT(d.id) FILTER (WHERE d.stage='Won') AS won, "
        "  COALESCE(SUM(d.value) FILTER (WHERE d.stage='Won'), 0) AS won_value "
        "FROM public.graha_contacts c "
        "LEFT JOIN public.graha_deals d ON d.contact_id = c.id AND d.is_active=TRUE "
        "WHERE c.org_id=$1::uuid AND c.created_at > $2 AND c.is_active=TRUE "
        "GROUP BY COALESCE(c.source, 'unknown') ORDER BY leads DESC",
        org_id, cutoff)

    reps = []
    if include_reps:
        # Joined to `users` because the screen's version groups by `assigned_to`
        # and returns no name at all — a report that names people by id is not a
        # report anybody can act on.
        # Same correction as the conversion block above, and it matters more
        # here because these figures sit against a PERSON'S NAME. `Deals` is
        # the allocation — what was opened and routed to them in the window —
        # and `Won`/`Won value` are what they CLOSED in it, on `won_at`. Under
        # one created_at window a rep who closed a long deal opened last year
        # showed zero beside their own name.
        reps = await pool.fetch(
            # THE LADDER STOPS BEFORE THE EMAIL. It read `COALESCE(u.full_name,
            # u.name, u.email, 'Unassigned')`, so a rep with no name recorded
            # had their EMAIL printed as the row label of a CRM performance
            # report — and this report is exported to CSV and XLSX, so the
            # address travelled out of the product. That is a contact detail
            # rendered as a label, and it inverts the rule that Aekam must not
            # see a customer's member emails. The owner's ruling (2026-08-23)
            # is that a display-name ladder must never end at an email address.
            #
            # This agrees with the note above rather than contradicting it: a
            # report that names people by id is not actionable, and neither is
            # one that names them by inbox. Both are the identifier standing in
            # for the person; only the NAME is the label.
            #
            # MEASURED BEFORE REMOVING IT: 0 of 35 live accounts have neither
            # `full_name` nor `name`, so the email rung has never fired on real
            # data and this changes nothing anybody can see.
            #
            # `'Unassigned'` STAYS — it is this report's own terminal and says
            # more here than a generic label would, so only the email rung came
            # out. It also cannot be swapped for `display_name()` for a second
            # reason: this expression is repeated verbatim in the GROUP BY
            # below, and the two must stay character-identical or the query
            # fails at runtime. Edit both or neither.
            "SELECT COALESCE(u.full_name, u.name, 'Unassigned') AS person, "
            "  COUNT(*) FILTER (WHERE d.created_at > $2) AS total_deals, "
            "  COUNT(*) FILTER (WHERE d.stage='Won'  AND d.won_at  > $2) AS won, "
            "  COUNT(*) FILTER (WHERE d.stage='Lost' AND d.lost_at > $2) AS lost, "
            "  COALESCE(SUM(d.value) FILTER (WHERE d.stage='Won' AND d.won_at > $2), 0) "
            "    AS won_value "
            "FROM public.graha_deals d "
            "LEFT JOIN users u ON u.user_id = d.assigned_to "
            "WHERE d.org_id=$1::uuid AND d.assigned_to IS NOT NULL "
            # The other half of the pair — kept character-identical to the
            # SELECT expression above.
            "GROUP BY COALESCE(u.full_name, u.name, 'Unassigned') "
            # A person whose whole window is zeros is dropped rather than
            # printed as a line of noughts beside their name: the query now
            # spans the table, so every rep who ever held a deal would appear.
            "HAVING COUNT(*) FILTER (WHERE d.created_at > $2) > 0 "
            "    OR COUNT(*) FILTER (WHERE d.stage='Won'  AND d.won_at  > $2) > 0 "
            "    OR COUNT(*) FILTER (WHERE d.stage='Lost' AND d.lost_at > $2) > 0 "
            "ORDER BY won_value DESC",
            org_id, cutoff)

    # The raw rows. This is the whole of the CSV and the last sheet of the
    # workbook, and it is what makes the report auditable — every aggregate
    # above can be recomputed from it.
    deals = await pool.fetch(
        "SELECT d.title, d.stage, d.value, d.probability, d.created_at, "
        "  d.expected_close_date, d.won_at, d.lost_at, "
        "  COALESCE(cl.name, c.company, '') AS company, "
        "  COALESCE(c.name, '') AS contact, "
        "  COALESCE(c.source, '') AS source, "
        "  COALESCE(tr.name, '') AS territory, "
        # Keeps its own `''` terminal: this is the raw CSV/XLSX sheet, where
        # every other unresolved column beside it is an empty cell too, and a
        # sudden prose label in one column would read as data. Email rung only.
        "  COALESCE(u.full_name, u.name, '') AS owner "
        "FROM public.graha_deals d "
        "LEFT JOIN public.graha_contacts c ON c.id = d.contact_id "
        # Scoped on `org_id` as well as `id`, as the routers now are: this is
        # the sheet the whole report is auditable from, so a company name that
        # belongs to another organisation would be exported to a CSV, mailed
        # out, and re-read months later with nothing left to question it.
        "LEFT JOIN public.graha_clients cl "
        "       ON cl.id = d.client_id AND cl.org_id = d.org_id "
        # Org-scoped on `org_id` as well as `id`, like the client join
        # directly above — and this line was not, until Phase 7.1a. The
        # reasoning is the client join's, only worse here: this is the
        # sheet the whole report is auditable from, so a TERRITORY name
        # belonging to another organisation would be exported to a CSV,
        # mailed out, and re-read months later with nothing left to
        # question it. `graha_territories.id` is unique table-wide and
        # migration 023 gave the column a non-composite foreign key.
        "LEFT JOIN public.graha_territories tr "
        "       ON tr.id = d.territory_id AND tr.org_id = d.org_id "
        "LEFT JOIN users u ON u.user_id = d.assigned_to "
        "WHERE d.org_id=$1::uuid AND d.is_active=TRUE AND d.created_at > $2 "
        "ORDER BY d.created_at DESC",
        org_id, cutoff)

    conv = dict(conversion) if conversion else {}
    total = conv.get("total_deals") or 0
    # BOTH derived figures are built from the CREATED cohort, never from the
    # closed-in-period counts beside them. `open` used to subtract `won` and
    # `lost` from `total_deals`; once those two moved onto `won_at`/`lost_at`
    # that subtraction started mixing populations and could go NEGATIVE — a
    # period in which more deals were won than were opened is ordinary, and it
    # would have printed "Still open: -2".
    conv["open"] = total - (conv.get("cohort_won") or 0) - (conv.get("cohort_lost") or 0)
    conv["conversion_rate"] = (round((conv.get("cohort_won") or 0) / total * 100, 1)
                               if total else 0)

    return {
        "org": dict(org) if org else {},
        "period_days": days,
        "generated_at": datetime.now(timezone.utc),
        "conversion": conv,
        "forecast": [dict(r) for r in forecast],
        "velocity": [dict(r) for r in velocity],
        "sources": [dict(r) for r in sources],
        "reps": [dict(r) for r in reps],
        "deals": [dict(r) for r in deals],
        "includes_reps": include_reps,
    }


DEAL_COLUMNS = [
    ("title", "Deal"), ("company", "Company"), ("contact", "Contact"),
    ("stage", "Stage"), ("value", "Value"), ("probability", "Probability %"),
    ("owner", "Owner"), ("source", "Source"), ("territory", "Territory"),
    ("created_at", "Created"), ("expected_close_date", "Expected close"),
    ("won_at", "Won"), ("lost_at", "Lost"),
]


def _cell(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def to_csv(data: dict) -> bytes:
    """The deals, and only the deals. See the module docstring."""
    buf = io.StringIO(newline="")
    w = csv.writer(buf)
    w.writerow([label for _, label in DEAL_COLUMNS])
    for row in data["deals"]:
        w.writerow([_cell(row.get(key)) for key, _ in DEAL_COLUMNS])
    # BOM so Excel opens a UTF-8 CSV without mangling names.
    return buf.getvalue().encode("utf-8-sig")


def to_excel(data: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    def sheet(title, headers, rows):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, start=1):
            width = max([len(str(h))] + [len(_cell(r[i - 1])) for r in rows] or [0])
            ws.column_dimensions[get_column_letter(i)].width = min(42, max(12, width + 2))
        return ws

    wb.remove(wb.active)

    conv = data["conversion"]
    sheet("Summary", ["Figure", "Value"], [
        ["Organisation", (data["org"].get("name") or "")],
        ["Period", f"Last {data['period_days']} days"],
        ["Generated", data["generated_at"].strftime("%Y-%m-%d %H:%M UTC")],
        # Every label says WHICH cohort and over WHAT window. Two of these rows
        # count deals OPENED in the period and four count deals CLOSED in it;
        # they are different deals and the sheet has to say so, because a
        # column of six numbers under one "Period" line reads as one population.
        ["Deals opened in period", conv.get("total_deals") or 0],
        ["— of those, won so far", conv.get("cohort_won") or 0],
        ["— of those, lost so far", conv.get("cohort_lost") or 0],
        ["— of those, still open", conv.get("open") or 0],
        ["Conversion rate % (of deals opened in period)",
         conv.get("conversion_rate") or 0],
        ["Won in period (on won date)", conv.get("won") or 0],
        ["Won value in period (on won date)", float(conv.get("won_value") or 0)],
        ["Lost in period (on lost date)", conv.get("lost") or 0],
        ["Average cycle, opened to won (days)", conv.get("avg_cycle_days") or 0],
        ["Won deals with no won date (all time)", conv.get("won_undated") or 0],
        ["Lost deals with no lost date (all time)", conv.get("lost_undated") or 0],
    ])
    sheet("Forecast", ["Stage", "Deals", "Pipeline value", "Weighted value"],
          [[r["stage"], r["count"], float(r["total_value"]), float(r["weighted_value"])]
           for r in data["forecast"]])
    sheet("Velocity", ["Stage", "Deals", "Value", "Average days in stage"],
          [[r["stage"], r["count"], float(r["total_value"]), r["avg_days_in_stage"] or 0]
           for r in data["velocity"]])
    sheet("Sources", ["Source", "Leads", "Deals", "Won to date", "Won value to date"],
          [[r["source"], r["leads"], r["deals"], r["won"], float(r["won_value"])]
           for r in data["sources"]])
    if data["includes_reps"]:
        sheet("By person", ["Person", "Deals opened", "Won in period", "Lost in period",
               "Won value in period"],
              [[r["person"], r["total_deals"], r["won"], r["lost"], float(r["won_value"])]
               for r in data["reps"]])
    sheet("Deals", [label for _, label in DEAL_COLUMNS],
          [[_cell(row.get(key)) for key, _ in DEAL_COLUMNS] for row in data["deals"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _table(headers: list[str], rows: list[list], numeric_from: int = 1) -> str:
    if not rows:
        return '<p class="none">Nothing in this period.</p>'
    head = "".join(
        f'<th class="{"n" if i >= numeric_from else ""}">{_esc(h)}</th>'
        for i, h in enumerate(headers))
    body = "".join(
        "<tr>" + "".join(
            f'<td class="{"n" if i >= numeric_from else ""}">{_esc(c)}</td>'
            for i, c in enumerate(r)) + "</tr>"
        for r in rows)
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def _bars(rows: list[tuple[str, float]]) -> str:
    """A bar per row, drawn in CSS. No chart library: one is a large dependency
    for decoration, and a div with a width is a bar."""
    if not rows:
        return ""
    top = max((v for _, v in rows), default=0) or 1
    out = []
    for label, value in rows:
        pct = max(1, round(value / top * 100))
        out.append(
            f'<div class="bar"><span class="bar__l">{_esc(label)}</span>'
            f'<span class="bar__t"><i style="width:{pct}%"></i></span>'
            f'<span class="bar__v">{_inr(value)}</span></div>')
    return f'<div class="bars">{"".join(out)}</div>'


def _html(data: dict) -> str:
    org = data["org"]
    conv = data["conversion"]

    ids = []
    if org.get("gstin"):
        ids.append(f"GSTIN {_esc(org['gstin'])}")
    if org.get("pan"):
        ids.append(f"PAN {_esc(org['pan'])}")
    addr = _esc(org.get("billing_address") or "")

    # The stat strip is the most-quoted part of this document, so each caption
    # carries its own basis. "Won" and "Won value" are deals CLOSED in the
    # period (on `won_at`); "Deals opened" and "Conversion" are the deals
    # OPENED in it. Three words per caption is the whole cost of not having
    # somebody quote the wrong population in a board pack.
    stats = [
        ("Opened in period", str(conv.get("total_deals") or 0)),
        ("Won in period", str(conv.get("won") or 0)),
        ("Lost in period", str(conv.get("lost") or 0)),
        ("Conversion of opened", f"{conv.get('conversion_rate') or 0}%"),
        ("Won value in period", f"₹{_inr(conv.get('won_value'))}"),
        ("Avg cycle to win", f"{conv.get('avg_cycle_days') or 0} days"),
    ]
    stat_html = "".join(
        f'<div class="stat"><b>{_esc(v)}</b><span>{_esc(k)}</span></div>'
        for k, v in stats)

    reps_section = ""
    if data["includes_reps"]:
        reps_section = (
            "<h2>By person</h2>"
            + _table(["Person", "Deals opened", "Won in period", "Lost in period",
               "Won value in period"],
                     [[r["person"], r["total_deals"], r["won"], r["lost"],
                       "₹" + _inr(r["won_value"])] for r in data["reps"]]))

    return f"""<!doctype html>
<html><head><meta charset="utf-8"><style>
  @page {{ size: A4; margin: 16mm 14mm 18mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font: 10.5pt/1.45 "Helvetica Neue", Helvetica, Arial, sans-serif; color: #1a2230; }}
  .head {{ display: flex; justify-content: space-between; align-items: flex-start;
           border-bottom: 2px solid #1a2230; padding-bottom: 8pt; margin-bottom: 14pt; }}
  .org {{ font-size: 15pt; font-weight: 700; letter-spacing: -.2pt; }}
  .sub {{ color: #6e7b91; font-size: 8.5pt; margin-top: 2pt; white-space: pre-line; }}
  .title {{ text-align: right; }}
  .title b {{ font-size: 12pt; display: block; }}
  h2 {{ font-size: 10pt; text-transform: uppercase; letter-spacing: .08em;
        color: #6e7b91; margin: 16pt 0 5pt; border-bottom: 1px solid #e2dcc9;
        padding-bottom: 3pt; }}
  .stats {{ display: flex; flex-wrap: wrap; gap: 6pt; }}
  .stat {{ flex: 1 1 28%; border: 1px solid #e2dcc9; border-radius: 4pt; padding: 7pt 9pt; }}
  .stat b {{ display: block; font-size: 14pt; }}
  .stat span {{ font-size: 8pt; color: #6e7b91; text-transform: uppercase;
                letter-spacing: .05em; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 9.5pt; }}
  th, td {{ text-align: left; padding: 4pt 6pt; border-bottom: 1px solid #efeade; }}
  th {{ color: #6e7b91; font-size: 8pt; text-transform: uppercase; letter-spacing: .05em; }}
  td.n, th.n {{ text-align: right; }}
  .none {{ color: #6e7b91; font-style: italic; font-size: 9.5pt; }}
  .bars {{ margin-top: 6pt; }}
  .bar {{ display: flex; align-items: center; gap: 6pt; margin-bottom: 3pt; font-size: 9pt; }}
  .bar__l {{ width: 26%; color: #6e7b91; }}
  .bar__t {{ flex: 1; background: #f2eee2; height: 8pt; border-radius: 4pt; overflow: hidden; }}
  .bar__t i {{ display: block; height: 100%; background: #0082c6; }}
  .bar__v {{ width: 18%; text-align: right; }}
  .foot {{ position: fixed; bottom: -10mm; left: 0; right: 0; font-size: 8pt;
           color: #6e7b91; border-top: 1px solid #e2dcc9; padding-top: 4pt;
           display: flex; justify-content: space-between; }}
</style></head><body>
  <div class="head">
    <div>
      <div class="org">{_esc(org.get("name") or "CRM report")}</div>
      <div class="sub">{addr}{("\n" + " · ".join(ids)) if ids else ""}</div>
    </div>
    <div class="title">
      <b>CRM report</b>
      <span class="sub">Last {data["period_days"]} days<br>
      Generated {data["generated_at"].strftime("%d %b %Y")}</span>
    </div>
  </div>

  <h2>Summary</h2>
  <div class="stats">{stat_html}</div>

  <h2>Forecast — open pipeline</h2>
  {_table(["Stage", "Deals", "Pipeline value", "Weighted value"],
          [[r["stage"], r["count"], "₹" + _inr(r["total_value"]),
            "₹" + _inr(r["weighted_value"])] for r in data["forecast"]])}
  {_bars([(r["stage"], float(r["weighted_value"])) for r in data["forecast"]])}

  <h2>Pipeline velocity</h2>
  {_table(["Stage", "Deals", "Value", "Avg days in stage"],
          [[r["stage"], r["count"], "₹" + _inr(r["total_value"]),
            r["avg_days_in_stage"] or 0] for r in data["velocity"]])}

  <h2>Where the deals came from</h2>
  {_table(["Source", "Leads", "Deals", "Won to date", "Won value to date"],
          [[r["source"], r["leads"], r["deals"], r["won"],
            "₹" + _inr(r["won_value"])] for r in data["sources"]])}
  {_bars([(r["source"], float(r["won_value"])) for r in data["sources"]])}

  {reps_section}

  <div class="foot">
    <span>{_esc(org.get("name") or "")} · CRM report</span>
    <span>Generated {data["generated_at"].strftime("%d %b %Y %H:%M UTC")}</span>
  </div>
</body></html>"""


def to_pdf(data: dict) -> bytes:
    # OSError as well as ImportError: WeasyPrint imports fine and then fails to
    # load libgobject on a host without the native libraries, which is exactly
    # what a Windows dev machine looks like. Both mean "no PDF here" and both
    # must reach the caller as that, not as a 500.
    try:
        from weasyprint import HTML
        return HTML(string=_html(data), base_url=None).write_pdf()
    except (ImportError, OSError) as exc:
        raise RuntimeError(
            "PDF rendering is not available on this server — ask for the Excel "
            "or CSV format instead.") from exc


def render(data: dict, fmt: str) -> tuple[bytes, str, str]:
    """(content, media_type, extension) for one of `FORMATS`."""
    if fmt == "csv":
        return to_csv(data), "text/csv; charset=utf-8", "csv"
    if fmt == "excel":
        return (to_excel(data),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "xlsx")
    return to_pdf(data), "application/pdf", "pdf"
