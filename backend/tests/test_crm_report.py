"""The CRM report as a file. Approved 2026-08-09 with proposal 47.

Rendered against a fabricated data dict rather than the database, because what
can go wrong here is in the rendering: an unescaped company name, a rupee figure
grouped in thousands, or a PDF failure taking the whole download with it.
"""
from datetime import datetime, timezone

import pytest

from routers import graha
from services import crm_report


def _data(**over):
    d = {
        "org": {"name": "Aekam Inc", "gstin": "27AAAAA0000A1Z5", "pan": "AAAAA0000A",
                "billing_address": "12 Test Road\nMumbai 400001"},
        "period_days": 90,
        "generated_at": datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
        "conversion": {"total_deals": 10, "won": 3, "lost": 2, "open": 5,
                       "conversion_rate": 30.0, "won_value": 1234567,
                       "avg_cycle_days": 21},
        "forecast": [{"stage": "Proposal", "count": 2, "total_value": 500000,
                      "weighted_value": 250000}],
        "velocity": [{"stage": "New", "count": 4, "total_value": 100000,
                      "avg_days_in_stage": 9}],
        "sources": [{"source": "referral", "leads": 6, "deals": 4, "won": 2,
                     "won_value": 900000}],
        "reps": [{"person": "Keval Shah", "total_deals": 5, "won": 2, "lost": 1,
                  "won_value": 700000}],
        "deals": [{"title": "A & B deal", "company": "X Ltd", "contact": "Y",
                   "stage": "Won", "value": 100, "probability": 100,
                   "owner": "Keval Shah", "source": "referral", "territory": "West",
                   "created_at": datetime(2026, 7, 1, tzinfo=timezone.utc),
                   "expected_close_date": None,
                   "won_at": datetime(2026, 7, 20, tzinfo=timezone.utc),
                   "lost_at": None}],
        "includes_reps": True,
    }
    d.update(over)
    return d


def test_rupees_are_grouped_the_indian_way():
    """12,34,567 — not 1,234,567. A figure grouped in thousands reads as the
    wrong number to the person the report is sent to."""
    assert crm_report._inr(1234567) == "12,34,567"
    assert crm_report._inr(999) == "999"
    assert crm_report._inr(-100000) == "-1,00,000"
    assert crm_report._inr(None) == "0"


def test_the_html_escapes_what_the_user_typed():
    html = crm_report._html(_data(org={"name": "<script>x</script>", "gstin": "",
                                       "pan": "", "billing_address": ""}))
    assert "<script>" not in html
    assert "&lt;script&gt;" in html


def test_a_missing_gstin_omits_the_line_rather_than_refusing():
    """GSTIN and PAN are NON-MANDATORY in this product and block nothing."""
    html = crm_report._html(_data(org={"name": "No Ids Ltd", "gstin": "", "pan": "",
                                       "billing_address": ""}))
    assert "GSTIN" not in html and "PAN" not in html
    assert "No Ids Ltd" in html


def test_an_empty_period_still_renders():
    html = crm_report._html(_data(forecast=[], velocity=[], sources=[], reps=[],
                                  deals=[], includes_reps=False))
    assert "Nothing in this period." in html


def test_the_by_person_section_is_omitted_when_it_may_not_be_seen():
    """A member who cannot see per-rep numbers gets a report without that
    section, NOT a 403 on the whole download."""
    assert "By person" in crm_report._html(_data())
    assert "By person" not in crm_report._html(_data(includes_reps=False, reps=[]))


def test_the_csv_is_the_deals_and_only_the_deals():
    """A CSV cannot hold six sections, and pretending otherwise produces a file
    nothing can read."""
    text = crm_report.to_csv(_data()).decode("utf-8-sig")
    lines = text.strip().splitlines()
    assert len(lines) == 2, "header plus one deal"
    assert lines[0].startswith("Deal,Company,Contact,Stage,Value")
    assert "Conversion" not in text


def test_the_csv_carries_a_bom():
    """Otherwise Excel mangles any non-ASCII name in it."""
    assert crm_report.to_csv(_data()).startswith(b"\xef\xbb\xbf")


def test_the_workbook_has_a_sheet_per_section_and_the_raw_deals():
    import io

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(crm_report.to_excel(_data())))
    assert wb.sheetnames == ["Summary", "Forecast", "Velocity", "Sources",
                             "By person", "Deals"]


def test_the_workbook_drops_the_person_sheet_when_it_may_not_be_seen():
    import io

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(
        crm_report.to_excel(_data(includes_reps=False, reps=[]))))
    assert "By person" not in wb.sheetnames


def test_a_pdf_failure_does_not_take_excel_and_csv_with_it():
    """WeasyPrint's native libraries are absent on a Windows dev box: it imports
    and then raises OSError. Both that and ImportError must reach the route as a
    503 naming the alternative, not as a 500."""
    import inspect
    code = inspect.getsource(crm_report.to_pdf)
    assert "ImportError, OSError" in code
    assert "Excel" in code or "CSV" in code


def test_the_route_refuses_an_unknown_format():
    code = " ".join(inspect_src(graha.download_crm_report).split())
    assert "crm_report.FORMATS" in code and "400" in code


def test_the_route_asks_the_permission_question_before_gathering():
    code = " ".join(inspect_src(graha.download_crm_report).split())
    assert "held_module_levels" in code and "include_reps" in code


def inspect_src(fn) -> str:
    import inspect
    return inspect.getsource(fn)


@pytest.mark.parametrize("fmt", ["csv", "excel"])
def test_render_returns_content_type_and_extension(fmt):
    content, media, ext = crm_report.render(_data(), fmt)
    assert content and isinstance(content, bytes)
    assert media and ext in ("csv", "xlsx")


# ══════════════════════════════════════════════════════════════════════════
# "Won value" measured the wrong deals — proposal 73, defect 5
# ══════════════════════════════════════════════════════════════════════════
#
# `gather` had ONE window, `created_at > cutoff`, in the WHERE clause, and it
# summed `won_value` inside it. So the figure headed "Won value" for the last
# 90 days meant "deals CREATED in the last 90 days that are Won today" — a deal
# opened in March and won last week was missing from it entirely.
#
# Measured live 2026-08-22, last 90 days: E2E Rs53,13,648 shown against
# Rs66,37,948 actually won; Unicode Group Rs11,22,500 against Rs15,72,500. On
# the financial year to date Unicode is 39% low — Rs7,30,000 of real wins under
# a heading that claims to be the period's wins. `won_at` is filled on 33 of 33
# won deals, so nothing blocked the correct question being asked.

def test_won_value_windows_on_the_won_date_not_the_created_date():
    """The defect itself, pinned on the predicate.

    Structural rather than live: the suite is offline by design and the pool is
    a MagicMock, so the SQL predicate IS the population and comparing it is the
    honest offline form of the check.
    """
    import inspect
    sql = " ".join(inspect.getsource(crm_report.gather).split())
    assert "FILTER (WHERE stage='Won' AND won_at > $2)" in sql, (
        "the won-value figure is no longer windowed on `won_at`. A deal won "
        "inside the period but opened before it drops out of a figure headed "
        "'Won value' — 39% short on a real org when this was last measured."
    )
    assert "FROM public.graha_deals WHERE org_id=$1::uuid," not in sql, (
        "the WHERE clause has regained a `created_at` filter. Filtering the "
        "table on the created date puts every deal won-but-not-opened in the "
        "window beyond reach of the FILTERs that follow."
    )


def test_both_cohorts_are_kept_and_the_derived_figures_use_the_created_one():
    """Two populations, and the derived numbers must not straddle them.

    `conversion_rate` divides wins by openings, so both halves have to come
    from the deals OPENED in the window. Built from the closed-in-period count
    instead it can exceed 100%, and `open` — openings minus wins minus losses —
    can go NEGATIVE, because a period in which more deals close than open is
    completely ordinary.
    """
    import inspect
    src = " ".join(inspect.getsource(crm_report.gather).split())
    for name in ("cohort_won", "cohort_lost", "won_undated", "lost_undated"):
        assert name in src, f"{name} is gone from the conversion query"
    assert 'conv.get("cohort_won")' in src and 'conv.get("cohort_lost")' in src, (
        "`open` and `conversion_rate` are no longer derived from the created "
        "cohort — they can now print a negative or a rate above 100%."
    )


def test_every_conversion_figure_states_its_window_on_the_face_of_the_report():
    """The standing rule for this whole page: a number under a heading that
    does not say what it counts is the failure mode. Six figures sit in one
    strip and four of them count a different set of deals from the other two,
    so each caption carries its own basis.
    """
    data = _data(conversion={"total_deals": 10, "cohort_won": 3, "cohort_lost": 2,
                             "open": 5, "conversion_rate": 30.0, "won": 4,
                             "lost": 1, "won_value": 1234567, "avg_cycle_days": 21,
                             "won_undated": 0, "lost_undated": 0})
    html = crm_report._html(data)
    assert "Won value in period" in html and "Opened in period" in html
    assert "Conversion of opened" in html
    # The bare captions are what a reader quotes out of context. Checked
    # inside the stat strip only — the Sources table legitimately carries a
    # "Won to date" column over a different (lead-cohort) window.
    strip = html.split('<div class="stat">', 1)[1].split("</section>", 1)[0]
    assert "<span>Won</span>" not in strip
    assert "<span>Conversion</span>" not in strip

    content, _, _ = crm_report.render(data, "excel")
    assert content
