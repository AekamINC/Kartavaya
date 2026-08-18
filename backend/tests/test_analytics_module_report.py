"""GET /v1/analytics/module-report — one file per module page (proposal 65 S2).

The two promises under test:

  · the report resolves the SAME arrangement the module's analytics tab
    shows this caller — personal default > org default > first applicable
    preset > the registry-derived default the frontend's autoLayout draws —
    and runs each widget through the metric's OWN registry builder, so the
    file and the screen can never disagree about what a figure means;
  · a metric that cannot answer says so — a declared absence or a retired
    key renders as a STATED absence row in every format, never a convincing
    zero, and is never dropped silently (proposal 62 §10).

Same harness as test_analytics_client_report.py: a RecordingPool answering
registry SQL shapes, `_grant` faking `require_module` — the route's REAL
gate — and direct calls into the route function.
"""
from __future__ import annotations

import asyncio
import io
import json

import pytest
from fastapi import HTTPException

from routers import analytics as ax

USER = {"user_id": "user_aaa111"}
ORG = "22222222-2222-2222-2222-222222222222"
FROM, TO = "2026-05-01", "2026-08-17"


def run(coro):
    return asyncio.run(coro)


class FakeRequest:
    class _State:
        _auth_user = USER
    state = _State()
    method = "GET"


class RecordingPool:
    """Answers every query with a benign shape; records the SQL."""

    def __init__(self):
        self.calls = []
        self.views_rows = []       # what staging.analytics_views holds
        self.metric_rows = []      # what every registry builder's query returns

    async def fetch(self, sql, *args):
        self.calls.append((" ".join(sql.split()), list(args)))
        if "analytics_views" in sql:
            return self.views_rows
        return self.metric_rows

    async def fetchrow(self, sql, *args):
        self.calls.append((" ".join(sql.split()), list(args)))
        return None

    async def fetchval(self, sql, *args):
        self.calls.append((" ".join(sql.split()), list(args)))
        return 0


@pytest.fixture
def pool(monkeypatch):
    p = RecordingPool()

    async def _get_pool():
        return p

    monkeypatch.setattr(ax, "get_pool", _get_pool)
    return p


def _grant(monkeypatch, *modules):
    """Fake `require_module` — the route's REAL gate (subscription state,
    sensitive-module refusal, audit row); the seam mirrors
    test_analytics_client_report.py exactly."""
    def _rm(module):
        async def _gate(request, org_id):
            if module not in modules:
                raise HTTPException(403, f"{module} is not enabled")
        return _gate

    monkeypatch.setattr(ax, "require_module", _rm)


@pytest.fixture
def ganit_only(monkeypatch):
    _grant(monkeypatch, "ganit")


@pytest.fixture
def nothing_granted(monkeypatch):
    _grant(monkeypatch)


@pytest.fixture
def no_presets(monkeypatch):
    """Empty the preset shelf so resolution falls through to the derived
    default — every real module currently sits under some preset's modules
    tuple or has too few metrics to exercise the cap."""
    monkeypatch.setattr(ax, "PRESETS", {})


def report(pool, **over):
    kw = dict(module="ganit", date_from=FROM, date_to=TO, user=USER, org_id=ORG)
    kw.update(over)
    return run(ax.module_report(FakeRequest(), **kw))


#: A personal saved view: one real metric, one declared-absent metric. The
#: layout is a JSON STRING — asyncpg hands jsonb back that way on some paths,
#: and the resolver must parse rather than trust the driver's mood.
PERSONAL_VIEW = {
    "user_id": USER["user_id"],
    "name": "Mine",
    "layout": json.dumps([
        {"metric": "ganit.invoiced", "viz": "trend", "w": 2},
        {"metric": "ganit.tds_by_section", "viz": "kpi", "w": 1},
    ]),
}

ORG_VIEW = {
    "user_id": None,
    "name": "House style",
    "layout": [{"metric": "ganit.outstanding", "viz": "kpi", "w": 1}],
}


# ── refusals ─────────────────────────────────────────────────────────────────

def test_a_period_is_required(pool, ganit_only):
    with pytest.raises(HTTPException) as e:
        report(pool, date_from="", date_to="")
    assert e.value.status_code == 400


def test_one_bound_alone_is_still_400(pool, ganit_only):
    with pytest.raises(HTTPException) as e:
        report(pool, date_from=FROM, date_to="")
    assert e.value.status_code == 400


def test_the_five_year_cap_applies(pool, ganit_only):
    with pytest.raises(HTTPException) as e:
        report(pool, date_from="2019-01-01", date_to="2026-08-17")
    assert e.value.status_code == 400
    assert "maximum" in e.value.detail


def test_an_unknown_module_is_404(pool, ganit_only):
    with pytest.raises(HTTPException) as e:
        report(pool, module="astrology")
    assert e.value.status_code == 404


def test_an_unknown_format_is_400_never_silent_json(pool, ganit_only):
    with pytest.raises(HTTPException) as e:
        report(pool, format="docx")
    assert e.value.status_code == 400


def test_entitlement_refusal_is_403(pool, nothing_granted):
    with pytest.raises(HTTPException) as e:
        report(pool)
    assert e.value.status_code == 403


def test_core_never_touches_require_module(pool, nothing_granted):
    """core is the deliberately ungated surface — require_module('core') would
    refuse every org its own task counts. The fake gate refuses EVERYTHING,
    so this passing proves the gate was never consulted for core."""
    out = report(pool, module="core")
    assert out["module"] == "core"
    assert out["widgets"], "the core page resolved to nothing"


# ── resolution precedence ────────────────────────────────────────────────────

def test_a_personal_default_beats_the_org_default(pool, ganit_only):
    pool.views_rows = [ORG_VIEW, PERSONAL_VIEW]
    out = report(pool)
    assert out["source"] == "personal"
    assert [w["metric"] for w in out["widgets"]] == \
        ["ganit.invoiced", "ganit.tds_by_section"]


def test_the_org_default_stands_in_when_no_personal_one_exists(pool, ganit_only):
    pool.views_rows = [ORG_VIEW]
    out = report(pool)
    assert out["source"] == "org"
    assert [w["metric"] for w in out["widgets"]] == ["ganit.outstanding"]


def test_a_preset_stands_in_cut_to_this_module(pool, ganit_only):
    """Nothing saved: the first preset whose modules include ganit (founder),
    cut to its ganit widgets — never the other modules' widgets, which this
    caller was not gated for."""
    out = report(pool)
    assert out["source"] == "preset:founder"
    assert out["widgets"], "the preset cut left nothing"
    for w in out["widgets"]:
        assert ax.REGISTRY[w["metric"]].module == "ganit", w


def test_the_derived_default_caps_at_nine_in_registry_order(pool, ganit_only, no_presets):
    candidates = [m.key for m in ax.REGISTRY.values()
                  if m.module == "ganit" and not m.absent]
    assert len(candidates) > 9, "ganit no longer exercises the cap — pick a bigger module"
    out = report(pool)
    assert out["source"] == "derived"
    assert [w["metric"] for w in out["widgets"]] == candidates[:9]
    for w in out["widgets"]:
        m = ax.REGISTRY[w["metric"]]
        # autoLayout's mapping, mirrored: a flow draws as a trend, a stock
        # as a figure.
        assert w["viz"] == ("trend" if m.grain == "flow" else "kpi"), w


# ── the widgets run the registry's own SQL ───────────────────────────────────

def test_widgets_run_the_registry_builders_verbatim(pool, ganit_only, no_presets):
    """Each widget's query is character-for-character the builder's output —
    the report writes no SQL of its own."""
    from analytics.registry import MetricRequest
    import datetime

    out = report(pool)
    ran = [sql for sql, _ in pool.calls if "analytics_views" not in sql]
    win = ax.aw.parse(FROM, TO)
    for w, got in zip(out["widgets"], ran):
        m = ax.REGISTRY[w["metric"]]
        expected, params = m.sql(MetricRequest(
            org_id=ORG, window=win if m.grain == "flow" else None,
            bucket="month", group_by=None))
        assert got == " ".join(expected.split()), w["metric"]
        assert params[0] == ORG, "every query is org-scoped by its own binds"
        if m.grain == "flow":
            assert params[1:] and datetime.date(2026, 5, 1) in params, w["metric"]


def test_a_stock_widget_binds_no_window(pool, ganit_only, no_presets):
    report(pool)
    for sql, args in pool.calls:
        if "analytics_views" in sql:
            continue
        if "ganit_invoices" in sql and "ageing" not in sql and "BETWEEN" not in sql:
            # a stock (outstanding-shaped) query carries the org bind only
            assert len(args) == 1, (sql, args)


def test_a_foreign_module_widget_is_withheld_not_served(pool, ganit_only):
    """A saved ganit view naming a graha metric must not become a side door:
    the widget is gated on ITS module and a refusal withholds it — stated,
    and its SQL never runs."""
    pool.views_rows = [{
        "user_id": USER["user_id"], "name": "Sneaky",
        "layout": [
            {"metric": "ganit.invoiced", "viz": "trend", "w": 2},
            {"metric": "graha.pipeline_by_stage", "viz": "bars", "w": 2},
        ],
    }]
    out = report(pool)
    withheld = out["widgets"][1]
    assert "absent" in withheld and "graha" in withheld["absent"]
    assert "data" not in withheld
    assert not any("graha_" in sql for sql, _ in pool.calls), \
        "the refused module's SQL ran anyway"


# ── the stated absences ──────────────────────────────────────────────────────

def test_an_absent_metric_is_a_stated_absence_in_json(pool, ganit_only):
    pool.views_rows = [PERSONAL_VIEW]
    out = report(pool)
    absent = out["widgets"][1]
    assert absent["metric"] == "ganit.tds_by_section"
    assert "section column" in absent["absent"]
    assert "data" not in absent, "an absence must not also carry data"
    # and it never reached a query — sql is None on a declared-absent metric
    assert out["window"]["windowed"] == ["ganit.invoiced"]


def test_the_absence_survives_into_the_csv(pool, ganit_only):
    pool.views_rows = [PERSONAL_VIEW]
    resp = report(pool, format="csv")
    body = resp.body.decode("utf-8")
    assert "Not yet measurable" in body
    assert "section column" in body, "the reason must ride along, not just a flag"


def test_a_retired_key_is_stated_not_dropped(pool, ganit_only):
    pool.views_rows = [{
        "user_id": USER["user_id"], "name": "Old",
        "layout": [{"metric": "ganit.retired_metric", "viz": "kpi", "w": 1}],
    }]
    out = report(pool)
    assert len(out["widgets"]) == 1
    assert "no longer measured" in out["widgets"][0]["absent"]


# ── the files carry the same numbers ─────────────────────────────────────────

def test_csv_reuses_the_same_queries_not_a_second_pipeline(pool, ganit_only):
    """format is a parameter: the CSV runs the same SQL as the JSON — one
    query set, two renderings, byte-for-byte the same period."""
    report(pool)
    json_calls = [sql for sql, _ in pool.calls]
    pool.calls.clear()
    resp = report(pool, format="csv")
    csv_calls = [sql for sql, _ in pool.calls]
    assert json_calls == csv_calls
    body = resp.body.decode("utf-8")
    assert "Finance" in body, "the identity header must name the module"
    assert f"{FROM} to {TO}" in body


def test_the_filename_carries_module_and_period_never_an_id(pool, ganit_only):
    resp = report(pool, format="csv")
    cd = resp.headers["content-disposition"]
    assert f'filename="module-report_ganit_{FROM}_{TO}.csv"' in cd
    assert ORG not in cd and USER["user_id"] not in cd
    cd.encode("ascii")             # Starlette encodes headers latin-1; be stricter


def test_xlsx_has_a_sheet_per_widget_and_pk_magic(pool, ganit_only):
    """Four widgets — one metric twice (dedup) and one absent — make four
    sheets, uniquely named, identity on the first."""
    pool.views_rows = [{
        "user_id": USER["user_id"], "name": "Mine",
        "layout": [
            {"metric": "ganit.invoiced", "viz": "trend", "w": 2},
            {"metric": "ganit.invoiced", "viz": "kpi", "w": 1},
            {"metric": "ganit.receivables_ageing", "viz": "bars", "w": 2},
            {"metric": "ganit.tds_by_section", "viz": "kpi", "w": 1},
        ],
    }]
    resp = report(pool, format="xlsx")
    assert resp.body[:2] == b"PK", "not a zip container — not an xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.body))
    assert len(wb.sheetnames) == 4
    assert len(set(wb.sheetnames)) == 4, "duplicate labels must dedup, not collide"
    first = wb.worksheets[0]
    assert first["A1"].value == "Module report — Finance"
    assert first["A2"].value == f"{FROM} to {TO}"


def test_formula_shaped_strings_are_neutralised_in_csv_and_xlsx(pool, ganit_only):
    """A label like `=HYPERLINK(...)` — a client-typed string riding through a
    registry row — must open as text everywhere: CSV gets the apostrophe,
    and no xlsx cell may be a live formula (openpyxl treats `=`-leading
    strings as formulas)."""
    pool.metric_rows = [{"label": '=HYPERLINK("http://evil","x")', "value": 3}]
    resp = report(pool, format="csv")
    body = resp.body.decode("utf-8")
    assert "'=HYPERLINK" in body
    assert not any(line.startswith("=") for line in body.splitlines())
    resp = report(pool, format="xlsx")
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.body))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != "f", \
                    f"live formula cell {cell.coordinate} on {ws.title!r}"


# ── the sheet-title helper ───────────────────────────────────────────────────

def test_sheet_titles_are_capped_cleaned_and_deduped():
    used: set = set()
    long = ax._sheet_title("x" * 60, used)
    assert len(long) <= 31
    cleaned = ax._sheet_title("Bad[]:*?/\\name", used)
    assert not any(c in cleaned for c in "[]:*?/\\")
    a = ax._sheet_title("Invoiced", used)
    b = ax._sheet_title("Invoiced", used)
    c = ax._sheet_title("INVOICED", used)   # Excel compares case-insensitively
    assert a == "Invoiced" and b == "Invoiced (2)"
    assert len({a.lower(), b.lower(), c.lower()}) == 3
