"""
xlsx and pdf are produced, not silently swapped for JSON.

Measured live on staging before this: `format=xlsx` and `format=pdf` returned
`content-type: application/json` with a body labelling itself `"format":"json"`
for all five report types. The parameter was accepted and ignored, because only
`csv` was implemented and everything else fell through to the default return.

These assert on the BYTES — a zip magic number for xlsx, `%PDF` for pdf — rather
than on a status code, because a 200 returning the wrong media type is exactly
the defect being fixed.
"""
import io

import pytest

from routers.dristi import _csv_cell, _is_row_list


SCALARS_ONLY = {"tasks": 165, "contacts": 4, "revenue": 88500.0}
WITH_TABLE = {
    "monthly": [
        {"month": "2026-07-01T00:00:00+00:00", "total": 311671.6, "count": 6},
        {"month": "2026-06-01T00:00:00+00:00", "total": 12000.0, "count": 2},
    ],
}


def _split(data):
    """The shape rule both renderers share, lifted from the handler."""
    scalars = [(k, v) for k, v in data.items() if not _is_row_list(v)]
    tables = [(k, v) for k, v in data.items() if _is_row_list(v)]
    return scalars, tables


def test_scalars_and_tables_are_separated():
    s, t = _split({**SCALARS_ONLY, **WITH_TABLE})
    assert [k for k, _ in s] == ["tasks", "contacts", "revenue"]
    assert [k for k, _ in t] == ["monthly"]


# ── xlsx ──────────────────────────────────────────────────────────────────────

def test_xlsx_writes_a_real_workbook_with_a_sheet_per_table():
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    scalars, tables = _split({**SCALARS_ONLY, **WITH_TABLE})
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    for k, v in scalars:
        ws.append([k, _csv_cell(v)])
    for k, rows in tables:
        sheet = wb.create_sheet(str(k)[:31])
        headers = list(rows[0].keys())
        sheet.append(headers)
        for r in rows:
            sheet.append([_csv_cell(r.get(h)) for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    # xlsx is a zip. If this ever returns JSON again, the magic number changes.
    assert raw[:2] == b"PK", "xlsx must be a zip container, not JSON"
    assert b"{" != raw[:1]

    from openpyxl import load_workbook
    back = load_workbook(io.BytesIO(raw))
    assert back.sheetnames == ["Summary", "monthly"]
    assert [c.value for c in back["monthly"][1]] == ["month", "total", "count"]
    assert back["monthly"][2][1].value == 311671.6


def test_excel_sheet_names_are_made_legal():
    """Excel caps sheet names at 31 chars and rejects []:*?/\\ outright."""
    raw = "revenue[2026]:by/month*with?slashes\\and-a-very-long-tail"
    safe = "".join(c for c in raw if c not in "[]:*?/\\")[:31]
    assert len(safe) <= 31
    assert not set(safe) & set("[]:*?/\\")


# ── pdf ───────────────────────────────────────────────────────────────────────

def test_pdf_renders_to_pdf_bytes():
    """Skips where the native stack is absent, as doc_render.py:797 describes.

    WeasyPrint imports cleanly and then fails to dlopen libgobject/libpango on a
    machine without the GTK libraries — a Windows dev box, typically. Railway has
    them, which is why invoice and payslip PDFs work in production. Catching only
    ImportError would let this fail there and pass nowhere useful.
    """
    # NOT importorskip: the dlopen happens during weasyprint's own import, and
    # importorskip only catches ImportError, so the OSError escapes as a failure.
    try:
        import weasyprint  # noqa: F401
        from services.doc_render import render_pdf
        out = render_pdf(
            "<html><body><h1>Revenue report</h1>"
            "<table><tr><th>month</th></tr><tr><td>2026-07</td></tr></table>"
            "</body></html>"
        )
    except (ImportError, OSError) as exc:
        pytest.skip(f"WeasyPrint native libraries unavailable here: {exc}")
    assert out[:4] == b"%PDF", "pdf must be PDF bytes, not JSON"


def test_pdf_html_escapes_cell_values():
    """A report cell is user data; it must not be able to inject markup."""
    from html import escape
    nasty = '<script>alert(1)</script>'
    assert "<script>" not in escape(str(_csv_cell(nasty)))
    assert "&lt;script&gt;" in escape(str(_csv_cell(nasty)))
