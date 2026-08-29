"""Row-level report sections — the framework PLUGGED IN, and the five books.

`services/report_defs/` and `module_report.report_section` shipped working,
tested and unreachable: no route could list a section, no layout could hold
one, and the one layout item that could carry one crashed the report in every
format. This file pins the three things that changed, and the registers that
now fill the frame.

  · BUG 1 — `/module-report` built `window.windowed` / `window.as_at` with
    `w["metric"]` over every entry carrying `"data"`. A section returns
    `{report, label, grain, data}` — NO `metric` key — and satisfies both
    filters, so the subscript raised KeyError. The docstring on
    `report_section` claimed only `?format=json` was affected; it was wrong,
    because that payload is built BEFORE the format branches, so csv, xlsx
    and pdf died with it. `test_a_section_survives_every_format` is the pin,
    and it is parametrized over all four ON PURPOSE — a json-only test would
    have passed against the bug's own docstring.

  · BUG 2 — `_clean_layout` read `metric` off every entry and 422'd on the
    None, so `{"report": …}` could not be SAVED at all. The framework was
    unreachable from the API even with bug 1 fixed.

  · BUILD — `/report-sections` lists what a caller may put on a page.
    `sections_for` had no caller; a UI cannot offer what it cannot enumerate.

The register tests are ratchets, not decoration. Every guard asserted here
changes a live figure if it is dropped, and every count in a docstring was
measured read-only against the live database on 2026-08-20 before the
definition was written.
"""
from __future__ import annotations

import asyncio
import ast
import inspect
import io
import json
import re
from datetime import date, timedelta

import pytest
from fastapi import HTTPException

from routers import analytics as ax
from services import module_report as mr
from services.analytics_window import Window
from services.report_defs import REPORT_DEFS, load_all
from services.report_defs import _shared as sh
from services.report_defs import expense_register as er
from services.report_defs import payables_ageing as pa
from services.report_defs import purchase_register as pr
from services.report_defs import receipts_register as rr
from services.report_defs import sales_register as sr

USER = {"user_id": "user_aaa111"}
ORG = "22222222-2222-2222-2222-222222222222"
FROM, TO = "2026-05-01", "2026-08-17"
TODAY = date(2026, 8, 20)

#: The five definitions this commit adds, and the one that was already there.
NEW_KEYS = (
    "ganit.sales_register",
    "ganit.purchase_register",
    "ganit.receipts_register",
    "ganit.payables_ageing_by_bill",
    "ganit.expense_register_by_category",
)
ALL_KEYS = NEW_KEYS + ("ganit.receivables_ageing_by_party",)

#: Every module this commit wrote, for the cross-cutting scans below.
NEW_MODULES = (sr, pr, rr, pa, er, sh)


def run(coro):
    return asyncio.run(coro)


def sql_of(module) -> str:
    """The one query a definition module owns, whitespace-flattened."""
    name = next(n for n in dir(module) if n.endswith("_SQL"))
    return " ".join(getattr(module, name).split())


# ══════════════════════════════════════════════════════════════════════════
# the harness — test_analytics_module_report.py's, extended for sections
# ══════════════════════════════════════════════════════════════════════════

class FakeRequest:
    class _State:
        _auth_user = USER
    state = _State()
    method = "GET"


class RecordingPool:
    """Answers views with `views_rows` and everything else with `rows`."""

    def __init__(self):
        self.calls = []
        self.views_rows = []
        self.rows = []

    async def fetch(self, sql, *args):
        self.calls.append((" ".join(sql.split()), list(args)))
        if "analytics_views" in sql:
            return self.views_rows
        return self.rows

    async def fetchrow(self, sql, *args):
        self.calls.append((" ".join(sql.split()), list(args)))
        return None

    async def fetchval(self, sql, *args):
        self.calls.append((" ".join(sql.split()), list(args)))
        return 0


@pytest.fixture
def pool(monkeypatch):
    p = RecordingPool()

    async def _get_pool():
        return p

    monkeypatch.setattr(ax, "get_pool", _get_pool)
    return p


def _grant(monkeypatch, *modules):
    """Fake `require_module` — the route's REAL gate."""
    def _rm(module):
        async def _gate(request, org_id):
            if module not in modules:
                raise HTTPException(403, f"{module} is not enabled")
        return _gate

    monkeypatch.setattr(ax, "require_module", _rm)


@pytest.fixture
def ganit_only(monkeypatch):
    _grant(monkeypatch, "ganit")


@pytest.fixture
def echo_pdf(monkeypatch):
    """WeasyPrint is not this file's subject: `render_pdf` echoes its HTML so
    the assertion can read the DOCUMENT."""
    from services import doc_render as R
    monkeypatch.setattr(R, "render_pdf", lambda html: html.encode("utf-8"))


def saved(layout, name="Mine"):
    return {"user_id": USER["user_id"], "name": name, "layout": layout}


def report(pool, **over):
    kw = dict(module="ganit", date_from=FROM, date_to=TO, user=USER, org_id=ORG)
    kw.update(over)
    return run(ax.module_report(FakeRequest(), **kw))


#: A layout mixing both producers. The section is a FLOW register; the widget
#: is the metric whose one number the register puts documents under.
MIXED = [
    {"metric": "ganit.invoiced", "viz": "trend", "w": 2},
    {"report": "ganit.sales_register", "w": 2},
]


# ══════════════════════════════════════════════════════════════════════════
# BUG 1 — a section in a layout used to KeyError in ALL FOUR formats
# ══════════════════════════════════════════════════════════════════════════

@pytest.mark.parametrize("fmt", ["json", "csv", "xlsx", "pdf"])
def test_a_section_survives_every_format(pool, ganit_only, echo_pdf, fmt):
    """The regression, parametrized over all four renderings deliberately.

    `payload` is built at the top of the route, FOUR LINES before the
    `format == "json"` branch, and csv / xlsx / pdf are all downstream of it —
    so the KeyError on `w["metric"]` took every download with it, not just the
    JSON. A json-only test would have agreed with the docstring that was
    wrong.
    """
    pool.views_rows = [saved([{"report": "ganit.sales_register", "w": 2}])]
    out = report(pool, format=fmt)
    if fmt == "json":
        assert out["widgets"][0]["report"] == "ganit.sales_register"
        assert out["widgets"][0]["label"] == "Sales register"
    else:
        assert out.body, f"{fmt} rendered nothing"


def test_a_windowed_section_is_named_in_the_windowed_list(pool, ganit_only):
    """`_entry_key` reads a section's IDENTITY off `report`.

    Filtering the Nones out and stopping there would be safe and would also
    quietly claim a register that WAS windowed was not — the list is what the
    page prints beside "for this period".
    """
    pool.views_rows = [saved([{"report": "ganit.sales_register", "w": 2}])]
    out = report(pool)
    assert out["window"]["windowed"] == ["ganit.sales_register"]
    assert out["window"]["as_at"] == []


def test_a_stock_section_is_named_in_the_as_at_list(pool, ganit_only):
    pool.views_rows = [saved([{"report": "ganit.payables_ageing_by_bill", "w": 2}])]
    out = report(pool)
    assert out["window"]["as_at"] == ["ganit.payables_ageing_by_bill"]
    assert out["window"]["windowed"] == []


def test_one_layout_holds_both_producers(pool, ganit_only):
    """The whole design: a row-level report needed no new renderer, no new PDF
    engine and no new export code — only a second producer of the widget
    shape, in the same list."""
    pool.views_rows = [saved(MIXED)]
    out = report(pool)
    assert [w.get("metric") or w.get("report") for w in out["widgets"]] == \
        ["ganit.invoiced", "ganit.sales_register"]
    assert set(out["window"]["windowed"]) == {"ganit.invoiced", "ganit.sales_register"}


def test_a_withheld_section_is_never_named_in_a_window_list(pool, monkeypatch):
    """An entry that was REFUSED had no window applied to it, because it was
    never run. `_rendered_keys` filters on `"data" in w` for exactly that."""
    _grant(monkeypatch, "graha")           # ganit refused
    pool.views_rows = [saved([{"report": "ganit.sales_register", "w": 2}])]
    out = report(pool, module="graha")
    assert "absent" in out["widgets"][0] and "ganit" in out["widgets"][0]["absent"]
    assert out["window"]["windowed"] == [] and out["window"]["as_at"] == []
    assert not any("ganit_invoices" in s for s, _ in pool.calls), \
        "the refused section's SQL ran anyway"


def test_a_retired_section_key_is_stated_not_a_crash(pool, ganit_only):
    """A section can be retired after a view named it — the same way a metric
    can. It says so in words and does not take the report down."""
    pool.views_rows = [saved([{"report": "ganit.gone_away", "w": 2}])]
    out = report(pool)
    assert "retired" in out["widgets"][0]["absent"]
    assert out["window"]["windowed"] == []


def test_the_section_reaches_the_letterhead_and_the_csv(pool, ganit_only, echo_pdf):
    """End to end through the SHARED renderer: the register's own label and a
    row of its own columns, in the document and in the file."""
    pool.rows = [{"doc_date": date(2026, 6, 1), "doc_number": "INV-0007",
                  "doc_type": "tax_invoice", "party": "Sharma Textiles Pvt Ltd",
                  "place_of_supply": "27", "taxable": 1000.0, "cgst": 90.0,
                  "sgst": 90.0, "igst": 0.0, "cess": 0.0, "total": 1180.0}]
    pool.views_rows = [saved([{"report": "ganit.sales_register", "w": 2}])]

    html = bytes(report(pool, format="pdf").body).decode("utf-8")
    assert "Sales register" in html and "Sharma Textiles Pvt Ltd" in html
    assert sr.TOTAL_ROW in html

    body = report(pool, format="csv").body.decode("utf-8")
    assert "Sales register" in body and "INV-0007" in body


def test_xlsx_gives_the_section_its_own_sheet(pool, ganit_only):
    pool.views_rows = [saved(MIXED)]
    resp = report(pool, format="xlsx")
    assert resp.body[:2] == b"PK"
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.body))
    assert "Sales register" in wb.sheetnames


def test_rendered_keys_drops_a_shape_carrying_neither_key():
    """The `is not None` half of the fix, on its own. A producer this router
    does not have must not be able to raise inside the response builder,
    where the failure costs the whole report in every format."""
    entries = [{"data": [], "grain": "flow"},                       # no identity
               {"metric": "a.b", "data": [], "grain": "flow"},
               {"report": "a.c", "data": [], "grain": "flow"}]
    assert ax._rendered_keys(entries, "flow") == ["a.b", "a.c"]


# ══════════════════════════════════════════════════════════════════════════
# BUG 2 — `_clean_layout` refused to admit a section at all
# ══════════════════════════════════════════════════════════════════════════

def test_a_section_can_now_be_saved():
    out = ax._clean_layout([{"report": "ganit.sales_register", "w": 3}])
    assert out == [{"report": "ganit.sales_register", "w": 3}]


def test_an_unknown_report_key_is_422_naming_the_catalogue():
    """A rule the builder cannot express must be unwritable — the promise
    `_clean_layout` already made for metrics, extended."""
    with pytest.raises(HTTPException) as e:
        ax._clean_layout([{"report": "ganit.not_a_report"}])
    assert e.value.status_code == 422
    assert "report-sections" in e.value.detail


def test_an_entry_naming_both_a_metric_and_a_report_is_refused():
    """Two entries wearing one hat: whichever producer ran it, the other half
    would be silently discarded."""
    with pytest.raises(HTTPException) as e:
        ax._clean_layout([{"report": "ganit.sales_register",
                           "metric": "ganit.invoiced"}])
    assert e.value.status_code == 422
    assert "not both" in e.value.detail


@pytest.mark.parametrize("junk", [
    {"viz": "trend"}, {"group_by": "invoice_type"}, {"columns": ["a"]},
])
def test_widget_only_fields_are_refused_on_a_section(junk):
    """Accepting them would save a preference `render_report_html` then
    ignores — which reads, to the person who set it, as the product
    forgetting what they asked for."""
    with pytest.raises(HTTPException) as e:
        ax._clean_layout([{"report": "ganit.sales_register", **junk}])
    assert e.value.status_code == 422
    assert "report section" in e.value.detail


def test_a_section_sits_on_the_same_grid_a_widget_does():
    out = ax._clean_layout([{"report": "ganit.sales_register",
                             "w": 6, "x": 6, "y": 2, "h": 4}])
    assert out == [{"report": "ganit.sales_register", "w": 6,
                    "x": 6, "y": 2, "h": 4}]
    with pytest.raises(HTTPException) as e:
        ax._clean_layout([{"report": "ganit.sales_register", "w": 8, "x": 6}])
    assert "12-column grid" in e.value.detail


def test_a_metric_widget_still_rebuilds_exactly_as_it_did():
    """The geometry and width checks moved into shared helpers so a section
    could reuse them. A legacy widget must still rebuild BYTE-IDENTICAL, or a
    re-save rewrites rows it did not touch."""
    assert ax._clean_layout([{"metric": "ganit.invoiced"}]) == \
        [{"metric": "ganit.invoiced", "viz": "kpi", "w": 1}]
    assert ax._clean_layout([
        {"metric": "ganit.invoiced", "viz": "trend", "w": 4,
         "x": 2, "y": 7, "h": 3, "group_by": "invoice_type"}]) == \
        [{"metric": "ganit.invoiced", "viz": "trend", "w": 4,
          "x": 2, "y": 7, "h": 3, "group_by": "invoice_type"}]
    for bad in ({"metric": "nope.nope"}, {"metric": "ganit.invoiced", "w": 13},
                {"metric": "ganit.invoiced", "y": -1}):
        with pytest.raises(HTTPException):
            ax._clean_layout([bad])


def test_the_saved_layout_and_the_render_agree_on_what_a_section_is():
    """`is_section` is the ONE test both sides use — the router IMPORTS the
    service's, it does not keep a copy — so a layout cannot save as one thing
    and resolve as another."""
    assert ax._is_section is mr.is_section
    item = ax._clean_layout([{"report": "ganit.sales_register"}])[0]
    assert ax._is_section(item)
    assert not ax._is_section(ax._clean_layout([{"metric": "ganit.invoiced"}])[0])
    assert not ax._is_section("not a dict")


def test_the_widget_ceiling_still_counts_sections():
    over = [{"report": "ganit.sales_register"}] * (ax.MAX_WIDGETS + 1)
    with pytest.raises(HTTPException) as e:
        ax._clean_layout(over)
    assert "at most" in e.value.detail


# ══════════════════════════════════════════════════════════════════════════
# the dispatcher — the seam the OTHER two doors still owe
# ══════════════════════════════════════════════════════════════════════════

class _EmptyPool:
    async def fetch(self, sql, *args):
        return []


def test_report_entry_sends_each_kind_to_its_own_producer():
    win = Window(date(2026, 5, 1), date(2026, 8, 17))
    section = run(mr.report_entry(_EmptyPool(), ORG, "ganit", win,
                                  {"report": "ganit.sales_register"}, None, {}))
    widget = run(mr.report_entry(_EmptyPool(), ORG, "ganit", win,
                                 {"metric": "ganit.invoiced"}, None, {}))
    assert section["report"] == "ganit.sales_register"
    assert "metric" not in section
    assert widget["metric"] == "ganit.invoiced"
    assert "report" not in widget


def test_the_wrong_producer_turns_a_register_into_a_wrong_sentence():
    """WHY the dispatcher exists, demonstrated rather than asserted.

    `routers/dristi.py` (scheduled run-now) and `services/niyam/actions.py`
    (`report.send`) still walk a saved layout calling `report_widget`
    DIRECTLY. Hand it a section and it reads `metric` off the entry, gets
    None, misses the registry, and renders the register as "This metric is no
    longer measured" under the label "None" — on a document that is EMAILED.
    Both are safe only while no saved layout holds a section; the fix is one
    line each, in files this commit does not own.
    """
    win = Window(date(2026, 5, 1), date(2026, 8, 17))
    wrong = run(mr.report_widget(_EmptyPool(), ORG, "ganit", win,
                                 {"report": "ganit.sales_register"}, None, {}))
    assert wrong["label"] == "None"
    assert "no longer measured" in wrong["absent"]
    assert "data" not in wrong
    # …and the dispatcher, handed exactly the same entry, prints the register.
    right = run(mr.report_entry(_EmptyPool(), ORG, "ganit", win,
                                {"report": "ganit.sales_register"}, None, {}))
    assert right["label"] == "Sales register" and "data" in right


# ══════════════════════════════════════════════════════════════════════════
# BUILD — GET /report-sections
# ══════════════════════════════════════════════════════════════════════════

def _hold(monkeypatch, *modules):
    async def _held(pool, user_id, org_id, code):
        return "full" if code in modules else None
    monkeypatch.setattr(ax, "held_level", _held)


def sections(pool):
    return run(ax.report_sections(user=USER, org_id=ORG))


def offered_to(*held) -> list[str]:
    """The keys a caller holding `held` may be offered, computed FROM THE RULE.

    THIS USED TO BE A FROZEN TUPLE and it did not survive contact with a second
    author. `ALL_KEYS` listed six ganit registers; the moment `core.workload_now`
    landed — declaring `reads={"core"}`, and `core` is in `UNGATED_MODULES`, so
    every org member may open it — three assertions here went red on a section
    that was behaving exactly as designed. A frozen list cannot tell "somebody
    added a section" apart from "the gate leaked", which is the only thing these
    tests are for.

    So the expectation is now the rule itself, evaluated over the live catalogue:
    a section is offered when everything it READS is either held or ungated.
    Add a section and this follows it. Widen what a grant reaches and it fails,
    which is the failure worth having.
    """
    reachable = set(held) | set(ax.UNGATED_MODULES)
    return sorted(k for k, d in REPORT_DEFS.items() if set(d.reads) <= reachable)


def test_the_catalogue_lists_every_section_a_ganit_holder_may_open(pool, monkeypatch):
    _hold(monkeypatch, "ganit")
    out = sections(pool)
    expected = offered_to("ganit")
    assert [s["key"] for s in out["sections"]] == expected
    # Every ganit register is in there — the specific thing this test was
    # written to prove, still asserted by name rather than by arithmetic.
    assert set(ALL_KEYS) <= set(expected)
    assert out["withheld_count"] == len(REPORT_DEFS) - len(expected)


def test_a_caller_with_no_finance_grant_is_offered_nothing(pool, monkeypatch):
    """Unreachable sections are ABSENT, not listed and disabled — the rule
    `catalogue_for` holds — and the count says how many were hidden, so a UI
    can say "6 more with other modules" instead of looking empty."""
    _hold(monkeypatch, "graha")
    out = sections(pool)
    keys = [s["key"] for s in out["sections"]]
    assert keys == offered_to("graha")
    # NOT A FINANCE SECTION IN SIGHT. That is the claim in the name, and it is
    # what must keep holding; whether an UNGATED `core` section is also listed
    # is a different question, asked below.
    assert not any(k.startswith("ganit.") for k in keys)
    assert out["withheld_count"] == len(REPORT_DEFS) - len(keys)


def test_an_ungated_section_is_offered_to_a_member_holding_nothing(pool, monkeypatch):
    """The consequence of `UNGATED_MODULES = {"core"}`, stated out loud.

    A section that reads only `core` is offered to any member of the org, with
    no module grant at all — that is what "ungated" means, and it is currently
    true by inheritance from the metric catalogue rather than by a decision
    anybody wrote down for REGISTERS. Pinning it means the day somebody wants
    core sections gated, this test is where the decision surfaces, instead of
    the change landing silently.
    """
    _hold(monkeypatch)                        # holds nothing whatsoever
    out = sections(pool)
    keys = [s["key"] for s in out["sections"]]
    assert keys == offered_to()
    for k in keys:
        assert set(REPORT_DEFS[k].reads) <= set(ax.UNGATED_MODULES), (
            f"{k} is offered to a caller holding nothing but reads "
            f"{sorted(REPORT_DEFS[k].reads)}"
        )


def test_the_catalogue_carries_no_id_and_no_callable(pool, monkeypatch):
    """These rows are printed in a UI. No user, member or org id belongs in
    one (decision_names_not_ids), and `run` would not survive JSON anyway."""
    _hold(monkeypatch, "ganit")
    out = sections(pool)
    for s in out["sections"]:
        assert set(s) == {"key", "module", "label", "grain", "sensitivity",
                          "description"}
        assert s["description"], f"{s['key']} offers no description"
    blob = json.dumps(out)
    assert ORG not in blob and USER["user_id"] not in blob


def test_listing_names_never_runs_the_data_door(pool, monkeypatch):
    """`held_level`, not `require_module`: listing a register's existence is
    not serving its rows, and `/catalogue` draws the same line. The fake gate
    refuses EVERYTHING, so this passing proves the data door was not used —
    and no register's SQL ran."""
    _hold(monkeypatch, "ganit")
    _grant(monkeypatch)                      # require_module refuses all
    out = sections(pool)
    assert len(out["sections"]) == len(offered_to("ganit"))
    assert not any("ganit_invoices" in s for s, _ in pool.calls)


def test_every_module_a_definition_reads_is_asked_about(pool, monkeypatch):
    """`_reachable` walks the metric registry, and nothing requires a module a
    ReportDef reads to carry a Dristi metric. Without `also=`, such a section
    would be withheld for a reason that is not the caller's entitlement."""
    asked = []

    async def _held(pool_, user_id, org_id, code):
        asked.append(code)
        return "full"

    monkeypatch.setattr(ax, "held_level", _held)
    sections(pool)
    declared = {c for d in REPORT_DEFS.values() for c in d.reads}
    assert declared <= (set(asked) | ax.UNGATED_MODULES)


# ══════════════════════════════════════════════════════════════════════════
# the five definitions — declaration
# ══════════════════════════════════════════════════════════════════════════

def test_the_five_registers_are_declared():
    load_all()
    for key in NEW_KEYS:
        assert key in REPORT_DEFS, key
        d = REPORT_DEFS[key]
        assert d.module == "ganit"
        assert d.reads == frozenset({"ganit"})
        assert d.sensitivity == "financial"
        assert d.grain in ("flow", "stock")
        assert len(d.description) > 60, f"{key}: the description says too little"


@pytest.mark.parametrize("key,grain", [
    ("ganit.sales_register", "flow"),
    ("ganit.purchase_register", "flow"),
    ("ganit.receipts_register", "flow"),
    ("ganit.expense_register_by_category", "flow"),
    # A balance is what is unpaid NOW, not at a period end. Declared stock, so
    # `report_section` hands it None and it cannot silently read a window it
    # does not honour.
    ("ganit.payables_ageing_by_bill", "stock"),
])
def test_each_register_declares_the_grain_its_answer_actually_has(key, grain):
    load_all()
    assert REPORT_DEFS[key].grain == grain


@pytest.mark.parametrize("module", [sr, pr, rr, er])
def test_a_flow_register_handed_no_window_fails_loudly(module):
    """The two answers worse than raising are inventing a period and
    returning no rows. Neither is visible to the reader; this is."""
    fn = next(getattr(module, n) for n in dir(module)
              if n == module.KEY.split(".", 1)[1])
    with pytest.raises(ValueError, match="no window"):
        run(fn(None, ORG, None))


# ══════════════════════════════════════════════════════════════════════════
# cross-cutting SQL discipline — one scan, every register
# ══════════════════════════════════════════════════════════════════════════

@pytest.mark.parametrize("module", [sr, pr, rr, pa, er])
def test_every_register_is_org_scoped_by_a_bind_in_sql(module):
    """Never filtered in Python afterwards, and never interpolated: `$1::uuid`
    is cast because PgBouncer turns an untyped parse error into an instant
    500 (the credits incident)."""
    sql = sql_of(module)
    assert "org_id = $1::uuid" in sql, module.__name__
    assert "%" not in sql and "format(" not in sql.lower()


@pytest.mark.parametrize("module", [sr, pr, rr, pa, er])
def test_every_register_qualifies_its_schema(module):
    """`search_path` on this database is `"$user", public, extensions`, so an
    unqualified table resolves to nothing — and a shadow table in `public` has
    bitten this repo before (migration 142)."""
    sql = sql_of(module)
    assert " FROM public." in sql, module.__name__
    for join in re.findall(r"JOIN\s+(\S+)", sql):
        assert join.startswith("public."), (module.__name__, join)
    assert "SELECT *" not in sql


@pytest.mark.parametrize("module", [sr, rr])
def test_the_crm_join_carries_an_org_predicate(module):
    """`graha_clients` joined on `id` alone can surface ANOTHER org's client
    name — the latent cross-tenant leak this repo already found once, and the
    shape `receivables_ageing.py` carried until it was scoped on 2026-08-20.
    `ganit_invoices.client_id` has a plain FK to `graha_clients(id)` with no
    composite `(id, org_id)` constraint, so the schema cannot refuse a foreign
    company id and only this predicate can."""
    sql = sql_of(module)
    for alias, col in (("cl", "client_id"), ("ct", "contact_id")):
        assert f"ON {alias}.id = i.{col} AND {alias}.org_id = i.org_id" in sql, \
            (module.__name__, alias)


@pytest.mark.parametrize("module", [pr, pa])
def test_the_vendor_join_carries_an_org_predicate(module):
    assert "ON v.id = b.vendor_id AND v.org_id = b.org_id" in sql_of(module)


def test_no_join_anywhere_in_the_package_lacks_an_org_predicate():
    """The ratchet, over EVERY definition including the ones this commit did
    not write. A join on `id` alone is not a bug in one file, it is a shape
    that keeps coming back — `receivables_ageing.py` carried it until
    2026-08-20 — and the schema cannot stop it: none of these foreign keys has
    a composite `(id, org_id)` constraint, so the predicate is the only guard
    there is. Nothing in this package may join without one, ever again.
    """
    from services.report_defs import receivables_ageing as ra

    for module in (sr, pr, rr, pa, er, ra):
        sql = sql_of(module)
        # Each ON clause runs to the next JOIN or the WHERE.
        for clause in re.findall(
                r"JOIN\s+public\.\w+\s+\w+\s+ON\s+(.*?)(?=\s+(?:LEFT\s+)?JOIN\s|\s+WHERE\s)",
                sql):
            assert "org_id" in clause, (module.__name__, clause)


@pytest.mark.parametrize("module", [sr, pr, rr, pa, er])
def test_no_register_renders_an_id(module):
    """These rows are printed on a page the firm hands to someone. A member,
    client or org UUID must never reach one (decision_names_not_ids)."""
    sql = sql_of(module)
    selected = sql[sql.index("SELECT"):sql.index(" FROM ")]
    for token in ("id AS", ".id,", "client_id AS", "vendor_id AS", "org_id AS"):
        assert token not in selected, (module.__name__, token)


@pytest.mark.parametrize("module", NEW_MODULES)
def test_nothing_in_a_register_module_writes(module):
    """A report reads. The patterns are SQL-shaped rather than bare verbs
    because the prose in these files legitimately says things like "would
    silently drop those rows" — banning the word would ban the explanation."""
    src = inspect.getsource(module).lower()
    for verb in ("insert into", "delete from", "update public",
                 "drop table", "alter table", "truncate"):
        assert verb not in src, (module.__name__, verb)


@pytest.mark.parametrize("module", [sr, pr, rr, pa, er])
def test_every_register_query_is_a_bare_select(module):
    sql = sql_of(module).lower()
    assert sql.startswith("select ")
    for verb in ("insert", "update", "delete", "drop", "alter", "truncate",
                 "grant", ";"):
        assert verb not in sql, (module.__name__, verb)


@pytest.mark.parametrize("module", NEW_MODULES)
def test_no_statutory_fact_is_a_literal(module):
    """`services/statute.py` is the ONLY source of a form number, a section
    reference or a due date — `obligation(pool, key, *, as_of=...)`, where
    `as_of` is keyword-only with no default on purpose. These are COMMERCIAL
    registers: the firm's own books. They name no form, and the day one of
    them needs to, it asks.

    Docstrings are stripped before the scan, exactly as
    `tests/test_payroll_statutory.py` does it, so prose that merely says the
    words "form number" is not the thing being banned — a LITERAL is.
    """
    tree = ast.parse(inspect.getsource(module))
    for node in ast.walk(tree):
        if isinstance(node, (ast.Module, ast.FunctionDef, ast.AsyncFunctionDef,
                             ast.ClassDef)):
            body = getattr(node, "body", [])
            if (body and isinstance(body[0], ast.Expr)
                    and isinstance(body[0].value, ast.Constant)
                    and isinstance(body[0].value.value, str)):
                node.body = body[1:]
    literals = [n.value for n in ast.walk(tree)
                if isinstance(n, ast.Constant) and isinstance(n.value, str)]
    banned = re.compile(r"\bForm[\s-]?\d|\b\d{2}[A-Z]Q\b|\b24Q\b|\b26Q\b"
                        r"|\bu/s\b|\bsection\s+\d", re.I)
    for lit in literals:
        assert not banned.search(lit), (module.__name__, lit)
    numbers = {n.value for n in ast.walk(tree)
               if isinstance(n, ast.Constant) and isinstance(n.value, int)}
    # The MSMED 45-day limit and the ESI wage ceiling are the two statutory
    # constants nearest this code; neither belongs in a commercial register.
    assert 21000 not in numbers and 45 not in numbers


# ══════════════════════════════════════════════════════════════════════════
# sales register
# ══════════════════════════════════════════════════════════════════════════

def _doc(number="INV-1", kind="tax_invoice", taxable=1000.0, cgst=90.0,
         sgst=90.0, igst=0.0, cess=0.0, total=1180.0, party="Menon Traders LLP"):
    return {"doc_date": date(2026, 6, 1), "doc_number": number,
            "doc_type": kind, "party": party, "place_of_supply": "27",
            "taxable": taxable, "cgst": cgst, "sgst": sgst, "igst": igst,
            "cess": cess, "total": total}


def test_a_credit_note_is_signed_across_every_money_column():
    """22 of the 685 live issued documents are credit notes, stored with
    POSITIVE totals. Summing them as issued ADDS reversals to turnover;
    negating the total but not the tax split is how the rows stop adding up
    across. `ganit.invoiced` already negates the same set."""
    # Every money column non-zero, so "signed" is proved on all six rather
    # than on whichever ones the fixture happened to fill.
    rows = sr.build_rows([_doc(kind="credit_note", taxable=1000.0, cgst=90.0,
                               sgst=90.0, igst=45.0, cess=10.0, total=1235.0)])
    note = rows[0]
    assert note["Type"] == "Credit note"
    for col in sr.MONEY_COLUMNS:
        assert note[col] < 0, col
    assert note["Taxable value"] + note["CGST"] + note["SGST"] + \
        note["IGST"] + note["Cess"] == note["Total"]


def test_a_debit_note_is_not_negated():
    """A debit note increases what the customer owes — the same direction as
    an invoice."""
    assert sr.build_rows([_doc(kind="debit_note")])[0]["Total"] > 0


def test_the_footer_is_the_sum_of_the_rows_above_it():
    """The one figure a reader cross-checks. Rounding is per CELL for exactly
    this reason: summing raw floats and rounding at the end is how a register
    ends up a paisa off its own columns."""
    rows = sr.build_rows([_doc(), _doc(number="INV-2"),
                          _doc(number="CN-1", kind="credit_note")])
    total = rows[-1]
    assert total[sr.LABEL_COLUMN] == sr.TOTAL_ROW
    assert total["Total"] == 1180.0
    assert total["Date"] == "", "a footer in a date column is not a date"
    for col in sr.MONEY_COLUMNS:
        assert total[col] == round(sum(r[col] for r in rows[:-1]), 2), col


def test_an_empty_register_prints_no_footer():
    """`render_report_html` prints "No rows for this period" for an empty
    list, which is the honest page. A lone row of zeros reads as a register
    that ran and found nothing happened, when it may equally have found
    nothing at all."""
    assert sr.build_rows([]) == []


def test_the_party_falls_back_the_way_the_invoice_itself_does():
    """`client_id` is NULL on 189 of the 685 issued documents. The invoice
    DOCUMENT already prints the contact's name and company in that case, so
    the register resolves the same chain rather than blanking 27% of a
    column. Nine documents in the whole database name nobody."""
    sql = sql_of(sr)
    assert ("COALESCE(NULLIF(TRIM(cl.name), ''), NULLIF(TRIM(ct.company), ''), "
            "NULLIF(TRIM(ct.name), ''), $4::text) AS party") in sql
    assert sr.build_rows([_doc(party=None)])[0][sr.PARTY_COLUMN] == sr.UNLINKED


@pytest.mark.parametrize("guard", [
    # Soft delete.
    "i.is_active = TRUE",
    # NEVER `= 'final'`: doc_status defaults to 'final' and the live values are
    # final/viewed/sent/draft — an equality test drops 155 issued documents.
    "i.doc_status <> 'draft'",
    # Two columns record a cancellation and neither is authoritative alone.
    "i.cancelled_at IS NULL",
    "i.payment_status <> 'cancelled'",
    # Proformas and quotations are OFFERS. 0 live rows; the guard is what
    # stops the first one landing in the sales book.
    "NOT (i.invoice_type = ANY($5::text[]))",
    # The document's own date, not the date it was typed in.
    "i.invoice_date BETWEEN $2::date AND $3::date",
])
def test_the_sales_guards_are_all_present(guard):
    assert guard in sql_of(sr), guard


def test_the_offer_types_are_bound_not_interpolated():
    assert sr.OFFER_TYPES == ("proforma", "quotation")
    rows = run(_capture(sr.sales_register))
    assert rows["args"][4] == ["proforma", "quotation"]


def test_the_taxable_value_is_subtotal_less_discount():
    """Verified on all 685 live issued rows: subtotal - discount + the four
    tax columns equals total, 0 mismatches. So the register adds across, which
    is the check a reader performs first."""
    assert ("(COALESCE(i.subtotal, 0) - COALESCE(i.discount, 0))::float AS taxable"
            in sql_of(sr))


def test_an_unmapped_document_type_still_reads():
    """A sixth type added to the CHECK prints readably on day one instead of
    waiting for the label map to be updated."""
    assert sr.type_label("self_billed_invoice") == "Self billed invoice"
    assert sr.type_label(None) == "Document"


# ══════════════════════════════════════════════════════════════════════════
# purchase register
# ══════════════════════════════════════════════════════════════════════════

def _bill(number="BILL-1", gstin="27AAAAA0000A1Z5", status="unpaid"):
    return {"bill_date": date(2026, 6, 2), "bill_number": number,
            "vendor": "Kulkarni Supplies", "vendor_gstin": gstin,
            "status": status, "taxable": 500.0, "cgst": 45.0, "sgst": 45.0,
            "igst": 0.0, "cess": 0.0, "total": 590.0}


def test_a_vendor_with_no_gstin_gets_a_blank_cell_not_a_dropped_row():
    """Only 51 of the 80 live vendors carry a GSTIN, leaving 16 of the 189
    bills with an empty cell. GSTIN is NON-MANDATORY in this product and
    blocks nothing: an unregistered vendor is a real vendor. Dropping the
    column makes the register useless to the 51; dropping the rows makes it a
    lie."""
    rows = pr.build_rows([_bill(gstin=None), _bill(number="BILL-2")])
    assert rows[0]["Vendor GSTIN"] == ""
    assert rows[1]["Vendor GSTIN"] == "27AAAAA0000A1Z5"
    assert rows[-1]["Total"] == 1180.0, "the blank-GSTIN bill still counts"


def test_a_cancelled_bill_is_not_a_purchase():
    sql = sql_of(pr)
    assert "b.status <> $5::text" in sql
    assert pr.CANCELLED_STATUS == "cancelled"


def test_the_purchase_footer_ties():
    rows = pr.build_rows([_bill(), _bill(number="BILL-2")])
    total = rows[-1]
    assert total[pr.LABEL_COLUMN] == pr.TOTAL_ROW
    for col in pr.MONEY_COLUMNS:
        assert total[col] == round(sum(r[col] for r in rows[:-1]), 2), col
    assert total["Taxable value"] + total["CGST"] + total["SGST"] + \
        total["IGST"] + total["Cess"] == total["Total"]


def test_bill_status_reads_as_english():
    assert pr.status_label("partially_paid") == "Part paid"
    assert pr.status_label("unpaid") == "Unpaid"
    assert pr.status_label("some_new_state") == "Some new state"


# ══════════════════════════════════════════════════════════════════════════
# receipts register
# ══════════════════════════════════════════════════════════════════════════

def _receipt(ref="UTR-9911", method="bank_transfer", amount=5000.0):
    return {"paid_on": date(2026, 6, 3), "reference": ref, "method": method,
            "doc_number": "INV-1", "party": "Iyer Consulting Industries",
            "amount": amount}


def test_a_receipt_with_no_reference_still_appears():
    """11 of the 506 live payments carry an empty-string reference. A receipt
    recorded without a UTR or cheque number is still money in the bank, and
    this register has to tie to a bank statement."""
    rows = rr.build_rows([_receipt(ref="")])
    assert rows[0][rr.LABEL_COLUMN] == ""
    assert rows[0]["Amount"] == 5000.0


def test_every_payment_method_the_check_admits_reads_as_english():
    """The CHECK admits six values and all six are mapped, so no register cell
    ever prints a database enum at a person."""
    for raw in ("cash", "bank_transfer", "upi", "cheque", "card", "other"):
        assert rr.METHOD_LABELS[raw][0].isupper()
    assert rr.method_label("bank_transfer") == "Bank transfer"
    assert rr.method_label("standing_order") == "Standing order"


def test_the_receipts_join_is_inner_and_org_scoped_on_both_sides():
    """`invoice_id` resolves on all 506 live payments and 0 point at another
    org's invoice. The org predicate is belt-and-braces on the joined row —
    fail-closed beats trusting a foreign key one hop away."""
    sql = sql_of(rr)
    assert "JOIN public.ganit_invoices i ON i.id = p.invoice_id AND i.org_id = p.org_id" in sql
    assert "LEFT JOIN public.ganit_invoices" not in sql


def test_the_receipts_register_does_not_filter_on_doc_status():
    """4 of the 506 live payments sit against a DRAFT invoice, and they are in
    the register on purpose: the money is in the bank either way. This product
    has no payment gateway and never will — "paid" only ever comes from bank
    reconciliation — so a register that misses those four does not tie to the
    statement and is worth nothing."""
    assert "doc_status" not in sql_of(rr)


def test_the_receipts_footer_ties():
    rows = rr.build_rows([_receipt(), _receipt(amount=2500.5)])
    assert rows[-1][rr.LABEL_COLUMN] == rr.TOTAL_ROW
    assert rows[-1]["Amount"] == 7500.5
    assert rows[-1]["Date"] == ""


# ══════════════════════════════════════════════════════════════════════════
# payables ageing by bill
# ══════════════════════════════════════════════════════════════════════════

def _open_bill(days_overdue, balance, number="BILL-1", vendor="Kulkarni Supplies"):
    return {"vendor": vendor, "bill_number": number,
            "bill_date": TODAY - timedelta(days=days_overdue + 30),
            "due_date": TODAY - timedelta(days=days_overdue),
            "bill_total": balance, "paid": 0.0, "balance_due": balance}


def test_the_bucketing_is_asked_of_the_shared_ager_not_reimplemented():
    """The payables page, the receivables page and the client statement all
    bucket through ONE function, so they cannot disagree about what "61–90"
    means. If this fails because someone inlined a CASE ladder, the failure is
    the point."""
    from services import statement_pdf
    assert pa.age_receivables is statement_pdf.age_receivables
    assert pa.AGEING_BUCKETS is statement_pdf.AGEING_BUCKETS


@pytest.mark.parametrize("days,bucket", [
    (-7, "Current"), (0, "Current"), (1, "1–30 days"), (30, "1–30 days"),
    (31, "31–60"), (60, "31–60"), (61, "61–90"), (90, "61–90"),
    (91, "90+"), (400, "90+"),
])
def test_every_bucket_boundary(days, bucket):
    """Day zero is not day one: a boundary that slips by one prints a vendor
    as late on the morning the payment is due."""
    assert pa.bucket_of(1000.0, TODAY - timedelta(days=days), TODAY) == bucket


def test_ageing_anchors_on_the_due_date_not_the_bill_date():
    """Measured on the live open payables: anchoring on `bill_date` puts
    ₹33,02,025.16 in 90+ where the due date puts ₹28,69,924.48 — ₹4,32,100.68
    reported as three months late that is not. The row below is 100 days past
    its BILL date and 70 past its due date."""
    row = pa.build_rows([_open_bill(70, 1000.0)], TODAY)[0]
    assert row["Ageing"] == "61–90"
    assert row["Days overdue"] == 70


def test_a_bill_with_no_due_date_ages_from_the_bill_date():
    """16 of the 95 open bills carry no due date. Ageing on `due_date` alone
    would drop a sixth of the payables off the page, and a ledger that omits
    rows reconciles against nothing while still looking complete."""
    assert "COALESCE(b.due_date, b.bill_date) AS due_date" in sql_of(pa)


def test_days_overdue_never_goes_negative():
    """The column is headed "Days overdue"; -7 in it is a different fact
    wearing the same header. How far away it is stays visible in the Due date
    column beside it."""
    row = pa.build_rows([_open_bill(-7, 1000.0)], TODAY)[0]
    assert row["Days overdue"] == 0 and row["Ageing"] == "Current"
    assert row["Due date"] == TODAY + timedelta(days=7)


def test_the_page_is_ordered_to_be_worked_down():
    """Most overdue first, then biggest. A bill 400 days late for ₹9,000 is
    still the one a firm settles before a ₹9 lakh bill not due until next
    week."""
    rows = pa.build_rows([
        _open_bill(-5, 900000.0, "BILL-NEW"),
        _open_bill(400, 9000.0, "BILL-OLD"),
        _open_bill(400, 90000.0, "BILL-OLD-BIG"),
    ], TODAY)
    assert [r["Bill"] for r in rows[:-1]] == \
        ["BILL-OLD-BIG", "BILL-OLD", "BILL-NEW"]


def test_the_balance_is_arithmetic_never_the_status_flag():
    """`ganit_vendor_bills` has no `balance_due` column to be tempted by, and
    `status` is a label somebody sets while a balance is a subtraction."""
    sql = sql_of(pa)
    assert "(COALESCE(b.total, 0) - COALESCE(b.amount_paid, 0))::float AS balance_due" in sql
    assert "b.total - COALESCE(b.amount_paid, 0) > 0" in sql


def test_the_payables_footer_ties():
    rows = pa.build_rows([_open_bill(10, 100.0), _open_bill(20, 250.25, "B2")],
                         TODAY)
    total = rows[-1]
    assert total[pa.LABEL_COLUMN] == pa.TOTAL_ROW
    assert total["Balance"] == 350.25
    assert total["Ageing"] == "" and total["Days overdue"] == ""


def test_the_stock_register_binds_no_window():
    """Declared `grain='stock'`, so `report_section` hands it None — and the
    query carries no date bind to apply one with."""
    sql = sql_of(pa)
    assert "BETWEEN" not in sql and "::date" not in sql


# ══════════════════════════════════════════════════════════════════════════
# expense register
# ══════════════════════════════════════════════════════════════════════════

def _expense(category="Rent", title="June office rent", vendor="Landlord LLP",
             amount=100000.0, tax=18000.0, total=118000.0):
    return {"spent_on": date(2026, 6, 1), "category": category, "title": title,
            "vendor": vendor, "amount": amount, "tax_amount": tax,
            "total": total}


def test_the_category_master_is_never_joined():
    """`ganit_expenses.category` is FREE TEXT with no FK, and the two disagree
    on live data: 12 of the 14 distinct strings match a category row, 2 do
    not. An INNER join would silently drop those rows and understate spend; a
    LEFT join would add a NULL column that invites the reader to think the row
    is broken."""
    assert "ganit_expense_categories" not in sql_of(er)
    assert "JOIN" not in sql_of(er), "the expense register joins nothing at all"


def test_the_register_is_ordered_by_category_then_date():
    """"By category" is an ORDERING, not a grouping key resolved against a
    master list — each category reads as one block a person can subtotal by
    eye, and the order is what makes the row ceiling a deterministic cut."""
    assert "ORDER BY category, e.expense_date, e.title" in sql_of(er)


def test_an_expense_with_no_category_prints_a_word():
    rows = er.build_rows([_expense(category=None)])
    assert rows[0][er.LABEL_COLUMN] == er.UNCATEGORISED


def test_the_expense_footer_ties_and_the_columns_add_across():
    """`amount + tax_amount` equals `total` on all 378 live rows."""
    rows = er.build_rows([_expense(), _expense(title="July rent")])
    total = rows[-1]
    assert total[er.LABEL_COLUMN] == er.TOTAL_ROW
    assert total["Amount"] + total["Tax"] == total["Total"]
    assert total["Total"] == 236000.0


def test_no_tds_column_is_invented():
    """`tds_amount` is 0 on all 378 live rows. A column of 378 zeroes reads as
    "no tax was deducted" rather than "nothing has ever been recorded here" —
    and the section and rate that would go with it come from
    `services/statute.py`, never from a literal here."""
    assert "tds_amount" not in sql_of(er)


# ══════════════════════════════════════════════════════════════════════════
# the row ceiling — shared by every register
# ══════════════════════════════════════════════════════════════════════════

def test_the_ceiling_admits_itself_in_the_table():
    """Not a silent truncation. A register that quietly stops at row N
    reconciles against nothing while still looking complete."""
    rows = sr.build_rows([_doc(number=f"INV-{i}") for i in range(3)], dropped=12)
    assert rows[-2][sr.LABEL_COLUMN] == sr.TOTAL_ROW
    notice = rows[-1][sr.LABEL_COLUMN]
    assert "Only the first 3 rows are listed" in notice
    assert "12+ more" in notice
    assert "listed rows only" in notice


def test_no_ceiling_no_notice():
    rows = sr.build_rows([_doc()])
    assert len(rows) == 2, "one document, one footer, nothing else"


def test_the_ceiling_is_a_runaway_guard_not_a_working_limit():
    """The largest register any org can produce today is 595 rows."""
    assert sh.ROW_CAP >= 5000


def test_every_register_fetches_one_more_row_than_it_prints():
    """`ROW_CAP + 1`, so the overflow is known without a second COUNT."""
    for module in (sr, pr, rr, pa, er):
        assert "LIMIT $" in sql_of(module), module.__name__


# ── the harness for binds ───────────────────────────────────────────────────

async def _capture(fn):
    """Run a register against a pool that records its binds and returns
    nothing — the org scoping and the parameter shapes, without a database."""
    seen: dict = {}

    class P:
        async def fetch(self, sql, *args):
            seen["sql"], seen["args"] = sql, list(args)
            return []

    await fn(P(), ORG, Window(date(2026, 5, 1), date(2026, 8, 17)))
    return seen


@pytest.mark.parametrize("fn", [
    sr.sales_register, pr.purchase_register, rr.receipts_register,
    er.expense_register_by_category,
])
def test_a_flow_register_binds_the_org_first_and_then_the_period(fn):
    seen = run(_capture(fn))
    assert seen["args"][0] == ORG
    assert seen["args"][1] == date(2026, 5, 1)
    assert seen["args"][2] == date(2026, 8, 17)
    assert seen["args"][-1] == sh.ROW_CAP + 1


def test_the_stock_register_binds_the_org_and_no_dates():
    seen: dict = {}

    class P:
        async def fetch(self, sql, *args):
            seen["args"] = list(args)
            return []

    run(pa.payables_ageing_by_bill(P(), ORG, None))
    assert seen["args"][0] == ORG
    assert not any(isinstance(a, date) for a in seen["args"])


# ══════════════════════════════════════════════════════════════════════════
# the shape the whole frame depends on
# ══════════════════════════════════════════════════════════════════════════

@pytest.mark.parametrize("key", NEW_KEYS)
def test_every_register_returns_the_widget_shape(key):
    """`render_report_html` discriminates on `label` + `data`/`absent` and
    nothing else. That is why a row-level report needed no new renderer."""
    class P:
        async def fetch(self, sql, *args):
            return []

    win = Window(date(2026, 5, 1), date(2026, 8, 17))
    out = run(mr.report_section(P(), ORG, "ganit", win, {"report": key}, None, {}))
    assert set(out) == {"report", "label", "grain", "data"}
    assert out["report"] == key
